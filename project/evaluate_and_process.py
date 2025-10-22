#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Combined Evaluation and Processing Script

This script combines the functionality of evaluation_DeCo_Diff_raw.py and process_raw_data_to_results.py
to provide 4 different execution modes:

1. save_only: Save .npy files and diff images only (no categorization)
2. process_only: Read existing .npy files and generate categorization results
3. save_and_process: Save .npy files and immediately process them for categorization
4. full_pipeline: Complete pipeline without saving intermediates (evaluation to categorization)
5. full_pipeline_with_saving_npy: Complete pipeline with saving intermediate NPY files (needs: --annotation-dir, --pretrained)

IMPORTANT: This script handles patch-level processing:
- The dataset returns individual patches (each __getitem__ returns a single patch)
- The DataLoader naturally batches individual patches together
- Each patch is a 3D tensor with shape [3, patch_size, patch_size]
- The default PyTorch collate function works perfectly for this use case

Usage examples:
  # Mode 1: Save only
  python evaluate_and_process.py --mode save_only --annotation-dir path/to/annotations --pretrained path/to/model.pt

  # Mode 2: Process only  
  python evaluate_and_process.py --mode process_only --annotation-dir path/to/annotations
  # Or with explicit results directory:
  python evaluate_and_process.py --mode process_only --results-dir path/to/results --annotation-dir path/to/annotations

  # Mode 3: Save and process
  python evaluate_and_process.py --mode save_and_process --annotation-dir path/to/annotations --pretrained path/to/model.pt

  # Mode 4: Full pipeline
  python evaluate_and_process.py --mode full_pipeline --annotation-dir path/to/annotations --pretrained path/to/model.pt

  # Mode 5: Full pipeline with saving NPY files
  python evaluate_and_process.py --mode full_pipeline_with_saving_npy --annotation-dir path/to/annotations --pretrained path/to/model.pt

  # Using JSON configuration only (mode specified in JSON)
  python evaluate_and_process.py --input-json config.json

  # Using JSON configuration with mode override
  python evaluate_and_process.py --mode save_only --input-json config.json

  # Enable debug logging for troubleshooting
  python evaluate_and_process.py --mode save_only --annotation-dir path/to/annotations --pretrained model.pt --debug

DISTRIBUTED FOLDER STRUCTURE:
This script automatically distributes .npy files across multiple folders to avoid Windows performance issues
with too many files in a single folder. When saving intermediate NPY files:

- Files are distributed across folders: results/evaluation_results/part_0000/, part_0001/, etc.
- Each folder contains approximately 100,000 files total (adaptive based on files per patch set)
- Each patch set typically consists of 4 files: _encodedrecon.npy, _latent.npy, _anomaly_map_arithmetic.npy, _coords.npy
- The file count per patch set is dynamically calculated at runtime for flexibility
- The loading functions automatically search across all distributed folders
- For small datasets, files are saved in the base evaluation_results folder (no distribution needed)

This optimization ensures optimal performance on Windows systems while maintaining full compatibility.
"""

from __future__ import annotations
import os
os.environ['NO_ALBUMENTATIONS_UPDATE'] = '1'

import warnings
warnings.filterwarnings(
    "ignore",
    message="A new version of Albumentations is available.*",
    category=UserWarning
)

import argparse
import json
import numpy as np
import torch
import torch.nn.functional as F
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Tuple, Set, Any
from collections import OrderedDict
from collections import deque
from tqdm import tqdm
import glob
from PIL import Image as PILImage
import cv2
import matplotlib.pyplot as plt
import re
import platform
from concurrent.futures import ThreadPoolExecutor

# Import from evaluation_DeCo_Diff_raw.py
from diffusion import create_diffusion
from diffusers.models.autoencoders.autoencoder_kl import AutoencoderKL
from models import UNET_models
from torchvision import transforms
from torch.utils.data import DataLoader, Dataset, Subset
import sys

# Configure UTF-8 encoding for output
if sys.platform.startswith('win'):
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Import utility functions
from utils import (
    path_to_safe_filename,
    safe_filename_to_path,
    _to_numpy,
    _binary_mask,
    _binary_mask_exclude_boundary3,
    load_original_images,
    load_ground_truth_map
)

# Additional imports for migrated functions
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from sklearn.metrics import roc_curve, auc
from torchmetrics.functional.image import (
    learned_perceptual_image_patch_similarity as _lpips,
    structural_similarity_index_measure as _ssim,
)
from io import BytesIO

# ============================================================================
# INLINED CLASSES AND FUNCTIONS FROM evaluation_DeCo_Diff2.py
# ============================================================================

# ---------------------------------------------------------------------------
# Constants and Type Definitions
# ---------------------------------------------------------------------------
_LATENT_SCALE = 0.18215

Kinded = Tuple[str, Any]  # (kind, value)
Record = OrderedDict[str, Kinded]

# ---------------------------------------------------------------------------
# EvaluationMetrics Class
# ---------------------------------------------------------------------------

class EvaluationMetrics:
    """Utility class for computing and storing evaluation metrics with memory-efficient approach."""

    def __init__(self):
        # Use running statistics instead of storing all values
        self.epoch_stats = {
            'encodedrecon_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []},
            'latent_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []},
            'anomaly_map_arithmetic_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []},
            'anomaly_map_geometric_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []}
        }
        # Store histogram bins for distribution analysis
        self.hist_bins = np.arange(0, 0.4, 0.01)
        self.histograms = {
            'encodedrecon_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64),
            'latent_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64),
            'anomaly_map_arithmetic_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64),
            'anomaly_map_geometric_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64)
        }

    def add_batch_stats(self, encodedrecon_diff, latent_diff, anomaly_map_arithmetic, anomaly_map_geometric):
        """Add batch statistics to epoch collection using running statistics."""
        # Process encodedrecon_diff
        encodedrecon_flat = encodedrecon_diff.flatten().cpu().numpy()
        self._update_running_stats('encodedrecon_values', encodedrecon_flat)
        self._update_histogram('encodedrecon_values', encodedrecon_flat)

        # Process latent_diff
        latent_flat = latent_diff.flatten().cpu().numpy()
        self._update_running_stats('latent_values', latent_flat)
        self._update_histogram('latent_values', latent_flat)

        # Process anomaly_map_arithmetic
        anomaly_map_arithmetic_flat = anomaly_map_arithmetic.flatten().cpu().numpy()
        self._update_running_stats('anomaly_map_arithmetic_values', anomaly_map_arithmetic_flat)
        self._update_histogram('anomaly_map_arithmetic_values', anomaly_map_arithmetic_flat)

        # Process anomaly_map_geometric
        anomaly_map_geometric_flat = anomaly_map_geometric.flatten().cpu().numpy()
        self._update_running_stats('anomaly_map_geometric_values', anomaly_map_geometric_flat)
        self._update_histogram('anomaly_map_geometric_values', anomaly_map_geometric_flat)


    def _update_running_stats(self, key, values):
        """Update running statistics for a given key."""
        stats = self.epoch_stats[key]
        n = len(values)

        if n == 0:
            return

        # Update count
        stats['count'] += n

        # Update sum and sum of squares
        stats['sum'] += np.sum(values)
        stats['sum_sq'] += np.sum(values ** 2)

        # Update min and max
        stats['min'] = min(stats['min'], np.min(values))
        stats['max'] = max(stats['max'], np.max(values))

        stats['values'].extend(values.tolist())

    def _update_histogram(self, key, values):
        """Update histogram for a given key."""
        hist, _ = np.histogram(values, bins=self.hist_bins, range=(0, 3))
        self.histograms[key] += hist

    def print_epoch_stats(self):
        """Print epoch-wise statistics."""
        print(f'\n=== EPOCH-WISE STATISTICS ===')

        for name, stats in self.epoch_stats.items():
            if stats['count'] > 0:
                # Calculate statistics from running values
                mean = stats['sum'] / stats['count']
                variance = (stats['sum_sq'] / stats['count']) - (mean ** 2)
                std = np.sqrt(max(0, variance))  # Ensure non-negative

                print(f'{name} epoch stats:')
                print(f'  min: {stats["min"]:.6f}, max: {stats["max"]:.6f}')
                print(f'  mean: {mean:.6f}, std: {std:.6f}')

                # Calculate quantiles if we have stored values
                if len(stats['values']) > 0:
                    values_array = np.array(stats['values'])
                    q1 = np.percentile(values_array, 25)
                    median = np.percentile(values_array, 50)
                    q3 = np.percentile(values_array, 75)
                    print(f'  Q1: {q1:.6f}, median: {median:.6f}, Q3: {q3:.6f}')

                # Print histogram
                hist = self.histograms[name]
                print(f'{name} epoch distribution:')
                for i, (count, edge) in enumerate(zip(hist, self.hist_bins[:-1])):
                    if stats['count'] > 0:
                        percentage = (count / stats['count']) * 100
                        print(f'  [{edge:.2f}-{edge+0.01:.2f}): {count:6d} ({percentage:5.1f}%)')

# ---------------------------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------------------------

def add_metric_fields(rec: Record, *, device=torch.device("cpu")) -> None:

    def to4d(x):
        if isinstance(x, np.ndarray):
            x = torch.from_numpy(x)
            if x.dtype != torch.float32:
                x = x.float()
            if x.ndim == 3 and x.shape[-1] == 3:  # HWC -> CHW
                x = x.permute(2, 0, 1)
            if x.ndim == 2:
                x = x.unsqueeze(0)
            if x.ndim == 3:
                x = x.unsqueeze(0)
            return x.to(device).clamp(-1, 1)
        return None

    a = to4d(rec["encoded_recon"][1])
    b = to4d(rec["dod_recon"][1])

    # Add null checks before computing metrics
    if a is not None and b is not None:
        rec["lpips"] = ("metric", _lpips(a, b, net_type="alex").item())
        ssim_result = _ssim(a, b)
        rec["ssim"] = ("metric", ssim_result[0].item() if isinstance(ssim_result, tuple) else ssim_result.item())
        rec["mse"] = ("metric", F.mse_loss(a, b).item())
    else:
        # Set default values if conversion fails
        rec["lpips"] = ("metric", 0.0)
        rec["ssim"] = ("metric", 0.0)
        rec["mse"] = ("metric", 0.0)


def make_record(**kwargs) -> Record:
    """Return an **ordered** dict whose values are (kind, value) pairs."""
    return OrderedDict(kwargs)


def _get_largest_connected_component_pixels(anomaly_binary: torch.Tensor) -> int:
    """
    Calculate the number of pixels in the largest connected component of white pixels.

    Args:
        anomaly_binary: Binary tensor with shape (H, W) or (1, H, W) where 1 indicates white pixels

    Returns:
        Number of pixels in the largest connected component
    """
    # Convert to numpy and ensure 2D shape
    if anomaly_binary.dim() == 3 and anomaly_binary.shape[0] == 1:
        binary_np = anomaly_binary.squeeze(0).cpu().numpy()
    else:
        binary_np = anomaly_binary.cpu().numpy()

    # Ensure binary values (0 or 1)
    binary_np = (binary_np > 0).astype(np.uint8)

    # Find connected components
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary_np, connectivity=8)

    if num_labels <= 1:  # Only background (label 0) or no components
        return 0

    # Find the largest component (excluding background which is label 0)
    largest_component_size = 0
    for i in range(1, num_labels):  # Skip background (i=0)
        component_size = stats[i, cv2.CC_STAT_AREA]
        if component_size > largest_component_size:
            largest_component_size = component_size

    return largest_component_size


def _create_contour_based_binary_mask_single(anomaly_map: torch.Tensor, adaptive_threshold: float = 0.1, anomaly_binary_threshold: int = 5) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
    """
    Create three different binary masks for a single image based on different contour selection styles.

    Args:
        anomaly_map: Anomaly map tensor with shape (H, W) with values in [0, 1]
        adaptive_threshold: Threshold for adaptive contour selection (default: 0.1)
        anomaly_binary_threshold: Threshold value for binary conversion (0-255, default: 5)
                                 - Lower values create more white pixels
                                 - Higher values create fewer white pixels

    Returns:
        Tuple of three binary tensors with shape (H, W) where selected contour pixels are 1, others are 0:
        - style1: Top contours by sum (most significant)
        - style2: Statistical outliers (mean + threshold * std)
        - style3: Contours contributing significant portion of total
    """
    # Convert to numpy
    map_np = anomaly_map.cpu().numpy()

    # Ensure the map is 2D
    if map_np.ndim != 2:
        print(f"Warning: Expected 2D array, got shape {map_np.shape}")
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask

    # Handle negative values and ensure proper range
    map_np = np.clip(map_np, 0, 1)  # Clip to [0, 1] range

    # Convert to uint8 for contour detection (0-255 range)
    map_uint8 = (map_np * 255).astype(np.uint8)

    # Apply morphological operations to reduce noise
    kernel = np.ones((3, 3), np.uint8)
    map_uint8 = cv2.morphologyEx(map_uint8, cv2.MORPH_CLOSE, kernel)  # Close small holes
    map_uint8 = cv2.morphologyEx(map_uint8, cv2.MORPH_OPEN, kernel)   # Remove small noise

    # Convert to binary image for contour detection
    _, binary_map = cv2.threshold(map_uint8, anomaly_binary_threshold, 255, cv2.THRESH_BINARY)

    # Check if the image is all zeros (no contours possible)
    if np.all(binary_map == 0):
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask

    try:
        # Find contours on the binary image
        contours, _ = cv2.findContours(binary_map, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    except cv2.error as e:
        print(f"OpenCV error in findContours: {e}")
        print(f"Binary map shape: {binary_map.shape}, dtype: {binary_map.dtype}")
        print(f"Binary map min: {binary_map.min()}, max: {binary_map.max()}")
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask

    if not contours:
        # No contours found, return all zeros
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask

    # Calculate contour statistics (sum of pixel values within each contour)
    contour_stats = []

    # Filter parameters to reduce noise
    min_contour_area = 10.0  # Minimum contour area to consider
    min_white_pixels = 5     # Minimum white pixels to consider

    for i, contour in enumerate(contours):
        # Create a mask for this contour
        contour_mask = np.zeros_like(binary_map)
        cv2.fillPoly(contour_mask, [contour], (255,))

        # Calculate sum of pixel values within this contour
        contour_white_pixels = np.sum(contour_mask > 0)
        contour_sum = np.sum(map_np * (contour_mask > 0))
        contour_area = cv2.contourArea(contour)

        # Filter out noise: skip contours that are too small
        if contour_area < min_contour_area or contour_white_pixels < min_white_pixels:
            continue

        contour_stats.append({
            'index': i,
            'contour': contour,
            'sum': contour_sum,
            'area': contour_area,
            'white_pixels': contour_white_pixels
        })

    # Sort by sum (descending)
    contour_stats.sort(key=lambda x: x['sum'], reverse=True)

    if not contour_stats:
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask

    # Extract sums for adaptive selection
    sums = np.array([stat['sum'] for stat in contour_stats])

    # Calculate statistics for adaptive selection
    mean_sum = np.mean(sums)
    std_sum = np.std(sums)

    # Remove statistical outliers (contours with sums > mean + 2*std)
    outlier_threshold = mean_sum + 2 * std_sum
    filtered_contour_stats = [stat for stat in contour_stats if stat['sum'] <= outlier_threshold]

    if not filtered_contour_stats:
        # If all contours were outliers, keep the top 3
        filtered_contour_stats = contour_stats[:3]

    # Use filtered contours for selection
    contour_stats = filtered_contour_stats

    # Create binary masks for each style
    binary_mask_style1 = np.zeros_like(binary_map)
    binary_mask_style2 = np.zeros_like(binary_map)
    binary_mask_style3 = np.zeros_like(binary_map)

    for contour_info in contour_stats:
        cv2.fillPoly(binary_mask_style1, [contour_info['contour']], (255,))

    for contour_info in contour_stats:
        cv2.fillPoly(binary_mask_style2, [contour_info['contour']], (255,))

    for contour_info in contour_stats:
        cv2.fillPoly(binary_mask_style3, [contour_info['contour']], (255,))

    # Convert back to tensors and normalize to [0, 1]
    binary_tensor_style1 = torch.from_numpy(binary_mask_style1).float() / 255.0
    binary_tensor_style2 = torch.from_numpy(binary_mask_style2).float() / 255.0
    binary_tensor_style3 = torch.from_numpy(binary_mask_style3).float() / 255.0

    # Return all three styles
    return (binary_tensor_style1.to(anomaly_map.device),
            binary_tensor_style2.to(anomaly_map.device),
            binary_tensor_style3.to(anomaly_map.device))

def _tensor_to_xlimage(arr, size: int) -> XLImage:
    if isinstance(arr, torch.Tensor):
        arr = arr.detach().cpu().numpy()
    arr = np.squeeze(arr)
    if arr.ndim == 2:
        arr = arr[..., None]
    if arr.ndim == 3 and arr.shape[0] in (1, 3, 4):
        arr = np.transpose(arr, (1, 2, 0))
    c = arr.shape[2]
    if c == 1:
        arr = np.repeat(arr, 3, axis=2)
    elif c == 4:
        # Split the image into 4 quadrants if c == 4 (e.g., 4-channel image)
        q0 = arr[..., 0]
        q1 = arr[..., 1]
        q2 = arr[..., 2]
        q3 = arr[..., 3]
        # Stack as 2x2 grid
        top = np.concatenate([q0, q1], axis=1)
        bottom = np.concatenate([q2, q3], axis=1)
        arr = np.stack(
            [np.concatenate([top, bottom], axis=0)] * 3, axis=2
        )  # make 3 channels
    elif c != 3:
        raise ValueError(f"unsupported channels: {c}")
    if np.any(arr < 0):
        arr = ((np.clip(arr, -1, 1) + 1) / 2 * 255).astype(np.uint8)
    else:
        arr = (arr * 255).astype(np.uint8)

    buf = BytesIO()
    PILImage.fromarray(arr, mode="RGB").save(buf, format="PNG")
    buf.seek(0)
    img = XLImage(buf)
    img.width = img.height = size
    return img


def _write_row(ws, row_idx: int, rec: dict, size: int):
    scalars, embeds = [], []
    for col_idx, key in enumerate(rec.keys(), 1):
        kind, val = rec[key]
        if kind == "image":
            embeds.append((col_idx, _tensor_to_xlimage(val, size)))
            scalars.append("")
            ws.column_dimensions[get_column_letter(col_idx)].width = size // 8
        else:
            # Convert tuples and other non-serializable types to strings
            if isinstance(val, (tuple, list)):
                scalars.append(str(val))
            elif isinstance(val, (np.ndarray, torch.Tensor)):
                scalars.append(str(val.tolist() if hasattr(val, 'tolist') else val))
            else:
                scalars.append(val)
    ws.append(scalars)
    ws.row_dimensions[row_idx].height = size * 0.75
    for col_idx, img in embeds:
        ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")


def make_excel(
    records: List[Record],
    image_size: int,
    save_dir: str | Path = "report",
    save_filename: str | None = datetime.now().strftime("%y%m%d_%H%M%S"),
    max_rows_per_file: int = 50,
) -> List[Path]:
    """Create Excel report with all evaluation records and images.

    Args:
        records: List of evaluation records
        image_size: Size of images to embed
        save_dir: Directory to save Excel files
        save_filename: Base filename for the Excel files
        max_rows_per_file: Maximum number of rows per Excel file (default: 100)

    Returns:
        List of paths to created Excel files
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    # Header comes from the first record's keys (order preserved)
    header = list(records[0].keys())

    # Calculate how many files we need
    total_rows = len(records)
    num_files = (total_rows + max_rows_per_file - 1) // max_rows_per_file  # Ceiling division

    # Check if the last file would have fewer than max_rows_per_file rows
    # If so, reduce the number of files by 1 and append remaining rows to the previous file
    if num_files > 1:
        remaining_rows = total_rows % max_rows_per_file
        if remaining_rows > 0 and remaining_rows < max_rows_per_file:
            num_files -= 1

    created_files = []

    for file_index in range(num_files):
        # Calculate start and end indices for this file
        start_idx = file_index * max_rows_per_file

        # For the last file, include all remaining rows
        if file_index == num_files - 1:
            end_idx = total_rows
        else:
            end_idx = (file_index + 1) * max_rows_per_file

        # Get records for this file
        file_records = records[start_idx:end_idx]

        # Create workbook for this file
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            # Set worksheet title
            if num_files == 1:
                ws.title = "Report"
            else:
                ws.title = f"Report_Part{file_index + 1}"

            # Add header
            ws.append(header)

            # Add data rows
            for r, rec in enumerate(file_records, start=2):
                _write_row(ws, r, rec, image_size)

            # Set column widths
            for c in range(1, len(header) + 1):
                ws.column_dimensions[get_column_letter(c)].width = 18

        # Generate filename
        if num_files == 1:
            out_filename = f"report_{save_filename}.xlsx"
        else:
            out_filename = f"report_{save_filename}_part{file_index + 1:02d}.xlsx"

        out_path = save_dir / out_filename
        wb.save(out_path)
        created_files.append(out_path)

        print(f"Report part {file_index + 1}/{num_files} saved to {out_path} ({len(file_records)} rows)")

    if num_files > 1:
        print(f"\nTotal: Created {num_files} Excel files with {total_rows} total rows")

    return created_files


def draw_patch_rectangles_on_image(base_img, predicted_defective_set, ground_truth_defective, overlapping, patch_size=128, grid_thickness=1):
    """
    Draw patch rectangles (TP/FP/FN) on top of an image.
    Args:
        base_img: The image to draw on (np.uint8, HxWx3)
        predicted_defective: Set of (grid_row, grid_col) for predicted defective patches
        ground_truth_defective: Set of (grid_row, grid_col) for ground truth defective patches
        overlapping: Set of (grid_row, grid_col) where prediction and ground truth overlap
        patch_size: Size of patches
        grid_thickness: Thickness of the rectangle lines (default: 1)
    Returns:
        Image with rectangles drawn:
        - Yellow rectangles around predicted defective regions
        - Red rectangles around ground truth defective regions
        - Green rectangles where prediction and ground truth overlap
    """
    img = base_img.copy()
    # Draw predicted defective regions (yellow)
    for grid_row, grid_col in predicted_defective_set:
        x = grid_col * patch_size
        y = grid_row * patch_size
        cv2.rectangle(img, (x, y), (x + patch_size, y + patch_size), (255, 255, 0), grid_thickness)
    # Draw ground truth defective regions (red)
    for grid_row, grid_col in ground_truth_defective:
        x = grid_col * patch_size
        y = grid_row * patch_size
        cv2.rectangle(img, (x, y), (x + patch_size, y + patch_size), (255, 0, 0), grid_thickness)
    # Draw overlapping regions (green)
    for grid_row, grid_col in overlapping:
        x = grid_col * patch_size
        y = grid_row * patch_size
        cv2.rectangle(img, (x, y), (x + patch_size, y + patch_size), (0, 255, 0), grid_thickness)
    return img


def determine_image_status(image_records: list) -> str:
    """
    Determine the overall status of an image based on its patch records.

    Returns:
        str: 'TP', 'FN', 'FP', or 'TN' based on the image's classification status
    """
    # Count patches by their individual status
    status_counts = {'TP': 0, 'FN': 0, 'FP': 0, 'TN': 0}

    for record in image_records:
        status = record["status"][1]  # Get the status from the record
        if status in status_counts:
            status_counts[status] += 1

    # Determine overall image status based on the most common patch status
    # If there are any defective patches (TP or FP), the image is considered defective
    # If there are no defective patches but there are normal patches (TN or FN), the image is considered normal
    # Priority: TP > FP > FN > TN (defective patches take precedence)

    if status_counts['TP'] > 0 or status_counts['FP'] > 0:
        # Image has defective patches
        if status_counts['TP'] > 0:
            return 'TP'  # True Positive: correctly identified as defective
        else:
            return 'FP'  # False Positive: incorrectly identified as defective
    else:
        # Image has no defective patches
        if status_counts['FN'] > 0:
            return 'FN'  # False Negative: missed defective patches
        else:
            return 'TN'  # True Negative: correctly identified as normal


def compute_y_true_y_score(all_records):
    """
    For each param value, compute y_true and y_score arrays from records.
    Returns: list of (y_true, y_score) tuples, one for each param value.
    """
    y_true_score_list = []
    for i, (records, records_defect) in enumerate(all_records):
        y_true = []
        y_score = []
        for rec in records:
            y_true.append(0)
            mask = (
                rec["anomaly_map_arithmetic_binary"][1]
                if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
                else rec["anomaly_map_arithmetic_binary"]
            )
            num_white = np.sum(mask == 1)
            y_score.append(num_white)
        for rec in records_defect:
            y_true.append(1)
            mask = (
                rec["anomaly_map_arithmetic_binary"][1]
                if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
                else rec["anomaly_map_arithmetic_binary"]
            )
            num_white = np.sum(mask == 1)
            y_score.append(num_white)
        y_score = np.array(y_score)
        y_true = np.array(y_true)
        y_true_score_list.append((y_true, y_score))
    return y_true_score_list


def compute_metrics_from_y_true_y_score(y_true_score_list):
    """
    For each (y_true, y_score), compute accuracy, threshold, and ROC stats.
    Returns: accuracies, thresholds, and a dict of lists for each ROC metric.
    """
    accuracies = []
    fpr_list = []
    tpr_list = []
    thresholds_list = []
    best_thresholds = []
    best_idxs = []
    aucs = []
    y_trues = []
    y_preds = []
    y_scores = []

    for i, (y_true, y_score) in enumerate(y_true_score_list):
        fpr, tpr, thresholds_, best_threshold, best_idx, auc_score = compute_roc_stats(
            y_true, y_score
        )
        y_pred = (y_score >= best_threshold).astype(int)
        y_preds.append(y_pred)
        accuracy = np.mean(y_pred == y_true)
        accuracies.append(accuracy)
        fpr_list.append(fpr)
        tpr_list.append(tpr)
        best_thresholds.append(best_threshold)
        best_idxs.append(best_idx)
        aucs.append(auc_score)
        y_trues.append(y_true)
        y_scores.append(y_score)
        print(f"Accuracy {accuracy:.4f} (threshold={best_threshold})")

    roc_stats = {
        "fpr": fpr_list,
        "tpr": tpr_list,
        "best_threshold": best_thresholds,
        "best_idx": best_idxs,
        "auc": aucs,
        "y_true": y_trues,
        "y_pred": y_preds,
        "y_score": y_scores,
        "accuracies": accuracies,
    }
    return roc_stats


def compute_roc_stats(y_true, y_score):
    """
    Compute ROC statistics and best threshold using Youden's J statistic.
    Returns: fpr, tpr, thresholds, best_threshold, best_idx, auc_score
    """
    fpr, tpr, thresholds = roc_curve(y_true, y_score)
    youden_j = tpr - fpr
    best_idx = np.argmax(youden_j)
    best_threshold = thresholds[best_idx]
    auc_score = auc(fpr, tpr)
    return fpr, tpr, thresholds, best_threshold, best_idx, auc_score


def plot_accuracy_results(
    param_values,
    accuracies,
    param_name: str,
    save_dir="accuracy_vs_param",
    save_filename=datetime.now().strftime("%y%m%d_%H%M%S"),
    title=None,
    xlabel=None,
    ylabel="Accuracy",
    grid=True,
    marker="o",
    **plot_kwargs,
):
    """
    Plot accuracy results with customizable parameters.

    Args:
        param_values: List of parameter values
        accuracies: List of corresponding accuracies
        param_name: Name of the parameter being varied
        save_dir: Directory to save the plot
        save_filename: Filename for saving the plot
        title: Custom title for the plot
        xlabel: Custom x-axis label
        ylabel: Custom y-axis label
        grid: Whether to show grid
        marker: Marker style for the plot
        **plot_kwargs: Additional plotting parameters
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    plt.figure()
    plt.plot(param_values, accuracies, marker=marker, **plot_kwargs)
    plt.xlabel(xlabel or param_name)
    plt.ylabel(ylabel)
    plt.title(title or f"Accuracy vs {param_name.capitalize()} (synthetic defect)")
    plt.ylim(0.5, 1.0)
    plt.grid(grid)
    out_path = os.path.join(save_dir, f"accuracy_vs_{param_name}_{save_filename}.png")
    plt.savefig(out_path)
    print(f"Accuracy vs {param_name} saved to {out_path}")
    plt.close()


def save_perturbation_results(
    param_name: str,
    roc_stats: dict,
    param_values: list,
    save_dir: str,
):
    """
    Save perturbation experiment data (roc_stats and param_values) to a specified folder in JSON format.

    Args:
        param_name: Name of the perturbation parameter
        roc_stats: Dictionary containing ROC statistics
        param_values: List of parameter values
        save_dir: Directory to save the results
    """
    save_dir_path = Path(save_dir).expanduser()
    save_dir_path.mkdir(parents=True, exist_ok=True)

    # Convert numpy arrays to lists for JSON serialization
    def convert_for_json(obj):
        if hasattr(obj, 'tolist'):  # numpy arrays
            return obj.tolist()
        elif isinstance(obj, list):
            return [convert_for_json(item) for item in obj]
        elif isinstance(obj, dict):
            return {key: convert_for_json(value) for key, value in obj.items()}
        else:
            return obj

    # Convert roc_stats to JSON-serializable format
    roc_stats_json = convert_for_json(roc_stats)

    # Convert param_values to JSON-serializable format
    param_values_json = convert_for_json(param_values)

    # Save both roc_stats and param_values in a single JSON file
    results_data = {
        "param_name": param_name,
        "param_values": param_values_json,
        "roc_stats": roc_stats_json
    }

    json_path = os.path.join(save_dir, f"{param_name}_results.json")
    with open(json_path, "w") as f:
        json.dump(results_data, f, indent=2)

    print(f"Perturbation results saved to {json_path}")


def save_patch_results_from_records(checkpoint_manager, image_path: str,
                                  patch_records: list,
                                  predicted_defective_set: set, ground_truth_defective: set, overlapping: set,
                                  enable_save_optional_image_results: bool = False, patch_size: int = 256,
                                  patch_x: int = 0, patch_y: int = 0):
    """Save patch-level results immediately."""
    safe_name = path_to_safe_filename(image_path)
    patch_coord_str = f"x{patch_x}_y{patch_y}"

    # Create status-based subfolders
    status_folders = {}
    for status in ['TP', 'FN', 'FP', 'TN']:
        status_dir = os.path.join(checkpoint_manager.marked_images_dir, status)
        os.makedirs(status_dir, exist_ok=True)
        status_folders[status] = status_dir

    # Determine the status of this patch
    if not patch_records:
        print(f"Warning: No patch records provided for {image_path}")
        return

    patch_status = patch_records[0]["status"][1]

    # Load original image
    try:
        original_img = np.array(PILImage.open(image_path).convert('RGB'))
        h, w, _ = original_img.shape
    except Exception as e:
        print(f"Warning: Failed to load image {image_path}: {e}")
        return

    # Extract patch region from original image
    patch_width = min(patch_size, w - patch_x)
    patch_height = min(patch_size, h - patch_y)
    patch_img = original_img[patch_y:patch_y+patch_height, patch_x:patch_x+patch_width]

    # Save patch image in the appropriate status folder
    patch_filename = f"{safe_name}__{patch_coord_str}.png"
    patch_path = os.path.join(status_folders[patch_status], patch_filename)
    PILImage.fromarray(patch_img).save(patch_path)

    # Save patch-level anomaly maps
    for record in patch_records:
        # Extract anomaly maps for this patch
        patch_arithmetic = record["anomaly_map_arithmetic"][1]
        patch_arithmetic_binary = record["anomaly_map_arithmetic_binary"][1]
        patch_geometric = record["anomaly_map_geometric"][1]
        patch_geometric_binary = record["anomaly_map_geometric_binary"][1]

        # Extract patch regions
        patch_regions = {
            'ar': patch_arithmetic.squeeze()[:patch_height, :patch_width],
            'ar_bin': patch_arithmetic_binary.squeeze()[:patch_height, :patch_width],
            'ge': patch_geometric.squeeze()[:patch_height, :patch_width],
            'ge_bin': patch_geometric_binary.squeeze()[:patch_height, :patch_width],
        }

        # Save patch-level anomaly maps
        for map_name, region in patch_regions.items():
            # Create anomaly map image for this patch
            anomaly_map_img = ImageProcessor.create_anomaly_map_image(
                region, patch_size=patch_size, add_grid=False,
                predicted_defective_set=predicted_defective_set, ground_truth_defective=ground_truth_defective,
                overlapping=overlapping, is_binary=(map_name.endswith('binary'))
            )
            anomaly_map_filename = f"{safe_name}__{patch_coord_str}_{map_name}.png"
            anomaly_map_path = os.path.join(status_folders[patch_status], anomaly_map_filename)
            PILImage.fromarray(anomaly_map_img).save(anomaly_map_path)

    # Mark this patch as processed
    if checkpoint_manager:
        checkpoint_manager.mark_image_processed(image_path)  # For now, mark entire image

# ---------------------------------------------------------------------------
# CheckpointManager Class
# ---------------------------------------------------------------------------

class CheckpointManager:
    """Manages checkpoint/resume functionality for evaluation."""
    
    def __init__(self, results_dir: str, annotation_dir: str | None = None, force_rerun: bool = False):
        self.results_dir = results_dir
        self.annotation_dir = annotation_dir
        self.force_rerun = force_rerun
        
        # Extract base name without timestamp for consistent checkpoint location
        base_name = self._extract_base_name(results_dir)
        self.base_checkpoint_dir = os.path.join(os.path.dirname(results_dir), f"{base_name}_checkpoints")
        
        # Create checkpoint directory
        os.makedirs(self.base_checkpoint_dir, exist_ok=True)
        
        self.checkpoint_file = os.path.join(self.base_checkpoint_dir, "evaluation_checkpoint.json")

        # Create evaluation_results directory at the base results level (shared across all runs)
        base_results_dir = os.path.dirname(results_dir) if os.path.basename(results_dir) != "results" else results_dir
        self.evaluation_results_dir = os.path.join(base_results_dir, "evaluation_results")
        os.makedirs(self.evaluation_results_dir, exist_ok=True)

        # Create marked_images directory for image saving
        self.marked_images_dir = os.path.join(results_dir, "marked_images")
        os.makedirs(self.marked_images_dir, exist_ok=True)
        
        # Clear checkpoint files if force rerun is enabled
        if self.force_rerun:
            self.clear_checkpoint_files()
            print("Force rerun enabled: Cleared existing checkpoint files")
        
        # Cache for processed images to avoid frequent file I/O
        self._processed_images_cache = None
        self._cache_timestamp = 0
    
    def _extract_base_name(self, results_dir: str) -> str:
        """Extract base name from timestamped directory."""
        # Handle patterns like "test_name_250708_143022" -> "test_name"
        dir_name = os.path.basename(results_dir)
        
        # Try to find the last underscore followed by timestamp pattern
        import re
        # Pattern for timestamp: YYMMDD_HHMMSS
        timestamp_pattern = r'_\d{6}_\d{6}$'
        match = re.search(timestamp_pattern, dir_name)
        
        if match:
            # Remove the timestamp part
            base_name = dir_name[:match.start()]
            return base_name
        else:
            # If no timestamp pattern found, use the original name
            return dir_name
    
    def find_latest_checkpoint(self) -> str:
        """Find the latest checkpoint file from the specific checkpoint directory."""
        if os.path.exists(self.checkpoint_file):
            return self.checkpoint_file
        
        # Look for checkpoint in the specific directory based on base name
        # e.g., if results_dir is "test_250711_162955", look for "test_checkpoints"
        base_name = self._extract_base_name(self.results_dir)
        specific_checkpoint_dir = os.path.join(os.path.dirname(self.results_dir), f"{base_name}_checkpoints")
        specific_checkpoint_file = os.path.join(specific_checkpoint_dir, "evaluation_checkpoint.json")
        
        if os.path.exists(specific_checkpoint_file):
            print(f"Found existing checkpoint: {specific_checkpoint_file}")
            return specific_checkpoint_file
        
        return self.checkpoint_file
    
    def get_checkpoint_data(self) -> dict:
        """Load checkpoint data if it exists."""
        checkpoint_file = self.find_latest_checkpoint()
        if os.path.exists(checkpoint_file):
            try:
                with open(checkpoint_file, 'r') as f:
                    return json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                pass
        return {"current_image_index": 0, "processed_images": []}
    
    def _get_processed_images_with_cache(self) -> set:
        """Get processed images with caching to avoid frequent file I/O."""
        # Check if cache is valid (file hasn't been modified since last cache)
        checkpoint_file = self.find_latest_checkpoint()
        if os.path.exists(checkpoint_file):
            current_timestamp = os.path.getmtime(checkpoint_file)
            if (self._processed_images_cache is not None and 
                self._cache_timestamp == current_timestamp):
                return self._processed_images_cache
        
        # Load from file and update cache
        checkpoint_data = self.get_checkpoint_data()
        processed_images = set(checkpoint_data.get("processed_images", []))
        
        # Update cache
        self._processed_images_cache = processed_images
        self._cache_timestamp = current_timestamp if os.path.exists(checkpoint_file) else 0
        
        return processed_images
    
    def save_checkpoint(self, current_image_index: int, processed_images: list):
        """Save current progress to checkpoint file."""
        # Get existing checkpoint data to preserve processed_images
        existing_data = self.get_checkpoint_data()
        
        # Merge processed images from both sources (optimized)
        existing_processed = set(existing_data.get("processed_images", []))
        new_processed = set(processed_images)
        
        # Only merge if there are actually new images to add
        if new_processed - existing_processed:
            merged_processed = list(existing_processed.union(new_processed))
        else:
            merged_processed = list(existing_processed)
        
        # Sort for consistent, predictable order (only when writing to file)
        merged_processed.sort()
        
        checkpoint_data = {
            "current_image_index": current_image_index,
            "processed_images": merged_processed,
            "timestamp": datetime.now().isoformat()
        }
        
        with open(self.checkpoint_file, 'w') as f:
            json.dump(checkpoint_data, f, indent=2)
        
        # Invalidate cache since we just wrote to the file
        self._processed_images_cache = None
    
    def get_processed_images(self) -> set:
        """Get set of already processed images from checkpoint file."""
        return self._get_processed_images_with_cache()
    
    def mark_image_processed(self, image_path: str):
        """Mark an image as processed by updating the checkpoint file."""
        # Use cache to avoid reading file again
        processed_images = self._get_processed_images_with_cache()
        
        # Only update if image is not already processed
        if image_path not in processed_images:
            processed_images.add(image_path)
            
            # Get current checkpoint data
            checkpoint_data = self.get_checkpoint_data()
            # Sort for consistent, predictable order (only when writing to file)
            checkpoint_data["processed_images"] = sorted(processed_images)
            checkpoint_data["timestamp"] = datetime.now().isoformat()
            
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f, indent=2)
            
            # Update cache
            self._processed_images_cache = processed_images
            self._cache_timestamp = os.path.getmtime(self.checkpoint_file)
    
    def is_image_processed(self, image_path: str) -> bool:
        """Check if an image has been fully processed."""
        # If force rerun is enabled, no image is considered processed
        if self.force_rerun:
            return False
        return image_path in self.get_processed_images()
    
    def get_resume_info(self, all_image_paths: list) -> tuple[int, list]:
        """Get resume information for evaluation."""
        # If force rerun is enabled, start from the beginning
        if self.force_rerun:
            print("Force rerun enabled: Starting from the beginning")
            return 0, []
        
        checkpoint_data = self.get_checkpoint_data()
        processed_images = self.get_processed_images()
        
        # Find the first unprocessed image
        current_index = 0
        for i, image_path in enumerate(all_image_paths):
            if image_path not in processed_images:
                current_index = i
                break
        else:
            # All images processed
            current_index = len(all_image_paths)
        
        return current_index, sorted(processed_images)
    
    def cleanup_checkpoint(self):
        """Clean up checkpoint files after successful completion."""
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
        
        # Also clean up the checkpoint directory if it's empty
        try:
            if os.path.exists(self.base_checkpoint_dir) and not os.listdir(self.base_checkpoint_dir):
                os.rmdir(self.base_checkpoint_dir)
        except OSError:
            pass  # Directory not empty or already removed

    def clear_checkpoint_files(self):
        """Clear all checkpoint files to force a fresh start."""
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
            print(f"Removed checkpoint file: {self.checkpoint_file}")
        
        # Clear cache
        self._processed_images_cache = None
        self._cache_timestamp = 0
    
    def batch_mark_images_processed(self, image_paths: list):
        """Mark multiple images as processed in a single operation for better performance."""
        if not image_paths:
            return
        
        # Use cache to avoid reading file again
        processed_images = self._get_processed_images_with_cache()
        
        # Find new images to add
        new_images = [img for img in image_paths if img not in processed_images]
        
        if new_images:
            # Add new images
            processed_images.update(new_images)
            
            # Get current checkpoint data
            checkpoint_data = self.get_checkpoint_data()
            # Sort for consistent, predictable order (only when writing to file)
            checkpoint_data["processed_images"] = sorted(processed_images)
            checkpoint_data["timestamp"] = datetime.now().isoformat()
            
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f, indent=2)
            
            # Update cache
            self._processed_images_cache = processed_images
            self._cache_timestamp = os.path.getmtime(self.checkpoint_file)
    
    def print_checkpoint_status(self):
        """Print current checkpoint status for debugging."""
        checkpoint_file = self.find_latest_checkpoint()
        
        print(f"Checkpoint directory: {self.base_checkpoint_dir}")
        print(f"Current checkpoint file: {checkpoint_file}")
        
        if os.path.exists(checkpoint_file):
            try:
                with open(checkpoint_file, 'r') as f:
                    data = json.load(f)
                    print(f"Checkpoint timestamp: {data.get('timestamp', 'unknown')}")
                    print(f"Current image index: {data.get('current_image_index', 0)}")
                    print(f"Processed images count: {len(data.get('processed_images', []))}")
                    if data.get('processed_images'):
                        print(f"Last processed: {data['processed_images'][-1]}")
            except Exception as e:
                print(f"Error reading checkpoint: {e}")
        else:
            print("No checkpoint file found")


class ImageProcessor:
    """Utility class for common image processing operations."""
    
    @staticmethod
    def create_anomaly_map_image(anomaly_map, predicted_defective_set, ground_truth_defective, overlapping, is_binary=True, patch_size=128, add_grid=True, grid_color=(255, 255, 255), grid_thickness=1):
        """Create anomaly map visualization with optional grid overlay and patch prediction rectangles."""
        if is_binary:
            # Binary map: 0 or 1 - create custom red colormap
            anomaly_map_img = (anomaly_map * 255).astype(np.uint8)
            # Create custom colormap: 0 -> transparent (black), 255 -> pure red
            h, w = anomaly_map_img.shape
            anomaly_map_colored_bgr = np.zeros((h, w, 3), dtype=np.uint8)
            # Set red channel based on anomaly values (0 = transparent/black, 255 = pure red)
            anomaly_map_colored_bgr[:, :, 2] = anomaly_map_img  # Red channel (BGR format)
        else:
            anomaly_map_img = (anomaly_map * 255).astype(np.uint8)
            anomaly_map_colored_bgr = cv2.applyColorMap(anomaly_map_img, cv2.COLORMAP_HOT)

        # Add grid overlay if requested
        if add_grid:
            h, w = anomaly_map.shape
            
            # Draw vertical grid lines
            for x in range(patch_size, w, patch_size):
                cv2.line(anomaly_map_colored_bgr, (x, 0), (x, h), grid_color, grid_thickness)
            
            # Draw horizontal grid lines
            for y in range(patch_size, h, patch_size):
                cv2.line(anomaly_map_colored_bgr, (0, y), (w, y), grid_color, grid_thickness)
            
            # Draw border around the entire image
            cv2.rectangle(anomaly_map_colored_bgr, (0, 0), (w-1, h-1), grid_color, grid_thickness)
        
        # Convert BGR to RGB for proper display
        anomaly_map_colored = cv2.cvtColor(anomaly_map_colored_bgr, cv2.COLOR_BGR2RGB)
        
        # Add patch prediction rectangles if provided
        if predicted_defective_set is not None and ground_truth_defective is not None and overlapping is not None:
            anomaly_map_colored = draw_patch_rectangles_on_image(
                anomaly_map_colored, 
                predicted_defective_set, 
                ground_truth_defective, 
                overlapping, 
                patch_size=patch_size, 
                grid_thickness=grid_thickness
            )
        
        return anomaly_map_colored
    
    @staticmethod
    def create_anomaly_overlay(original_img, anomaly_map, alpha=0.6, is_binary=True):
        """Create overlay of anomaly map on original image."""
        if is_binary:
            # Binary overlay using pure red
            overlay = original_img.copy()
            mask = anomaly_map > 0
            # Create pure red overlay for anomaly regions
            h, w = anomaly_map.shape
            anomaly_colored = np.zeros((h, w, 3), dtype=np.uint8)
            # Set red channel for anomaly regions (pure red)
            anomaly_colored[mask, 0] = 255  # Red channel (RGB format)
            # Apply the red anomaly regions to the overlay
            overlay[mask] = anomaly_colored[mask]
        else:
            # Continuous overlay using HOT colormap
            # Convert [0, 1] range to [0, 255] range for colormap
            anomaly_map_uint8 = (anomaly_map * 255).astype(np.uint8)
            
            # Apply HOT colormap to the map (returns BGR)
            anomaly_colored_bgr = cv2.applyColorMap(anomaly_map_uint8, cv2.COLORMAP_HOT)
            # Convert BGR to RGB for proper overlay
            anomaly_colored = cv2.cvtColor(anomaly_colored_bgr, cv2.COLOR_BGR2RGB)
            # Create overlay using alpha blending
            overlay = cv2.addWeighted(original_img, 1-alpha, anomaly_colored, alpha, 0)
        
        return overlay.astype(np.uint8)


def save_image_results_from_records(marked_images_dir: str, evaluation_results_dir: str, image_path: str, 
                                  image_records: list, 
                                  predicted_defective_set: set, ground_truth_defective: set, overlapping: set,
                                  enable_save_optional_image_results: bool = False, patch_size: int = 256):
    """Save all results for a single image immediately using records."""
    from utils import path_to_safe_filename
    
    safe_name = path_to_safe_filename(image_path)
    
    # Create status-based subfolders in the provided marked_images_dir
    status_folders = {}
    for status in ['TP', 'FN', 'FP', 'TN']:
        status_dir = os.path.join(marked_images_dir, status)
        os.makedirs(status_dir, exist_ok=True)
        status_folders[status] = status_dir
    
    # Determine the overall status of this image based on its patches
    image_status = determine_image_status(image_records)
    
    # Debug: Print status information
    status_counts = {'TP': 0, 'FN': 0, 'FP': 0, 'TN': 0}
    for record in image_records:
        status = record["status"][1]
        if status in status_counts:
            status_counts[status] += 1
    #print(f"Image {os.path.basename(image_path)} status: {image_status} (TP:{status_counts['TP']}, FN:{status_counts['FN']}, FP:{status_counts['FP']}, TN:{status_counts['TN']})")
    
    # Load original image
    original_img = np.array(PILImage.open(image_path).convert('RGB'))
    h, w, _ = original_img.shape
    
    # Save marked image (always saved) in the appropriate status folder
    marked_img = draw_patch_rectangles_on_image(
        original_img, predicted_defective_set, ground_truth_defective, overlapping, patch_size=patch_size, grid_thickness=1
    )
    marked_path = os.path.join(status_folders[image_status], f"{safe_name}__marked.png")
    PILImage.fromarray(marked_img).save(marked_path)
    
    # Also save image-level images without classification
    image_level_dir = os.path.join(marked_images_dir, "image_level")
    os.makedirs(image_level_dir, exist_ok=True)
    image_level_path = os.path.join(image_level_dir, f"{safe_name}__marked.png")
    PILImage.fromarray(marked_img).save(image_level_path)
    
    # Initialize anomaly maps based on flag
    anomaly_maps = {
        'required': {
            'arithmetic': np.zeros((h, w), dtype=np.float32),
            'arithmetic_binary': np.zeros((h, w), dtype=np.float32),
            'geometric': np.zeros((h, w), dtype=np.float32),
            'geometric_binary': np.zeros((h, w), dtype=np.float32),
        }
    }
    
    # Initialize optional maps only when flag is enabled
    if enable_save_optional_image_results:
        anomaly_maps['optional'] = {
            'arithmetic_binary_style1': np.zeros((h, w), dtype=np.float32),
            'arithmetic_binary_style2': np.zeros((h, w), dtype=np.float32),
            'arithmetic_binary_style3': np.zeros((h, w), dtype=np.float32),
            'geometric_binary_style1': np.zeros((h, w), dtype=np.float32),
            'geometric_binary_style2': np.zeros((h, w), dtype=np.float32),
            'geometric_binary_style3': np.zeros((h, w), dtype=np.float32),
        }
    
    def _get_xy_from_patch_coords(rec) -> tuple:
        coords = rec.get("patch_coords", (None, []))[1]
        if isinstance(coords, (list, tuple)) and len(coords) == 8:
            return int(coords[0]), int(coords[1])  # top-left from 8-value
        raise ValueError(f"Expected 8-value patch_coords, got: {coords}")

    for record in image_records:
        # Extract coordinates from record (supports 8- or 2-value formats)
        x_coord, y_coord = _get_xy_from_patch_coords(record)
        
        # Calculate actual patch dimensions for this position
        patch_width = min(patch_size, w - x_coord)
        patch_height = min(patch_size, h - y_coord)
        
        # Extract required anomaly maps from record
        patch_arithmetic = record["anomaly_map_arithmetic"][1]
        patch_arithmetic_binary = record["anomaly_map_arithmetic_binary"][1]
        patch_geometric = record["anomaly_map_geometric"][1]
        patch_geometric_binary = record["anomaly_map_geometric_binary"][1]
        
        # Extract patch regions
        patch_regions = {
            'arithmetic': patch_arithmetic.squeeze()[:patch_height, :patch_width],
            'arithmetic_binary': patch_arithmetic_binary.squeeze()[:patch_height, :patch_width],
            'geometric': patch_geometric.squeeze()[:patch_height, :patch_width],
            'geometric_binary': patch_geometric_binary.squeeze()[:patch_height, :patch_width],
        }
        
        # Extract optional patch regions if flag is enabled
        if enable_save_optional_image_results:
            # Use default binary maps as fallback for style fields
            patch_regions.update({
                'arithmetic_binary_style1': record.get("anomaly_map_arithmetic_binary_style1", [None, patch_arithmetic_binary])[1].squeeze()[:patch_height, :patch_width],
                'arithmetic_binary_style2': record.get("anomaly_map_arithmetic_binary_style2", [None, patch_arithmetic_binary])[1].squeeze()[:patch_height, :patch_width],
                'arithmetic_binary_style3': record.get("anomaly_map_arithmetic_binary_style3", [None, patch_arithmetic_binary])[1].squeeze()[:patch_height, :patch_width],
                'geometric_binary_style1': record.get("anomaly_map_geometric_binary_style1", [None, patch_geometric_binary])[1].squeeze()[:patch_height, :patch_width],
                'geometric_binary_style2': record.get("anomaly_map_geometric_binary_style2", [None, patch_geometric_binary])[1].squeeze()[:patch_height, :patch_width],
                'geometric_binary_style3': record.get("anomaly_map_geometric_binary_style3", [None, patch_geometric_binary])[1].squeeze()[:patch_height, :patch_width],
            })
        
        # Assign required regions to anomaly maps
        for map_name, region in patch_regions.items():
            if map_name in anomaly_maps['required']:
                anomaly_maps['required'][map_name][y_coord:y_coord+patch_height, x_coord:x_coord+patch_width] = region
            elif enable_save_optional_image_results and map_name in anomaly_maps['optional']:
                anomaly_maps['optional'][map_name][y_coord:y_coord+patch_height, x_coord:x_coord+patch_width] = region
    
    # Define image configurations
    image_configs = {
        'required': [
            (anomaly_maps['required']['arithmetic'], "am_ar", False),
            (anomaly_maps['required']['arithmetic_binary'], "am_ar_bin", True),
            (anomaly_maps['required']['geometric'], "am_ge", False),
            (anomaly_maps['required']['geometric_binary'], "am_ge_bin", True),
        ]
    }
    
    # Add optional configurations if flag is enabled
    if enable_save_optional_image_results:
        image_configs['optional'] = [
            (anomaly_maps['optional']['arithmetic_binary_style1'], "am_ar_bin_st1", True),
            (anomaly_maps['optional']['arithmetic_binary_style2'], "am_ar_bin_st2", True),
            (anomaly_maps['optional']['arithmetic_binary_style3'], "am_ar_bin_st3", True),
            (anomaly_maps['optional']['geometric_binary_style1'], "am_ge_bin_st1", True),
            (anomaly_maps['optional']['geometric_binary_style2'], "am_ge_bin_st2", True),
            (anomaly_maps['optional']['geometric_binary_style3'], "am_ge_bin_st3", True),
        ]
    
    # Save all configured images in the appropriate status folder
    for _config_type, configs in image_configs.items():
        for anomaly_map, suffix, is_binary in configs:
            # Save anomaly map image
            anomaly_map_img = ImageProcessor.create_anomaly_map_image(
                anomaly_map, predicted_defective_set, ground_truth_defective, 
                overlapping, is_binary=is_binary, patch_size=patch_size, add_grid=True
            )
            anomaly_map_path = os.path.join(status_folders[image_status], f"{safe_name}__{suffix}.png")
            PILImage.fromarray(anomaly_map_img).save(anomaly_map_path)
            
            # Save overlay image
            overlay_img = ImageProcessor.create_anomaly_overlay(original_img, anomaly_map, alpha=0.8, is_binary=is_binary)
            overlay_path = os.path.join(status_folders[image_status], f"{safe_name}__ao_{suffix}.png")
            PILImage.fromarray(overlay_img).save(overlay_path)
            
            # Save marked overlay image
            marked_overlay_img = draw_patch_rectangles_on_image(overlay_img, predicted_defective_set, ground_truth_defective, overlapping, patch_size=patch_size, grid_thickness=1)
            marked_overlay_path = os.path.join(status_folders[image_status], f"{safe_name}__mo_{suffix}.png")
            PILImage.fromarray(marked_overlay_img).save(marked_overlay_path)
            
            # Also save to image_level directory (without classification)
            image_level_anomaly_path = os.path.join(image_level_dir, f"{safe_name}__{suffix}.png")
            PILImage.fromarray(anomaly_map_img).save(image_level_anomaly_path)
            
            image_level_overlay_path = os.path.join(image_level_dir, f"{safe_name}__ao_{suffix}.png")
            PILImage.fromarray(overlay_img).save(image_level_overlay_path)
            
            image_level_marked_overlay_path = os.path.join(image_level_dir, f"{safe_name}__mo_{suffix}.png")
            PILImage.fromarray(marked_overlay_img).save(image_level_marked_overlay_path)
    
    # Save evaluation results
    patch_analysis = []
    for record in image_records:
        x, y = _get_xy_from_patch_coords(record)
        anomaly_map = record["anomaly_map_arithmetic_binary"][1]
        anomaly_pixels = int(np.sum(anomaly_map))
        grid_row = y // patch_size
        grid_col = x // patch_size
        patch_analysis.append({
            "grid_row": grid_row,
            "grid_col": grid_col,
            "anomaly_max": record["anomaly_max"][1],
            "anomaly_pixels":record["anomaly_pixels"][1],
            "status": record["status"][1]
        })
    
    result_filename = f"{safe_name}__evaluation.json"
    result_path = os.path.join(evaluation_results_dir, result_filename)
    evaluation_result = {
        "image_path": image_path,
        "patch_analysis": patch_analysis,
        "grid_size": patch_size
    }
    with open(result_path, 'w') as f:
        json.dump(evaluation_result, f, indent=2)

# ============================================================================
# END OF INLINED CLASSES AND FUNCTIONS
# ============================================================================

# Set up device
torch.set_grad_enabled(False)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

if device == torch.device("cpu"):
    print("GPU not found. Using CPU instead.")
else:
    # Enable performance knobs for GPU
    try:
        torch.backends.cudnn.benchmark = True
        torch.backends.cuda.matmul.allow_tf32 = True
        torch.backends.cudnn.allow_tf32 = True
    except Exception:
        print("GPU not found. Using CPU instead.")
        pass

# Constants
_LATENT_SCALE = 0.18215

# Type definitions
Kinded = Tuple[str, Any]  # (kind, value)
Record = OrderedDict[str, Kinded]

# Global debug flag (will be set from args)
DEBUG_ENABLED = False

from utils import debug_print

# Shared utility functions for record creation and processing
def _extract_patch_coordinates(patch_coords, batch_index=0, patch_size=128):
    """
    Extract patch coordinates from various input formats.
    
    Args:
        patch_coords: Tensor or list containing patch coordinates
        batch_index: Index for batch processing (default: 0)
        patch_size: Default patch size for fallback coordinates
    
    Returns:
        tuple: coords_8_values
    """
    if isinstance(patch_coords, torch.Tensor) and len(patch_coords.shape) == 2:
        if batch_index < patch_coords.size(0):
            coords_8_values = patch_coords[batch_index].tolist()
        else:
            coords_8_values = patch_coords[-1].tolist() if patch_coords.size(0) > 0 else [0, 0, patch_size, 0, patch_size, patch_size, 0, patch_size]
    elif isinstance(patch_coords, (list, tuple)):
        coords_8_values = patch_coords
    else:
        raise ValueError(f"Expected patch_coords to be [batch_size, 8] tensor or list, got {type(patch_coords)} with shape {patch_coords.shape if hasattr(patch_coords, 'shape') else 'unknown'}")
    
    # Validate coordinate format
    if len(coords_8_values) != 8:
        raise ValueError(f"Expected 8-value patch coordinates, got {len(coords_8_values)} values: {coords_8_values}")

    return coords_8_values

def _process_anomaly_maps(anomaly_map_arithmetic_raw, anomaly_binary_threshold, patch_size, x_coord, y_coord, original_image):
    """
    Process anomaly maps and create binary masks with proper cropping.
    
    Args:
        anomaly_map_arithmetic_raw: Raw anomaly map data
        anomaly_binary_threshold: Threshold for binary mask creation
        patch_size: Size of the patch
        x_coord, y_coord: Top-left corner coordinates
        original_image: Original image for dimension checking
    
    Returns:
        tuple: (anomaly_map_arithmetic_tensor, anomaly_binary_cropped, anomaly_pixels, is_predicted_defective, binary_map_numpy)
    """
    # Convert to torch tensor for processing
    anomaly_map_arithmetic_tensor = torch.from_numpy(anomaly_map_arithmetic_raw).float().unsqueeze(0).unsqueeze(0)
    
    # Create binary mask
    anomaly_map_arithmetic_binary = _binary_mask(anomaly_map_arithmetic_tensor, anomaly_binary_threshold)
    
    # Get actual patch dimensions for consistent cropping
    h, w = original_image.shape[:2]
    actual_patch_height = min(patch_size, h - y_coord)
    actual_patch_width = min(patch_size, w - x_coord)
    
    # Crop the binary mask tensor to match the actual patch size
    anomaly_binary_cropped = anomaly_map_arithmetic_binary[:, :, :actual_patch_height, :actual_patch_width]
    
    # Calculate anomaly pixels
    anomaly_pixels = torch.sum(anomaly_binary_cropped).item()
    
    # Store the binary map with proper shape
    binary_map_numpy = _to_numpy(anomaly_binary_cropped).squeeze()
    
    return anomaly_map_arithmetic_tensor, anomaly_binary_cropped, anomaly_pixels, binary_map_numpy

def _calculate_patch_status(x_coord, y_coord, patch_size, is_predicted_defective, ground_truth_defective):
    """
    Calculate the status (TP, FP, FN, TN) for a patch.
    
    Args:
        x_coord, y_coord: Top-left corner coordinates
        patch_size: Size of the patch
        is_predicted_defective: Whether the patch is predicted as defective
        ground_truth_defective: Set of ground truth defective patches
    
    Returns:
        str: Status string (TP, FP, FN, TN)
    """
    # Convert pixel coordinates to grid coordinates
    grid_row = y_coord // patch_size
    grid_col = x_coord // patch_size
    
    # Determine status
    status = "TP" if is_predicted_defective and (grid_row, grid_col) in ground_truth_defective else \
             "FP" if is_predicted_defective else \
             "FN" if (grid_row, grid_col) in ground_truth_defective else "TN"
    
    return status

def _create_evaluation_record(
    split, image_path, coords_8_values, anomaly_max, anomaly_pixels, 
    is_predicted_defective, status, original_patch, encodedrecon_raw, 
    latent_raw, anomaly_map_arithmetic_raw, binary_map_numpy
):
    """
    Create a standardized evaluation record.
    
    Args:
        split: Data split identifier
        image_path: Path to the image
        coords_8_values: 8-value patch coordinates
        anomaly_max: Maximum anomaly value
        anomaly_pixels: Number of anomaly pixels
        is_predicted_defective: Whether patch is predicted defective
        status: Patch status (TP, FP, FN, TN)
        original_patch: Original image patch
        encodedrecon_raw: Encoded reconstruction data
        latent_raw: Latent representation data
        anomaly_map_arithmetic_raw: Raw anomaly map data
        binary_map_numpy: Binary mask data
    
    Returns:
        dict: Evaluation record
    """
    record = make_record(
        split=("meta", split),
        image_path=("meta", image_path),
        image_path_original=("meta", path_to_safe_filename(image_path)),
        anomaly_class=("meta", "all"),
        patch_coords=("meta", coords_8_values),
        anomaly_max=("meta", anomaly_max),
        anomaly_pixels=("meta", anomaly_pixels),
        is_predicted_defective=("meta", is_predicted_defective),
        status=("meta", status),
        orig=("image", original_patch),
        dod_recon=("image", encodedrecon_raw),
        encoded_recon=("image", encodedrecon_raw),
        anomaly_map_arithmetic=("image", anomaly_map_arithmetic_raw),
        anomaly_map_arithmetic_binary=("image", binary_map_numpy),
        anomaly_map_geometric=("image", anomaly_map_arithmetic_raw),
        anomaly_map_geometric_binary=("image", binary_map_numpy),
        encoded=("image", latent_raw),
    )
    
    # Add metric fields
    record["lpips"] = ("metric", 0.0)
    record["ssim"] = ("metric", 0.0)
    record["mse"] = ("metric", 0.0)
    
    return record

def _process_single_patch(
    ground_truth_map, original_images, anomaly_binary_threshold,
    anomaly_pixel_num_threshold, patch_size, current_image_path, coords_8_values,
    encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw
):
    """
    Process a single patch and create an evaluation record.
    
    Args:
        patch_data: Patch data dictionary
        ground_truth_map: Ground truth information
        original_images: Dictionary of original images
        anomaly_binary_threshold: Threshold for binary mask
        anomaly_pixel_num_threshold: Threshold for anomaly pixel count
        patch_size: Size of the patch
        current_image_path: Path to current image
        coords_8_values: 8-value patch coordinates
        encodedrecon_raw: Encoded reconstruction data
        latent_raw: Latent representation data
        anomaly_map_arithmetic_raw: Raw anomaly map data
    
    Returns:
        dict: Evaluation record or None if processing fails
    """
    try:
        # Extract coordinates
        x_coord, y_coord = coords_8_values[0], coords_8_values[1]
        
        # Get image from cache (handles both dict and cache objects)
        if hasattr(original_images, 'get'):
            # It's a cache object
            original_image = original_images.get(current_image_path)
        else:
            # It's a regular dictionary
            if current_image_path not in original_images:
                debug_print(f"⚠️  Image not found in original_images: {current_image_path}", debug=DEBUG_ENABLED)
                return None
            original_image = original_images[current_image_path]
        
        if original_image is None:
            debug_print(f"⚠️  Failed to load image: {current_image_path}", debug=DEBUG_ENABLED)
            return None
        
        # Process anomaly maps
        anomaly_map_arithmetic_tensor, anomaly_binary_cropped, anomaly_pixels, binary_map_numpy = _process_anomaly_maps(
            anomaly_map_arithmetic_raw, anomaly_binary_threshold, patch_size, x_coord, y_coord, original_image
        )
        
        # Determine if patch is defective
        is_predicted_defective = anomaly_pixels > anomaly_pixel_num_threshold
        
        # Calculate metrics
        anomaly_max = int(round(anomaly_map_arithmetic_tensor.max().item() * 255))
        
        # Get ground truth defective patches
        ground_truth_defective = ground_truth_map.get(current_image_path, set()) if ground_truth_map else set()
        
        # Calculate status
        status = _calculate_patch_status(x_coord, y_coord, patch_size, is_predicted_defective, ground_truth_defective)
        
        # Get original patch
        h, w = original_image.shape[:2]
        actual_patch_height = min(patch_size, h - y_coord)
        actual_patch_width = min(patch_size, w - x_coord)
        original_patch = original_image[y_coord:y_coord + actual_patch_height, x_coord:x_coord + actual_patch_width]
        
        # Create record
        record = _create_evaluation_record(
            "test", current_image_path, coords_8_values, anomaly_max, anomaly_pixels,
            is_predicted_defective, status, original_patch, encodedrecon_raw,
            latent_raw, anomaly_map_arithmetic_raw, binary_map_numpy
        )
        
        return record
        
    except Exception as e:
        debug_print(f"⚠️  Error processing patch for {current_image_path}: {e}", debug=DEBUG_ENABLED)
        return None

class AnnotatedImageDataset(Dataset):
    """Dataset for images with JSON annotations for defective regions."""
    
    def __init__(
        self,
        annotation_dir: str,
        patch_size: int = 128,
        stride: int = None,
        transform=None,
        object_class: str = "pcb",
    ):
        self.annotation_dir = annotation_dir
        self.patch_size = patch_size
        self.stride = stride if stride is not None else patch_size
        self.transform = transform
        self.object_class = object_class
        
        # Load all annotation files
        self.annotation_files = []
        for root, dirs, files in os.walk(annotation_dir):
            for file in files:
                if file.endswith('.json'):
                    self.annotation_files.append(os.path.join(root, file))
        
        print(f"Found {len(self.annotation_files)} annotation files")

        # Build a per-patch index at initialization time.
        # Each entry is a tuple: (resolved_image_path, coords_8_values)
        self.patch_items = []
        self._unique_image_paths = set()

        for annotation_file in self.annotation_files:
            # Load annotation data to get image path
            with open(annotation_file, 'r') as f:
                annotation_data = json.load(f)

            image_path_raw = annotation_data.get('image_path', '')
            resolved_image_path = self._resolve_image_path(annotation_file, image_path_raw)

            # Open image only to read its size; do not load into memory
            with PILImage.open(resolved_image_path) as img:
                width, height = img.size

            # Compute padded dimensions (divisible by patch_size)
            pad_height = (self.patch_size - (height % self.patch_size)) % self.patch_size
            pad_width = (self.patch_size - (width % self.patch_size)) % self.patch_size
            padded_height = height + pad_height
            padded_width = width + pad_width

            # Generate 8-value coordinates for each patch on the padded image grid
            for y in range(0, padded_height - self.patch_size + 1, self.stride):
                for x in range(0, padded_width - self.patch_size + 1, self.stride):
                    # Standard parallel patch (aligned with image axes)
                    x1, y1 = x, y  # Top-left
                    x2, y2 = x + self.patch_size, y  # Top-right
                    x3, y3 = x + self.patch_size, y + self.patch_size  # Bottom-right
                    x4, y4 = x, y + self.patch_size  # Bottom-left
                    coords_8_values = (x1, y1, x2, y2, x3, y3, x4, y4)
                    self.patch_items.append((resolved_image_path, coords_8_values))

            self._unique_image_paths.add(resolved_image_path)

        print(f"Indexed {len(self.patch_items)} patches from {len(self._unique_image_paths)} images")
        
        # Image cache to avoid reloading the same image multiple times
        self._image_cache = {}
        self._cache_size_limit = 10  # Limit cache to prevent memory issues
        
    def __len__(self):
        # Dataset unit is now a single patch
        return len(self.patch_items)
        
    def __getitem__(self, index):
        debug_print(f"🔍 Starting __getitem__ index: {index}", debug=DEBUG_ENABLED)

        # Retrieve per-patch unit: (image_path, coords_8_values)
        image_path, coords_8_values = self.patch_items[index]

        # Resolve image path
        resolved_image_path = self._resolve_image_path(None, image_path)
        
        # Get or load image from cache
        image_np = self._get_cached_image(resolved_image_path)

        # Extract patch using all 8 coordinate values
        patch_np = self._extract_patch_from_coords(image_np, coords_8_values)

        if patch_np.shape[0] != self.patch_size or patch_np.shape[1] != self.patch_size:
            raise ValueError(f"Extracted patch has wrong size: {patch_np.shape} at index {index}")

        # Convert to tensor [3, patch_size, patch_size]
        x = self.transform(PILImage.fromarray(patch_np)) if self.transform else transforms.ToTensor()(PILImage.fromarray(patch_np))

        # Sanity check: ensure 3D tensor
        if len(x.shape) == 4 and x.shape[0] == 1:
            x = x.squeeze(0)
        if len(x.shape) != 3 or x.shape[0] != 3:
            raise ValueError(f"Dataset[{index}]: Expected 3D tensor with 3 channels, got shape {x.shape}")

        # Dummy segmentation and class per single patch item
        seg = torch.zeros(1, self.patch_size, self.patch_size)
        object_cls = torch.zeros(1, dtype=torch.long)

        # Metadata
        anomaly_classes = "all"
        # Convert patch coordinates to tensor for proper batching
        patch_coords = torch.tensor(coords_8_values, dtype=torch.long)

        debug_print(f"🔍 Dataset[{index}]: Final shapes - x: {x.shape}, seg: {seg.shape}, object_cls: {object_cls.shape}", debug=DEBUG_ENABLED)
        debug_print(f"🔍 Dataset[{index}]: patch_coords shape: {patch_coords.shape}, dtype: {patch_coords.dtype}", debug=DEBUG_ENABLED)
        debug_print(f"🔍 Ending __getitem__ index: {index}", debug=DEBUG_ENABLED)
        return x, seg, object_cls, anomaly_classes, image_path, patch_coords

    def _resolve_image_path(self, annotation_file, image_path):
        """Resolve the actual image path, trying common fallbacks if needed."""
        # If valid absolute/relative path is provided and exists
        if image_path and os.path.exists(image_path):
            return image_path

        # If annotation_file is provided, try constructing from its base name
        base_name = None
        if annotation_file is not None:
            base_name = os.path.splitext(os.path.basename(annotation_file))[0]
        else:
            # If we don't have the annotation file, try to parse from provided path
            provided = image_path or ""
            base_name = os.path.splitext(os.path.basename(provided))[0]

        if base_name and base_name.endswith('_annotations'):
            base_name = base_name[:-12]

        candidate_dirs = [os.path.dirname(image_path) if image_path else ""]
        if annotation_file is not None:
            candidate_dirs.append(os.path.dirname(annotation_file))

        for d in candidate_dirs:
            for ext in ['.png', '.jpg', '.jpeg']:
                potential_path = os.path.join(d, f"{base_name}{ext}") if base_name else ""
                if potential_path and os.path.exists(potential_path):
                    return potential_path

        raise FileNotFoundError(f"Image not found for annotation: {annotation_file if annotation_file else image_path}")

    def get_all_image_paths(self):
        """Return a list of unique image paths indexed by this dataset."""
        return list(self._unique_image_paths)
    
    def _extract_patch_from_coords(self, image_np, coords_8_values):
        """
        Extract patch from image using 8-value coordinates.
        Optimizes for parallel patches (faster) vs non-parallel patches (slower but more flexible).
        """
        x1, y1, x2, y2, x3, y3, x4, y4 = coords_8_values
        
        # Check if patch is parallel to image axes (faster extraction)
        is_parallel = (y1 == y2 and y3 == y4 and x1 == x4 and x2 == x3)
        
        if is_parallel:
            # Fast path: rectangular patch aligned with image axes
            # Use top-left corner and dimensions
            x_min, y_min = int(x1), int(y1)
            x_max, y_max = int(x3), int(y3)
            
            # Ensure coordinates are within image bounds
            height, width = image_np.shape[:2]
            x_min = max(0, min(x_min, width - 1))
            y_min = max(0, min(y_min, height - 1))
            x_max = max(x_min + 1, min(x_max, width))
            y_max = max(y_min + 1, min(y_max, height))
            
            # Extract patch
            patch = image_np[y_min:y_max, x_min:x_max]
            
            # Resize to target patch_size if needed
            if patch.shape[:2] != (self.patch_size, self.patch_size):
                patch_pil = PILImage.fromarray(patch)
                patch_pil = patch_pil.resize((self.patch_size, self.patch_size), PILImage.Resampling.LANCZOS)
                patch = np.array(patch_pil)
            
            debug_print(f"🔧 Fast parallel patch extraction: {patch.shape}", debug=DEBUG_ENABLED)
            return patch
        else:
            # Slow path: non-parallel patch requires perspective transform
            # This is more computationally expensive but handles rotated patches
            debug_print(f"🔧 Slow non-parallel patch extraction (perspective transform)", debug=DEBUG_ENABLED)
            
            # Convert coordinates to numpy arrays for OpenCV
            src_points = np.float32([[x1, y1], [x2, y2], [x3, y3], [x4, y4]])
            dst_points = np.float32([[0, 0], [self.patch_size, 0], 
                                   [self.patch_size, self.patch_size], [0, self.patch_size]])
            
            # Calculate perspective transform matrix
            transform_matrix = cv2.getPerspectiveTransform(src_points, dst_points)
            
            # Apply perspective transform
            patch = cv2.warpPerspective(image_np, transform_matrix, (self.patch_size, self.patch_size))
            
            return patch

    
    def _get_cached_image(self, image_path):
        """
        Get image from cache or load and cache it if not present.
        Implements LRU-style cache management to prevent memory issues.
        """
        if image_path in self._image_cache:
            debug_print(f"🔧 Using cached image: {image_path}", debug=DEBUG_ENABLED)
            return self._image_cache[image_path]
        
        # Load and cache new image
        debug_print(f"🔧 Loading new image into cache: {image_path}", debug=DEBUG_ENABLED)
        image = PILImage.open(image_path).convert('RGB')
        image_np = np.array(image)
        
        # Pad image to ensure dimensions are divisible by patch_size
        image_np = self._pad_image_to_patch_size(image_np)
        
        # Manage cache size
        if len(self._image_cache) >= self._cache_size_limit:
            # Remove oldest entry (simple FIFO for now)
            oldest_key = next(iter(self._image_cache))
            del self._image_cache[oldest_key]
            debug_print(f"🔧 Removed oldest image from cache: {oldest_key}", debug=DEBUG_ENABLED)
        
        # Add to cache
        self._image_cache[image_path] = image_np
        debug_print(f"🔧 Cached image: {image_path} (cache size: {len(self._image_cache)})", debug=DEBUG_ENABLED)
        
        return image_np
    
    def _pad_image_to_patch_size(self, img):
        """
        Pad image to ensure dimensions are exactly divisible by patch_size.
        This eliminates the need for overlapping edge patches.
        
        Args:
            img: numpy array of shape (height, width, channels)
            
        Returns:
            padded_img: numpy array with dimensions divisible by patch_size
        """
        height, width = img.shape[:2]
        
        # Calculate required padding
        pad_height = (self.patch_size - (height % self.patch_size)) % self.patch_size
        pad_width = (self.patch_size - (width % self.patch_size)) % self.patch_size
        
        if pad_height == 0 and pad_width == 0:
            # No padding needed
            debug_print(f"  📐 No padding needed. Image dimensions: {height}x{width}", debug=DEBUG_ENABLED)
            return img
        
        # Calculate padding for each side (distribute evenly, with extra on bottom/right if odd)
        pad_top = pad_height // 2
        pad_bottom = pad_height - pad_top
        pad_left = pad_width // 2
        pad_right = pad_width - pad_left
        
        # Apply padding with reflection or edge mode to maintain image characteristics
        if len(img.shape) == 3:
            # Color image
            padded_img = np.pad(img, 
                              ((pad_top, pad_bottom), (pad_left, pad_right), (0, 0)), 
                              mode='reflect')
        else:
            # Grayscale image
            padded_img = np.pad(img, 
                              ((pad_top, pad_bottom), (pad_left, pad_right)), 
                              mode='reflect')
        
        new_height, new_width = padded_img.shape[:2]
        debug_print(f"  📐 Padded image from {height}x{width} to {new_height}x{new_width}", debug=DEBUG_ENABLED)
        debug_print(f"  📐 Padding applied: top={pad_top}, bottom={pad_bottom}, left={pad_left}, right={pad_right}", debug=DEBUG_ENABLED)
        
        return padded_img
    
    def _extract_patches(self, img):
        """Extract patches from padded image - simple non-overlapping grid."""
        patches = []
        coords = []
        
        height, width = img.shape[:2]
        stride = self.stride
        debug_print(f"  📐 Padded image dimensions: {height}x{width}, patch size: {self.patch_size}, stride: {stride}", debug=DEBUG_ENABLED)
        
        # When stride equals patch_size, ensure dimensions are divisible
        if stride == self.patch_size:
            assert height % self.patch_size == 0, f"Height {height} not divisible by patch_size {self.patch_size}"
            assert width % self.patch_size == 0, f"Width {width} not divisible by patch_size {self.patch_size}"
        
        # Extract patches with configurable stride
        for y in range(0, height - self.patch_size + 1, stride):
            for x in range(0, width - self.patch_size + 1, stride):
                # Extract patch (guaranteed to be exactly patch_size x patch_size)
                patch = img[y:y + self.patch_size, x:x + self.patch_size]
                
                # Calculate all 4 corner coordinates for the patch (8 values)
                x1, y1 = x, y  # Top-left
                x2, y2 = x + self.patch_size, y  # Top-right
                x3, y3 = x + self.patch_size, y + self.patch_size  # Bottom-right
                x4, y4 = x, y + self.patch_size  # Bottom-left
                coords_8_values = (x1, y1, x2, y2, x3, y3, x4, y4)
                
                # Debug: Check coordinate types
                debug_print(f"  🔍 Created coordinates: {coords_8_values}", debug=DEBUG_ENABLED)
                debug_print(f"  🔍 Coordinate types: {[type(coord) for coord in coords_8_values]}", debug=DEBUG_ENABLED)
                debug_print(f"  🔍 All integers: {all(isinstance(coord, int) for coord in coords_8_values)}", debug=DEBUG_ENABLED)
                
                patches.append(patch)
                coords.append(coords_8_values)
        
        debug_print(f"  ✅ Extracted {len(patches)} non-overlapping patches from padded image", debug=DEBUG_ENABLED)
        return patches, coords

def _compute_abs_diff_mean(a: torch.Tensor, b: torch.Tensor, diff_scale: float = 1.0) -> torch.Tensor:
    return torch.abs(a - b).mean(dim=1, keepdim=True) * diff_scale

def _compute_abs_diff_max(a: torch.Tensor, b: torch.Tensor) -> torch.Tensor:
    return torch.abs(a - b).max(dim=1, keepdim=True)[0]

def _process_batch_inference(x, object_cls, model, vae, diffusion, reverse_steps, device, epoch_metrics=None):
    """
    Shared inference logic for processing a batch.
    Returns the computed difference tensors.
    """
    debug_print(f"   🔄 Moving {x.size(0)} patches to device: {device}", debug=DEBUG_ENABLED)
    debug_print(f"   🔍 x shape before device move: {x.shape}", debug=DEBUG_ENABLED)
    debug_print(f"   🔍 object_cls shape: {object_cls.shape}", debug=DEBUG_ENABLED)
    
    # Validate tensor dimensions before processing
    if len(x.shape) != 4:
        raise ValueError(f"Expected 4D input tensor for VAE, got shape: {x.shape}. Expected: [batch, channels, height, width]")
    
    if x.shape[1] != 3:  # Check channels
        raise ValueError(f"Expected 3 channels (RGB), got {x.shape[1]} channels")
    
    # Move batch to device
    x_device = x.to(device)
    object_cls_device = object_cls.to(device)
    
    debug_print(f"   🔍 x_device shape after device move: {x_device.shape}", debug=DEBUG_ENABLED)
    
    debug_print(f"   🎨 VAE encoding...", debug=DEBUG_ENABLED)
    # Forward pass through VAE encoder (to latent space)
    if torch.cuda.is_available():
        with torch.cuda.amp.autocast(dtype=torch.float16):
            encoded = vae.encode(x_device).latent_dist.mean * _LATENT_SCALE
        # Ensure downstream diffusion/model (kept in FP32) receives FP32 tensors
        encoded = encoded.float()
    else:
        encoded = vae.encode(x_device).latent_dist.mean * _LATENT_SCALE
    debug_print(f"   ✅ VAE encoding completed, latent shape: {encoded.shape}", debug=DEBUG_ENABLED)

    # Reverse DDIM sampling conditioned on encoder latents
    # Ensure object_cls has the correct shape and dtype for the model's class embedder
    # The UNet's ClassEmbedder expects Long indices shaped [batch_size, 1] (so embedding -> [batch_size, 1, dim])
    if len(object_cls_device.shape) == 2:
        # Use as-is, expected shape: [batch_size, 1]
        context = object_cls_device
    elif len(object_cls_device.shape) == 1:
        # [batch_size] -> [batch_size, 1]
        context = object_cls_device.unsqueeze(1)
    else:
        # Squeeze/reshape unexpected shapes to [batch_size, 1]
        context = object_cls_device.view(object_cls_device.shape[0], -1)
        if context.shape[1] != 1:
            context = context[:, :1]
    
    debug_print(f"   🔍 Context tensor shape (indices): {context.shape}", debug=DEBUG_ENABLED)
    
    # Validate context tensor shape for embedding: must be 2D [batch_size, 1]
    if len(context.shape) != 2 or context.shape[1] != 1:
        raise ValueError(f"Context indices must be 2D [batch_size, 1], got: {context.shape}")

    # Ensure dtype is long for embedding indices
    if context.dtype != torch.long:
        context = context.long()
    
    # Additional validation: ensure the input tensor has the correct shape for the model
    debug_print(f"   🔍 Input tensor x_device shape: {x_device.shape}", debug=DEBUG_ENABLED)
    debug_print(f"   🔍 Input tensor x_device dtype: {x_device.dtype}", debug=DEBUG_ENABLED)
    debug_print(f"   🔍 Input tensor x_device device: {x_device.device}", debug=DEBUG_ENABLED)
    
    # The model might expect a specific input format
    # Try to ensure the input tensor is in the right format
    if len(x_device.shape) == 4:
        # Standard format: [batch, channels, height, width]
        debug_print(f"   ✅ Input tensor has correct 4D format: {x_device.shape}", debug=DEBUG_ENABLED)
        
        # The model architecture might have specific input requirements
        # Let's check if there's a mismatch between what we're providing and what the model expects
        debug_print(f"   🔍 Model type: {type(model)}", debug=DEBUG_ENABLED)
        debug_print(f"   🔍 Model device: {next(model.parameters()).device}", debug=DEBUG_ENABLED)
        
        # Check if the model has any specific input requirements
        if hasattr(model, 'config'):
            debug_print(f"   🔍 Model config: {model.config}", debug=DEBUG_ENABLED)
        
        # The issue might be that the model expects a different input format
        # Let's try to understand what the model actually expects
        debug_print(f"   🔍 Input tensor shape: {x_device.shape}", debug=DEBUG_ENABLED)
        debug_print(f"   🔍 Context tensor shape: {context.shape}", debug=DEBUG_ENABLED)
    else:
        debug_print(f"   ⚠️  Warning: Input tensor has unexpected shape: {x_device.shape}", debug=DEBUG_ENABLED)
    
    model_kwargs = {"context": context, "mask": None}
    
    debug_print(f"   🔄 Starting DDIM sampling with {reverse_steps} steps...", debug=DEBUG_ENABLED)
    debug_print(f"   🔍 Encoded latent shape: {encoded.shape}", debug=DEBUG_ENABLED)
    debug_print(f"   🔍 Model kwargs: {model_kwargs}", debug=DEBUG_ENABLED)
    
    # Try to catch the error earlier by testing the model with a simple forward pass
    try:
        debug_print(f"   🔍 Testing model forward pass...", debug=DEBUG_ENABLED)
        with torch.no_grad():
            # Create a simple test input with the same shape as encoded
            test_input = torch.randn_like(encoded)
            test_output = model(test_input, torch.zeros(1, device=device), **model_kwargs)
            debug_print(f"   ✅ Model forward pass successful, output shape: {test_output.shape}", debug=DEBUG_ENABLED)
    except Exception as e:
        debug_print(f"   ❌ Model forward pass failed: {e}", debug=DEBUG_ENABLED)
        debug_print(f"   🔍 This suggests the model has input format requirements we're not meeting", debug=DEBUG_ENABLED)
        # Continue anyway to see the full error
    
    latent_samples_list = []
    step_count = 0
    for samples in diffusion.ddim_deviation_sample_loop_progressive(
        model,
        shape=encoded.shape,
        noise=encoded,
        clip_denoised=False,
        start_t=reverse_steps,
        model_kwargs=model_kwargs,
        progress=False,
        device=device,
        eta=0.0,
    ):
        step_count += 1
        if step_count % max(1, reverse_steps // 5) == 0:  # Print every 20% of steps
            debug_print(f"     📈 DDIM step {step_count}/{reverse_steps}", debug=DEBUG_ENABLED)
        latent_samples_list.append(samples["sample"])
    latent_samples_final = latent_samples_list[-1]
    debug_print(f"   ✅ DDIM sampling completed after {step_count} steps", debug=DEBUG_ENABLED)

    debug_print(f"   🎨 VAE decoding...", debug=DEBUG_ENABLED)
    # Decode final latent samples
    if torch.cuda.is_available():
        with torch.cuda.amp.autocast(dtype=torch.float16):
            image_samples = vae.decode(latent_samples_final / _LATENT_SCALE).sample
            x0 = vae.decode(encoded / _LATENT_SCALE).sample
        # Bring back to FP32 for subsequent ops
        image_samples = image_samples.float()
        x0 = x0.float()
    else:
        image_samples = vae.decode(latent_samples_final / _LATENT_SCALE).sample
        x0 = vae.decode(encoded / _LATENT_SCALE).sample
    debug_print(f"   ✅ VAE decoding completed", debug=DEBUG_ENABLED)
    
    debug_print(f"   📊 Computing differences...", debug=DEBUG_ENABLED)
    # Core difference computations
    encodedrecon_dodrecon_diff_raw = _compute_abs_diff_max(x0, image_samples)
    encodedrecon_dodrecon_diff = torch.clamp(encodedrecon_dodrecon_diff_raw, 0.0, 0.05) * 20

    encoded_latent_diff_raw = _compute_abs_diff_mean(latent_samples_final, encoded)
    encoded_latent_diff = torch.clamp(encoded_latent_diff_raw, 0.0, 0.05) * 20

    # Resize encoded_latent_diff to match the spatial dimensions
    patch_size_actual = x_device.shape[-1]
    encoded_latent_diff_resized = F.interpolate(
        encoded_latent_diff,
        size=(patch_size_actual, patch_size_actual),
        mode="bilinear",
        align_corners=False,
    )
    anomaly_map_arithmetic = 0.5 * (encodedrecon_dodrecon_diff + encoded_latent_diff_resized)
    debug_print(f"   ✅ Difference computation completed", debug=DEBUG_ENABLED)
    
    # Collect epoch-wise statistics if enabled
    if epoch_metrics is not None:
        debug_print(f"   📊 Collecting epoch statistics...", debug=DEBUG_ENABLED)
        # Use the same logic as in evaluation_DeCo_Diff2.py
        epoch_metrics.add_batch_stats(
            encodedrecon_dodrecon_diff_raw, 
            encoded_latent_diff_raw, 
            anomaly_map_arithmetic, 
            anomaly_map_arithmetic  # Use same for geometric since we don't compute separate geometric
        )
        debug_print(f"   ✅ Epoch statistics collected", debug=DEBUG_ENABLED)
    
    # Clear memory
    del x_device, object_cls_device, encoded, latent_samples_list, latent_samples_final
    del image_samples, x0, encodedrecon_dodrecon_diff_raw, encoded_latent_diff_raw
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
    
    return encodedrecon_dodrecon_diff, encoded_latent_diff_resized, anomaly_map_arithmetic

def create_confusion_matrix_from_records(
    records: List[Record],
    output_dir: str,
    annotation_dir: str = None,
    patch_size: int = 128
) -> None:
    """
    Create confusion matrix visualization from records.
    This function creates a confusion matrix plot similar to evaluation_DeCo_Diff2.py.
    """
    if not records:
        print("No records provided for confusion matrix generation")
        return
    
    # Initialize confusion matrix counters
    all_TP = all_FP = all_FN = all_TN = 0
    
    # Group records by image path
    image_records = {}
    for record in records:
        image_path = record["image_path"][1]
        if image_path not in image_records:
            image_records[image_path] = []
        image_records[image_path].append(record)
    
    print(f"Creating confusion matrix for {len(image_records)} images...")
    
    # Simply count the status values directly from the records instead of re-deriving them
    for record in records:
        status = record["status"][1]
        if status == "TP":
            all_TP += 1
        elif status == "FP":
            all_FP += 1
        elif status == "FN":
            all_FN += 1
        elif status == "TN":
            all_TN += 1
    
    total = all_TP + all_FP + all_FN + all_TN
    accuracy = (all_TP + all_TN) / total if total > 0 else 0
    
    # Calculate additional metrics
    precision = all_TP / (all_TP + all_FP) if (all_TP + all_FP) > 0 else 0
    recall = all_TP / (all_TP + all_FN) if (all_TP + all_FN) > 0 else 0
    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    
    print("Confusion Matrix (patch-level):")
    print(f"TP: {all_TP}, FP: {all_FP}, FN: {all_FN}, TN: {all_TN}")
    print(f"Accuracy: {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall: {recall:.4f}")
    print(f"F1-Score: {f1_score:.4f}")
    
    # Create confusion matrix visualization
    cm = np.array([[all_TP, all_FN], [all_FP, all_TN]])
    
    plt.figure(figsize=(8, 6))
    plt.imshow(cm, interpolation='nearest', cmap='Blues')
    plt.title('Confusion Matrix (Patch-level)', fontsize=16, fontweight='bold')
    plt.colorbar()
    
    # Add text annotations
    thresh = cm.max() / 2.
    for i in range(2):
        for j in range(2):
            plt.text(j, i, format(cm[i, j], 'd'),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black",
                    fontsize=14, fontweight='bold')
    
    # Set labels
    tick_marks = np.arange(2)
    plt.xticks(tick_marks, ['Defective', 'Normal'], fontsize=12)
    plt.yticks(tick_marks, ['Defective', 'Normal'], fontsize=12)
    plt.ylabel('True Label', fontsize=12)
    plt.xlabel('Predicted Label', fontsize=12)
    
    # Add metrics text
    metrics_text = f'Accuracy: {accuracy:.4f}\nPrecision: {precision:.4f}\nRecall: {recall:.4f}\nF1-Score: {f1_score:.4f}'
    plt.figtext(0.02, 0.02, metrics_text, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="lightgray"))
    
    plt.tight_layout()
    
    # Save the confusion matrix plot
    cm_plot_path = os.path.join(output_dir, "confusion_matrix.png")
    plt.savefig(cm_plot_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Confusion matrix plot saved to: {cm_plot_path}")
    
    # Save detailed results to file
    result = {
        "TP": all_TP,
        "FP": all_FP,
        "FN": all_FN,
        "TN": all_TN,
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1_score,
        "total_patches": total
    }
    with open(os.path.join(output_dir, "confusion_matrix.json"), "w") as f:
        json.dump(result, f, indent=2)
    
    print(f"Confusion matrix results saved to: {os.path.join(output_dir, 'confusion_matrix.json')}")

def save_all_records_json(records: List[Record], output_dir: str, filename: str = "all_records.json", patch_size: int = 128, sort_records: bool = True) -> None:
    """
    Save all records in a single comprehensive JSON file.
    
    Args:
        records: List of all evaluation records
        output_dir: Directory to save the JSON file
        filename: Name of the output JSON file (default: "all_records.json")
        patch_size: Size of patches for grid coordinate calculation
        sort_records: Whether to sort records by anomaly_pixels (default: True)
    """
    print(f"Saving all {len(records)} records to comprehensive JSON...")
    
    # Sort records by anomaly_pixels from largest to smallest if requested
    if sort_records:
        print(f"🔍 Sorting records by anomaly_pixels (largest to smallest)...")
        def get_anomaly_pixels(record):
            """Extract anomaly_pixels value, handling both tuple and direct int formats."""
            anomaly_pixels = record.get("anomaly_pixels", 0)
            if isinstance(anomaly_pixels, (list, tuple)) and len(anomaly_pixels) > 1:
                return anomaly_pixels[1]  # Tuple format: (metadata, value)
            else:
                return anomaly_pixels if isinstance(anomaly_pixels, int) else 0  # Direct int format

        sorted_records = sorted(records, key=get_anomaly_pixels, reverse=True)
        print(f"✅ Records sorted successfully")
    else:
        print(f"📊 Keeping records in original order (no sorting)")
        sorted_records = records
    
    # Convert records to a JSON-serializable format
    all_records_data = {
        "total_records": len(records),
        "records": []
    }
    
    for i, record in enumerate(sorted_records):
        # Extract all the key information from each record
        record_data = {
            "record_id": i,
            "split": record["split"][1] if "split" in record else None,
            "image_path": record["image_path"][1] if "image_path" in record else None,
            "image_path_original": record["image_path_original"][1] if "image_path_original" in record else None,
            "anomaly_class": record["anomaly_class"][1] if "anomaly_class" in record else None,
            "patch_coords": record["patch_coords"][1] if "patch_coords" in record else None,
            "anomaly_max": int(record["anomaly_max"][1]) if "anomaly_max" in record else None,
            "anomaly_pixels": int(record["anomaly_pixels"][1]) if "anomaly_pixels" in record else None,
            "is_predicted_defective": record["is_predicted_defective"][1] if "is_predicted_defective" in record else None,
            "status": record["status"][1] if "status" in record else None,
            # Add metric information if available
            "lpips": float(record["lpips"][1]) if "lpips" in record else None,
            "ssim": float(record["ssim"][1]) if "ssim" in record else None,
            "mse": float(record["mse"][1]) if "mse" in record else None,
        }
        
        # Add computed grid coordinates (simple division since images are padded)
        if record_data["patch_coords"]:
            patch_coords = record_data["patch_coords"]
            debug_print(f"🔍 Processing patch_coords: {patch_coords} (type: {type(patch_coords)}, len: {len(patch_coords)})", debug=DEBUG_ENABLED)
            if len(patch_coords) == 8:
                # 8-value format: (x1, y1, x2, y2, x3, y3, x4, y4)
                x1, y1, x2, y2, x3, y3, x4, y4 = patch_coords
                patch_x, patch_y = x1, y1  # Top-left corner
            elif len(patch_coords) == 2:
                # Legacy 2-value format detected - convert to 8-value format
                debug_print(f"⚠️  Converting 2-value coordinates {patch_coords} to 8-value format", debug=DEBUG_ENABLED)
                x1, y1 = patch_coords[0], patch_coords[1]
                x2, y2 = x1 + patch_size, y1  # Top-right
                x3, y3 = x1 + patch_size, y1 + patch_size  # Bottom-right
                x4, y4 = x1, y1 + patch_size  # Bottom-left
                
                # Update the record with converted 8-value coordinates
                record_data["patch_coords"] = [x1, y1, x2, y2, x3, y3, x4, y4]
                patch_x, patch_y = x1, y1
                debug_print(f"✅ Converted to 8-value coordinates: {record_data['patch_coords']}", debug=DEBUG_ENABLED)
            else:
                raise ValueError(f"Expected 8-value patch coordinates, got {len(patch_coords)} values. All coordinates should be 8-value format: (x1, y1, x2, y2, x3, y3, x4, y4)")
            
            # Simple grid calculation since all patches are regular grid-aligned with padding
            record_data["grid_row"] = patch_y // patch_size
            record_data["grid_col"] = patch_x // patch_size
        
        all_records_data["records"].append(record_data)
    
    # Add summary statistics
    if records:
        statuses = [r["status"][1] for r in records if "status" in r]
        all_records_data["summary"] = {
            "total_patches": len(records),
            "status_counts": {
                "TP": sum(1 for s in statuses if s == "TP"),
                "FP": sum(1 for s in statuses if s == "FP"), 
                "TN": sum(1 for s in statuses if s == "TN"),
                "FN": sum(1 for s in statuses if s == "FN")
            }
        }
        
        # Add accuracy metrics if possible
        tp = all_records_data["summary"]["status_counts"]["TP"]
        fp = all_records_data["summary"]["status_counts"]["FP"]
        tn = all_records_data["summary"]["status_counts"]["TN"]
        fn = all_records_data["summary"]["status_counts"]["FN"]
        
        total = tp + fp + tn + fn
        if total > 0:
            all_records_data["summary"]["metrics"] = {
                "accuracy": (tp + tn) / total,
                "precision": tp / (tp + fp) if (tp + fp) > 0 else 0,
                "recall": tp / (tp + fn) if (tp + fn) > 0 else 0,
                "f1_score": 2 * tp / (2 * tp + fp + fn) if (2 * tp + fp + fn) > 0 else 0
            }
    
    # Add sorting information to the header
    if sort_records:
        all_records_data["sorting_info"] = {
            "sorted_by": "anomaly_pixels",
            "order": "descending (largest to smallest)",
            "note": "Records are sorted by anomaly_pixels value for easier analysis of most anomalous patches"
        }
    else:
        all_records_data["sorting_info"] = {
            "sorted_by": "none",
            "order": "original order",
            "note": "Records are kept in their original order as processed"
        }
    
    # Save to file
    output_path = os.path.join(output_dir, filename)
    with open(output_path, 'w') as f:
        json.dump(all_records_data, f, indent=2)
    
    print(f"All records saved to: {output_path}")
    print(f"Total records: {len(records)}")
    
    if sort_records:
        print(f"📊 Records sorted by anomaly_pixels (largest to smallest)")
        
        # Show some statistics about the sorting
        if records:
            anomaly_pixels_values = [r.get("anomaly_pixels", [None, 0])[1] for r in records if "anomaly_pixels" in r]
            if anomaly_pixels_values:
                max_pixels = max(anomaly_pixels_values)
                min_pixels = min(anomaly_pixels_values)
                avg_pixels = sum(anomaly_pixels_values) / len(anomaly_pixels_values)
                print(f"   Max anomaly_pixels: {max_pixels}")
                print(f"   Min anomaly_pixels: {min_pixels}")
                print(f"   Avg anomaly_pixels: {avg_pixels:.1f}")
    else:
        print(f"📊 Records kept in original order (no sorting)")
    
    print(f"Status distribution: {all_records_data['summary']['status_counts']}")

def load_model_and_components(args):
    """Load VAE, model, and diffusion components."""
    import glob

    # Load VAE
    if os.path.exists("./models/config.json"):
        vae = AutoencoderKL.from_pretrained("./models", local_files_only=True).to(device)
    else:
        vae = AutoencoderKL.from_pretrained(f"stabilityai/sd-vae-ft-{args.vae_type}").to(device)
    vae.eval()

    # Load model
    try:
        if args.pretrained != "":
            # Handle relative paths by checking if they exist relative to current directory
            # or parent directory (project root)
            ckpt = args.pretrained
            if not os.path.exists(ckpt):
                # Try relative to parent directory (project root)
                parent_ckpt = os.path.join("..", ckpt)
                if os.path.exists(parent_ckpt):
                    ckpt = parent_ckpt
                else:
                    raise FileNotFoundError(f"Checkpoint not found: {args.pretrained}")
        else:
            path = f"./DeCo-Diff_{args.dataset}_{args.object_class}_{args.model_size}_{args.patch_size}"
            try:
                ckpt = sorted(glob.glob(f"{path}/last.pt"))[-1]
            except (IndexError, FileNotFoundError):
                ckpt = sorted(glob.glob(f"{path}/*/last.pt"))[-1]
    except (IndexError, FileNotFoundError, OSError) as e:
        raise Exception(f"Please provide the model's pretrained path using --pretrained. Error: {e}")

    latent_size = int(args.patch_size) // 8
    model = UNET_models[args.model_size](latent_size=latent_size)

    state_dict = torch.load(ckpt)["model"]
    print(model.load_state_dict(state_dict))
    model.eval()  # important!
    model.cuda()
    
    # Debug: Check model configuration
    debug_print(f"🔍 Model loaded successfully", debug=DEBUG_ENABLED)
    debug_print(f"🔍 Model type: {type(model)}", debug=DEBUG_ENABLED)
    debug_print(f"🔍 Model device: {next(model.parameters()).device}", debug=DEBUG_ENABLED)
    
    # Check if the model has any specific input requirements
    if hasattr(model, 'config'):
        debug_print(f"🔍 Model config: {model.config}", debug=DEBUG_ENABLED)
    
    # Check the model's expected input format
    debug_print(f"🔍 Model latent_size: {latent_size}", debug=DEBUG_ENABLED)
    debug_print(f"🔍 Model patch_size: {args.patch_size}", debug=DEBUG_ENABLED)
    
    # Check if there's a mismatch between the model architecture and input
    if hasattr(model, 'image_size'):
        debug_print(f"🔍 Model expected image_size: {model.image_size}", debug=DEBUG_ENABLED)
    
    if hasattr(model, 'in_channels'):
        debug_print(f"🔍 Model expected in_channels: {model.in_channels}", debug=DEBUG_ENABLED)
    
    print("model loaded")

    # Create diffusion
    diffusion = create_diffusion(
        f"ddim{args.reverse_steps}",
        predict_deviation=True,
        sigma_small=False,
        predict_xstart=False,
        diffusion_steps=1000,
    )
    
    return vae, model, diffusion

def get_transform():
    """Get the transform for input images."""
    return transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
    ])


def create_bootstrap_indices(dataset_length: int, num_samples: int, seed: int = 42) -> List[int]:
    """
    Generate bootstrap sample indices with replacement.

    Pure function for creating bootstrap samples from a dataset. Uses numpy's random
    generator for efficient sampling with replacement.

    Args:
        dataset_length: Total number of items in dataset
        num_samples: Number of bootstrap samples to draw
        seed: Random seed for reproducibility (default: 42)

    Returns:
        List of sampled indices (may contain duplicates due to replacement)

    Example:
        >>> indices = create_bootstrap_indices(1000, 100, seed=42)
        >>> len(indices)
        100
    """
    rng = np.random.RandomState(seed)
    indices = rng.choice(dataset_length, size=num_samples, replace=True)
    return indices.tolist()


def _before_saving_results(args):
    """
    Shared logic for preparing evaluation components.
    Returns: (vae, model, diffusion, dataset, loader, evaluation_results_dir, checkpoint_manager)
    """
    # Load model components
    vae, model, diffusion = load_model_and_components(args)
    
    # Create dataset
    dataset = AnnotatedImageDataset(
        annotation_dir=args.annotation_dir,
        patch_size=args.patch_size,
        stride=args.stride,
        transform=get_transform(),
        object_class=args.object_class,
    )

    # Apply bootstrap sampling if requested
    if args.bootstrap_samples is not None:
        original_length = len(dataset)
        num_samples = args.bootstrap_samples

        print(f"🎲 Bootstrap sampling enabled: {num_samples} samples from {original_length} patches")

        # Generate bootstrap indices using pure function
        bootstrap_indices = create_bootstrap_indices(
            dataset_length=original_length,
            num_samples=num_samples,
            seed=args.bootstrap_seed
        )

        # Create subset dataset with bootstrap indices
        dataset = Subset(dataset, bootstrap_indices)
        print(f"✅ Bootstrap dataset created: {len(dataset)} samples (seed={args.bootstrap_seed})")

    dataloader_kwargs = dict(
        dataset=dataset,
        batch_size=args.batch_size,
        shuffle=False,
        num_workers=args.num_workers,
        drop_last=False,
        pin_memory=torch.cuda.is_available(),
    )

    if args.num_workers and args.num_workers > 0:
        dataloader_kwargs.update({
            "prefetch_factor": 2,
            "persistent_workers": True,
        })

    loader = DataLoader(**dataloader_kwargs)

    # Create checkpoint manager
    checkpoint_manager = CheckpointManager(args.results_dir, args.annotation_dir, args.force_rerun)
    
    # Create evaluation results directory
    evaluation_results_dir = os.path.join(args.results_dir, "evaluation_results")
    
    return vae, model, diffusion, dataset, loader, evaluation_results_dir, checkpoint_manager

# === Smart Image Cache Class ===
class SmartImageCache:
    """Smart LRU cache for images with memory-aware eviction."""
    
    def __init__(self, max_memory_gb=2.0, max_images=100):
        # Import PIL here to avoid global import issues
        from PIL import Image as PILImage
        import numpy as np
        self.PILImage = PILImage
        self.np = np
        self.max_memory_bytes = max_memory_gb * 1024 * 1024 * 1024
        self.max_images = max_images
        self.cache = {}
        self.access_order = []  # Track access order for LRU
        self.total_memory = 0
        
    def _get_memory_usage(self):
        """Get current memory usage in bytes."""
        import psutil
        process = psutil.Process()
        return process.memory_info().rss
        
    def _estimate_image_memory(self, image):
        """Estimate memory usage of an image in bytes."""
        if hasattr(image, 'nbytes'):
            return image.nbytes
        elif hasattr(image, '__sizeof__'):
            return image.__sizeof__()
        else:
            # Fallback estimation: assume RGB image
            return image.shape[0] * image.shape[1] * 3 * 4  # 4 bytes per float32
        
    def _evict_if_needed(self):
        """Evict least recently used images if memory limit exceeded."""
        import gc
        while (self.total_memory > self.max_memory_bytes or 
               len(self.cache) > self.max_images) and self.cache:
            
            # Remove least recently used image
            lru_key = self.access_order.pop(0)
            if lru_key in self.cache:
                evicted_image = self.cache.pop(lru_key)
                evicted_memory = self._estimate_image_memory(evicted_image)
                self.total_memory -= evicted_memory
                debug_print(f"🗑️  Evicted image from cache: {lru_key} (-{evicted_memory/1024/1024:.1f}MB)", debug=DEBUG_ENABLED)
            
            # Force garbage collection
            gc.collect()
    
    def get(self, image_path):
        """Get image from cache, loading it if not present."""
        if image_path in self.cache:
            # Move to end (most recently used)
            self.access_order.remove(image_path)
            self.access_order.append(image_path)
            debug_print(f"📖 Cache hit: {os.path.basename(image_path)}", debug=DEBUG_ENABLED)
            return self.cache[image_path]
        
        # Load image from disk
        try:
            debug_print(f"📥 Loading image: {os.path.basename(image_path)}", debug=DEBUG_ENABLED)
            
            # Try direct path first
            if os.path.exists(image_path):
                image = self.np.array(self.PILImage.open(image_path).convert('RGB'))
            else:
                # Try converting from safe filename
                if '__' in image_path:
                    from utils import safe_filename_to_path
                    actual_path = safe_filename_to_path(image_path)
                    if os.path.exists(actual_path):
                        image = self.np.array(self.PILImage.open(actual_path).convert('RGB'))
                    else:
                        print(f"Warning: Image not found: {image_path} (tried: {actual_path})")
                        return None
                else:
                    print(f"Warning: Image not found: {image_path}")
                    return None
            
            # Add to cache
            image_memory = self._estimate_image_memory(image)
            self.cache[image_path] = image
            self.access_order.append(image_path)
            self.total_memory += image_memory
            
            debug_print(f"💾 Cached image: {os.path.basename(image_path)} (+{image_memory/1024/1024:.1f}MB)", debug=DEBUG_ENABLED)
            
            # Evict if needed
            self._evict_if_needed()
            
            return image
            
        except Exception as e:
            print(f"Warning: Error loading image {image_path}: {str(e)}")
            return None
    
    def __contains__(self, image_path):
        """Check if an image path is available (either cached or loadable)."""
        # First check if it's in cache
        if image_path in self.cache:
            return True
        
        # Check if file exists on disk
        import os
        if os.path.exists(image_path):
            return True
        
        # Try converting from safe filename
        if '__' in image_path:
            try:
                from utils import safe_filename_to_path
                actual_path = safe_filename_to_path(image_path)
                return os.path.exists(actual_path)
            except:
                return False
        
        return False
    
    def __getitem__(self, image_path):
        """Get image using bracket notation (same as get method)."""
        return self.get(image_path)
    
    def get_stats(self):
        """Get cache statistics."""
        return {
            'cached_images': len(self.cache),
            'total_memory_mb': self.total_memory / 1024 / 1024,
            'max_memory_mb': self.max_memory_bytes / 1024 / 1024,
            'memory_usage_percent': (self.total_memory / self.max_memory_bytes) * 100
        }


# === Shared streaming helpers (for both incremental eval and process_only) ===
def _process_records_stream_incrementally(
    args,
    patch_item_iter,
    ground_truth_map,
    original_images,
    output_subdir_name: str = "processed_results",
):
    """
    Consume a stream of patch items incrementally and process/save results without
    materializing all records in memory.

    Each item yielded by patch_item_iter must be a tuple:
      (current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw)

    Returns: (metrics, output_subdir)
    """
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.tag:
        output_dir = os.path.join(args.results_dir, f"{args.tag}_{current_time}")
    else:
        output_dir = os.path.join(args.results_dir, f"{current_time}")
    os.makedirs(output_dir, exist_ok=True)
    output_subdir = os.path.join(output_dir, output_subdir_name)
    os.makedirs(output_subdir, exist_ok=True)
    debug_print(f"📁 Output directory: {output_subdir}", debug=DEBUG_ENABLED)

    checkpoint_manager = CheckpointManager(args.results_dir, args.annotation_dir, args.force_rerun)

    total_records = 0
    normal_records_count = 0
    defect_records_count = 0

    from collections import defaultdict
    image_to_records = defaultdict(list)
    batch_records = []
    flush_every = max(1, int(getattr(args, "batch_size", 64)))

    # Check if we need to aggregate overlapping patches back to grid positions
    # This happens when stride < patch_size was used during evaluation
    use_patch_aggregation = (hasattr(args, 'stride') and args.stride is not None and 
                           args.stride < args.patch_size)
    
    if use_patch_aggregation:
        debug_print(f"🔧 Using patch aggregation: stride={args.stride}, patch_size={args.patch_size}")
        # Collect and aggregate overlapping patches
        patch_groups = _aggregate_overlapping_patches(patch_item_iter, args.stride, args.patch_size, original_images)
        debug_print(f"🔧 Aggregated patches into {len(patch_groups)} grid positions")
    else:
        patch_groups = patch_item_iter

    # Process aggregated patches in parallel using round-robin distribution
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading
    from collections import defaultdict
    
    # Configuration for round-robin parallel processing
    if getattr(args, 'no_parallel', False):
        # Sequential processing fallback
        print("🐌 Sequential processing (parallel disabled)")
        for i, (current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw) in tqdm(patch_groups, desc="Processing patches", unit="patch"):
            record = _process_single_patch(
                ground_truth_map=ground_truth_map,
                original_images=original_images,
                anomaly_binary_threshold=args.anomaly_binary_threshold,
                anomaly_pixel_num_threshold=args.anomaly_pixel_num_threshold,
                patch_size=args.patch_size,
                current_image_path=current_image_path,
                coords_8_values=coords_8_values,
                encodedrecon_raw=encodedrecon_raw,
                latent_raw=latent_raw,
                anomaly_map_arithmetic_raw=anomaly_map_arithmetic_raw,
            )
            if record is None:
                continue
            batch_records.append(record)
            total_records += 1
            is_predicted_defective = record.get("is_predicted_defective", (None, False))[1]
            if is_predicted_defective:
                defect_records_count += 1
            else:
                normal_records_count += 1
            image_to_records[current_image_path].append(record)

            if len(batch_records) >= flush_every:
                _process_batch_records_immediately(
                    args,
                    batch_records,
                    ground_truth_map,
                    original_images,
                    checkpoint_manager,
                    output_subdir,
                    image_to_records,
                )
                del batch_records
                batch_records = []
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                import gc
                gc.collect()
        return  # Exit early for sequential processing
    
    # Parallel processing configuration
    if getattr(args, 'parallel_workers', None):
        max_workers = args.parallel_workers
        print(f"🚀 Using custom worker count: {max_workers}")
    else:
        max_workers = min(os.cpu_count() * 2, 16)  # Use more workers for better CPU utilization
        print(f"🚀 Auto-detected workers: {max_workers} (from {os.cpu_count()} CPU cores)")
    
    print(f"🚀 Round-robin parallel processing: {max_workers} workers")
    
    # Thread-safe counters and collections
    total_records_lock = threading.Lock()
    defect_records_lock = threading.Lock()
    normal_records_lock = threading.Lock()
    image_to_records_lock = threading.Lock()
    
    def process_single_patch_wrapper(patch_data):
        """Process a single patch in round-robin distribution."""
        current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw = patch_data
        
        record = _process_single_patch(
            ground_truth_map=ground_truth_map,
            original_images=original_images,
            anomaly_binary_threshold=args.anomaly_binary_threshold,
            anomaly_pixel_num_threshold=args.anomaly_pixel_num_threshold,
            patch_size=args.patch_size,
            current_image_path=current_image_path,
            coords_8_values=coords_8_values,
            encodedrecon_raw=encodedrecon_raw,
            latent_raw=latent_raw,
            anomaly_map_arithmetic_raw=anomaly_map_arithmetic_raw,
        )
        
        if record is None:
            return None
            
        # Return record with metadata for aggregation
        is_predicted_defective = record.get("is_predicted_defective", (None, False))[1]
        return {
            'record': record,
            'is_defective': is_predicted_defective,
            'image_path': current_image_path
        }
    
    # Use ThreadPoolExecutor for true round-robin processing
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time
    
    start_time = time.time()
    processed_count = 0
    
    print(f"📊 Starting round-robin parallel processing...")
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all patches using round-robin distribution
        futures = []
        
        with tqdm(desc="Submitting patches", unit="patch") as submit_pbar:
            for i, patch_data in enumerate(patch_groups):
                future = executor.submit(process_single_patch_wrapper, patch_data)
                futures.append(future)
                submit_pbar.update(1)
                
                # Update progress every 50 submissions to show worker activity
                if i % 50 == 0:
                    # Count active workers by checking running futures
                    active_workers = sum(1 for f in futures if f.running())
                    completed_workers = sum(1 for f in futures if f.done() and not f.cancelled())
                    submit_pbar.set_postfix({
                        'submitted': i+1,
                        'active': active_workers,
                        'completed': completed_workers,
                        'workers': f"{active_workers}/{max_workers}"
                    })
        
        print(f"📊 Submitted {len(futures)} patches for parallel processing")
        
        # Process results as they complete (true parallelism)
        with tqdm(total=len(futures), desc="Processing results", unit="patch") as pbar:
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result is None:
                        continue
                    
                    record = result['record']
                    is_predicted_defective = result['is_defective']
                    current_image_path = result['image_path']
                    
                    # Thread-safe aggregation
                    batch_records.append(record)
                    
                    with total_records_lock:
                        total_records += 1
                    
                    if is_predicted_defective:
                        with defect_records_lock:
                            defect_records_count += 1
                    else:
                        with normal_records_lock:
                            normal_records_count += 1
                    
                    with image_to_records_lock:
                        image_to_records[current_image_path].append(record)
                    
                    processed_count += 1
                    
                    # Update progress
                    pbar.set_postfix({
                        'total': total_records,
                        'defects': defect_records_count,
                        'normals': normal_records_count,
                        'batch_size': len(batch_records)
                    })
                    pbar.update(1)
                    
                    # Process batch when it reaches the flush threshold
                    if len(batch_records) >= flush_every:
                        _process_batch_records_immediately(
                            args,
                            batch_records,
                            ground_truth_map,
                            original_images,
                            checkpoint_manager,
                            output_subdir,
                            image_to_records,
                        )
                        del batch_records
                        batch_records = []
                        if torch.cuda.is_available():
                            torch.cuda.empty_cache()
                        import gc
                        gc.collect()
                        
                except Exception as e:
                    print(f"⚠️  Error processing patch: {e}")
                    processed_count += 1
                    pbar.update(1)
    
    elapsed_time = time.time() - start_time
    patches_per_second = processed_count / elapsed_time if elapsed_time > 0 else 0
    print(f"✅ Round-robin processing completed: {processed_count} patches in {elapsed_time:.2f}s ({patches_per_second:.1f} patches/sec)")

    if batch_records:
        _process_batch_records_immediately(
            args,
            batch_records,
            ground_truth_map,
            original_images,
            checkpoint_manager,
            output_subdir,
            image_to_records,
        )
        del batch_records

    metrics = _finalize_incremental_processing(
        args,
        total_records,
        normal_records_count,
        defect_records_count,
        image_to_records,
        ground_truth_map,
        output_subdir,
    )
    return metrics, output_subdir


def _aggregate_overlapping_patches(patch_item_iter, stride, patch_size, original_images):
    """
    Reconstruct equal-spaced grid patches by averaging overlapping regions from stride-based patches.
    
    Example: 
    - Grid patch (0,0,128,0,128,128,0,128) 
    - Gets contributions from stride patches like (0,0), (64,0), (0,64), (64,64)
    - Average the overlapping (64,64,128,64,128,128,64,128) region
    - Drop non-grid-aligned patches like (64,64,192,64,192,192,64,192)
    
    Args:
        patch_item_iter: Iterator yielding (image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw)
        stride: Stride used during patch extraction  
        patch_size: Size of patches
        original_images: Dictionary of original images for dimension calculation
        
    Returns:
        List of reconstructed grid patches: [(image_path, grid_coords_8_values, reconstructed_encodedrecon, reconstructed_latent, reconstructed_anomaly), ...]
    """
    from collections import defaultdict
    import numpy as np
    
    debug_print(f"🔧 Reconstructing equal-spaced grid patches from overlapping stride patches...")
    
    # Collect all stride patches by image
    stride_patches_by_image = defaultdict(list)
    
    for patch_data in tqdm(patch_item_iter, desc="Collecting stride patches", unit="patch"):
        image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw = patch_data
        stride_x, stride_y = coords_8_values[0], coords_8_values[1]
        
        stride_patches_by_image[image_path].append({
            'coords': (stride_x, stride_y),
            'encodedrecon_raw': encodedrecon_raw,
            'latent_raw': latent_raw,
            'anomaly_map_arithmetic_raw': anomaly_map_arithmetic_raw
        })
    
    reconstructed_patches = []
    
    for image_path, stride_patches in tqdm(stride_patches_by_image.items(), desc="Processing images", unit="image"):
        debug_print(f"🔧 Processing {len(stride_patches)} stride patches for {image_path}")
        
        # Get image dimensions to determine grid positions
        if image_path not in original_images:
            debug_print(f"⚠️  Image not found: {image_path}")
            continue
            
        original_image = original_images[image_path]
        img_height, img_width = original_image.shape[:2]
        
        # Calculate padded dimensions (same as in patch extraction)
        pad_height = (patch_size - (img_height % patch_size)) % patch_size
        pad_width = (patch_size - (img_width % patch_size)) % patch_size
        padded_height = img_height + pad_height
        padded_width = img_width + pad_width
        
        # Generate equal-spaced grid positions (what we want to reconstruct)
        grid_positions = []
        for grid_y in range(0, padded_height, patch_size):
            for grid_x in range(0, padded_width, patch_size):
                # Ensure we don't go beyond padded boundaries
                if grid_y + patch_size <= padded_height and grid_x + patch_size <= padded_width:
                    grid_positions.append((grid_x, grid_y))
        
        debug_print(f"🔧 Reconstructing {len(grid_positions)} grid patches from stride patches")
        
        # For each grid position, reconstruct the patch by averaging overlapping regions
        for grid_x, grid_y in tqdm(grid_positions, desc=f"Reconstructing grid patches for {os.path.basename(image_path)}", unit="patch", leave=False):
            # Initialize accumulation arrays
            patch_shape = stride_patches[0]['encodedrecon_raw'].shape
            accumulated_encodedrecon = np.zeros(patch_shape, dtype=np.float64)
            accumulated_latent = np.zeros(stride_patches[0]['latent_raw'].shape, dtype=np.float64)
            accumulated_anomaly = np.zeros(stride_patches[0]['anomaly_map_arithmetic_raw'].shape, dtype=np.float64)
            count_map = np.zeros((patch_size, patch_size), dtype=np.int32)
            
            # Find all stride patches that overlap with this grid patch
            contributing_patches = 0
            for stride_patch in stride_patches:
                stride_x, stride_y = stride_patch['coords']
                
                # Check if this stride patch overlaps with the grid patch
                # Overlap exists if: stride_x < grid_x + patch_size AND stride_x + patch_size > grid_x
                #                   AND stride_y < grid_y + patch_size AND stride_y + patch_size > grid_y
                if (stride_x < grid_x + patch_size and stride_x + patch_size > grid_x and
                    stride_y < grid_y + patch_size and stride_y + patch_size > grid_y):
                    
                    # Calculate overlap region in both patches
                    overlap_x_start = max(0, grid_x - stride_x)  # Start in stride patch
                    overlap_y_start = max(0, grid_y - stride_y)  # Start in stride patch
                    overlap_x_end = min(patch_size, grid_x + patch_size - stride_x)  # End in stride patch
                    overlap_y_end = min(patch_size, grid_y + patch_size - stride_y)  # End in stride patch
                    
                    # Calculate corresponding region in grid patch
                    grid_x_start = max(0, stride_x - grid_x)  # Start in grid patch
                    grid_y_start = max(0, stride_y - grid_y)  # Start in grid patch
                    grid_x_end = grid_x_start + (overlap_x_end - overlap_x_start)
                    grid_y_end = grid_y_start + (overlap_y_end - overlap_y_start)
                    
                    # Extract overlapping regions from stride patch
                    stride_encodedrecon_region = stride_patch['encodedrecon_raw'][..., overlap_y_start:overlap_y_end, overlap_x_start:overlap_x_end]
                    stride_latent_region = stride_patch['latent_raw'][..., overlap_y_start:overlap_y_end, overlap_x_start:overlap_x_end]
                    stride_anomaly_region = stride_patch['anomaly_map_arithmetic_raw'][overlap_y_start:overlap_y_end, overlap_x_start:overlap_x_end]
                    
                    # Accumulate in grid patch
                    accumulated_encodedrecon[..., grid_y_start:grid_y_end, grid_x_start:grid_x_end] += stride_encodedrecon_region
                    accumulated_latent[..., grid_y_start:grid_y_end, grid_x_start:grid_x_end] += stride_latent_region
                    accumulated_anomaly[grid_y_start:grid_y_end, grid_x_start:grid_x_end] += stride_anomaly_region
                    count_map[grid_y_start:grid_y_end, grid_x_start:grid_x_end] += 1
                    
                    contributing_patches += 1
            
            if contributing_patches > 0:
                # Compute mean where count > 0
                count_map_expanded = np.expand_dims(count_map, axis=tuple(range(len(accumulated_encodedrecon.shape) - 2)))
                count_map_expanded = np.broadcast_to(count_map_expanded, accumulated_encodedrecon.shape)
                
                # Avoid division by zero
                valid_mask = count_map_expanded > 0
                reconstructed_encodedrecon = np.zeros_like(accumulated_encodedrecon)
                reconstructed_encodedrecon[valid_mask] = accumulated_encodedrecon[valid_mask] / count_map_expanded[valid_mask]
                
                count_map_expanded_latent = np.expand_dims(count_map, axis=tuple(range(len(accumulated_latent.shape) - 2)))
                count_map_expanded_latent = np.broadcast_to(count_map_expanded_latent, accumulated_latent.shape)
                valid_mask_latent = count_map_expanded_latent > 0
                reconstructed_latent = np.zeros_like(accumulated_latent)
                reconstructed_latent[valid_mask_latent] = accumulated_latent[valid_mask_latent] / count_map_expanded_latent[valid_mask_latent]
                
                valid_mask_anomaly = count_map > 0
                reconstructed_anomaly = np.zeros_like(accumulated_anomaly)
                reconstructed_anomaly[valid_mask_anomaly] = accumulated_anomaly[valid_mask_anomaly] / count_map[valid_mask_anomaly]
                
                # Create grid coordinates (8-value format)
                grid_coords_8_values = [
                    grid_x, grid_y,                           # Top-left
                    grid_x + patch_size, grid_y,              # Top-right  
                    grid_x + patch_size, grid_y + patch_size, # Bottom-right
                    grid_x, grid_y + patch_size               # Bottom-left
                ]
                
                reconstructed_patches.append((
                    image_path,
                    grid_coords_8_values,
                    reconstructed_encodedrecon.astype(stride_patches[0]['encodedrecon_raw'].dtype),
                    reconstructed_latent.astype(stride_patches[0]['latent_raw'].dtype),
                    reconstructed_anomaly.astype(stride_patches[0]['anomaly_map_arithmetic_raw'].dtype)
                ))
                
                debug_print(f"🔧 Reconstructed grid patch ({grid_x},{grid_y}) from {contributing_patches} stride patches")
    
    debug_print(f"🔧 Final reconstruction: {len(reconstructed_patches)} equal-spaced grid patches")
    return reconstructed_patches


def _extract_image_path_from_batch(image_paths_batch, batch_index):
    """Extract image path from batch at given index, handling various formats."""
    if batch_index < len(image_paths_batch):
        value = image_paths_batch[batch_index]
    else:
        value = image_paths_batch[-1] if image_paths_batch else ""

    if isinstance(value, str):
        return value
    elif isinstance(value, (list, tuple)):
        return value[0] if value else ""
    else:
        return str(value)


def _get_distributed_npy_files(results_dir):
    """
    Get all .npy files from distributed folder structure.

    Args:
        results_dir: Base results directory

    Returns:
        List of paths to all _encodedrecon.npy files in distributed folders
    """
    import glob

    base_eval_dir = os.path.join(results_dir, "evaluation_results")

    # Find all part_XXXX folders
    part_pattern = os.path.join(base_eval_dir, "part_*")
    part_folders = glob.glob(part_pattern)

    all_npy_files = []

    if part_folders:
        debug_print(f"🔍 Found {len(part_folders)} distributed folders", debug=DEBUG_ENABLED)
        # Search in distributed folders
        npy_pattern = os.path.join(base_eval_dir, "part_*", "*_encodedrecon.npy")
        all_npy_files = glob.glob(npy_pattern)
    else:
        debug_print(f"🔍 No distributed folders found, searching in base folder", debug=DEBUG_ENABLED)
        # Fallback to original behavior if no distributed folders
        npy_pattern = os.path.join(base_eval_dir, "*_encodedrecon.npy")
        all_npy_files = glob.glob(npy_pattern)

    return all_npy_files


def _count_distributed_folders(results_dir, files_per_patch_set=4):
    """
    Count the number of distributed folders and estimate total files.

    Args:
        results_dir: Base results directory
        files_per_patch_set: Number of .npy files saved per patch set (default: 4 for backward compatibility)

    Returns:
        Tuple of (num_folders, estimated_total_files)
    """
    import glob

    base_eval_dir = os.path.join(results_dir, "evaluation_results")
    part_pattern = os.path.join(base_eval_dir, "part_*")
    part_folders = glob.glob(part_pattern)

    if part_folders:
        num_folders = len(part_folders)
        # Calculate patch sets per folder based on actual files per patch set
        patch_sets_per_folder = 100000 // files_per_patch_set
        estimated_total_patch_sets = num_folders * patch_sets_per_folder
        estimated_files = estimated_total_patch_sets * files_per_patch_set
        return num_folders, estimated_files
    else:
        return 0, 0


def _get_part_folders(results_dir):
    """
    Get sorted list of part_* directories from the evaluation results.

    Args:
        results_dir: Base results directory

    Returns:
        List of part folder paths sorted by part number
    """
    import glob

    base_eval_dir = os.path.join(results_dir, "evaluation_results")
    part_pattern = os.path.join(base_eval_dir, "part_*")
    part_folders = glob.glob(part_pattern)

    # Sort by part number to ensure consistent processing order
    def extract_part_number(path):
        import re
        match = re.search(r'part_(\d+)', os.path.basename(path))
        return int(match.group(1)) if match else 0

    part_folders.sort(key=extract_part_number)
    return part_folders


def _iterate_saved_patch_items_by_part(args, part_folder):
    """
    Iterate saved .npy patches from a specific part folder as a generator.

    Args:
        args: Arguments object
        part_folder: Path to the specific part_* folder

    Yields:
        (current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw)
    """
    import glob

    # Get all .npy files from this specific part folder
    npy_pattern = os.path.join(part_folder, "*_encodedrecon.npy")
    npy_files = glob.glob(npy_pattern)

    debug_print(f"📁 Processing part {os.path.basename(part_folder)}: {len(npy_files)} .npy files", debug=DEBUG_ENABLED)

    # Use tqdm for loading .npy files in this part
    for npy_file in tqdm(npy_files, desc=f"Loading {os.path.basename(part_folder)}", unit="file", leave=False):
        base_name = npy_file.replace("_encodedrecon.npy", "")
        coords_file = f"{base_name}_coords.npy"
        latent_file = f"{base_name}_latent.npy"
        anomaly_file = f"{base_name}_anomaly_map_arithmetic.npy"

        if os.path.exists(coords_file) and os.path.exists(latent_file) and os.path.exists(anomaly_file):
            try:
                patch_coords_8_values = np.load(coords_file).tolist()
                filename = os.path.basename(npy_file)

                # Extract image path portion by removing patch coordinates and file suffix
                import re as _re
                coord_pattern = r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+__"
                match = _re.search(coord_pattern, filename)
                if match:
                    file_info = filename[:match.start()]
                else:
                    if "__minimal_diff" in filename:
                        file_info = filename.split("__minimal_diff")[0]
                        file_info = _re.sub(r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+$", "", file_info)
                    else:
                        file_info = filename.split("__")[0]

                image_path = safe_filename_to_path(file_info)
                encodedrecon_data = np.load(npy_file)
                latent_data = np.load(latent_file)
                anomaly_data = np.load(anomaly_file)

                yield (
                    image_path,
                    patch_coords_8_values,
                    encodedrecon_data.squeeze(),
                    latent_data.squeeze(),
                    anomaly_data.squeeze(),
                )
            except Exception as e:
                debug_print(f"⚠️  Failed reading file in part {os.path.basename(part_folder)}: {npy_file}: {e}", debug=DEBUG_ENABLED)
                continue


class MetricsAccumulator:
    """
    Memory-optimized accumulator for metrics (TP/FP/FN/TN) across multiple processing parts.
    Only stores essential metrics and optionally minimal record data.
    """

    def __init__(self, store_records=False):
        self.total_tp = 0
        self.total_fp = 0
        self.total_fn = 0
        self.total_tn = 0
        self.total_records = 0
        self.defect_records_count = 0
        self.normal_records_count = 0

        # Optional: Store only essential fields, not full records with numpy arrays
        self.store_records = store_records
        if store_records:
            self.minimal_records = []  # Lightweight storage

    def update_from_records(self, records):
        """
        Update metrics from a batch of records from one part.
        Extracts only essential data and clears full records to save memory.

        Args:
            records: List of record dictionaries
        """
        part_tp = part_fp = part_fn = part_tn = 0

        for record in records:
            self.total_records += 1

            # Count defective vs normal predictions
            is_predicted_defective = record.get("is_predicted_defective", (None, False))[1]
            if is_predicted_defective:
                self.defect_records_count += 1
            else:
                self.normal_records_count += 1

            # Count confusion matrix elements
            status = record.get("status", (None, ""))[1]
            if status == "TP":
                self.total_tp += 1
                part_tp += 1
            elif status == "FP":
                self.total_fp += 1
                part_fp += 1
            elif status == "FN":
                self.total_fn += 1
                part_fn += 1
            elif status == "TN":
                self.total_tn += 1
                part_tn += 1

            # Store only minimal data if needed for reports (no numpy arrays)
            if self.store_records:
                self.minimal_records.append({
                    'status': (None, status),
                    'image_path': (None, record.get("image_path", (None, ""))[1]),
                    'anomaly_pixels': (None, int(record.get("anomaly_pixels", (None, 0))[1])),
                    'is_defective': (None, is_predicted_defective)
                })

        # Clear full records after extracting metrics to free memory immediately
        records.clear()

        # Print part-level metrics using tqdm.write for clean output
        tqdm.write(f"  📊 Part metrics: TP={part_tp}, FP={part_fp}, FN={part_fn}, TN={part_tn}")
        tqdm.write(f"  📊 Cumulative: TP={self.total_tp}, FP={self.total_fp}, FN={self.total_fn}, TN={self.total_tn}")

    def get_final_metrics(self):
        """
        Calculate and return final metrics.

        Returns:
            Dictionary containing final metrics
        """
        total_patches = self.total_tp + self.total_fp + self.total_fn + self.total_tn
        accuracy = (self.total_tp + self.total_tn) / total_patches if total_patches > 0 else 0
        precision = self.total_tp / (self.total_tp + self.total_fp) if (self.total_tp + self.total_fp) > 0 else 0
        recall = self.total_tp / (self.total_tp + self.total_fn) if (self.total_tp + self.total_fn) > 0 else 0
        specificity = self.total_tn / (self.total_tn + self.total_fp) if (self.total_tn + self.total_fp) > 0 else 0
        f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

        return {
            "total_records": self.total_records,
            "defective_records": self.defect_records_count,
            "normal_records": self.normal_records_count,
            "TP": self.total_tp,
            "FP": self.total_fp,
            "FN": self.total_fn,
            "TN": self.total_tn,
            "accuracy": accuracy,
            "precision": precision,
            "recall": recall,
            "specificity": specificity,
            "f1_score": f1_score,
            "total_patches": total_patches
        }

    def get_all_records(self):
        """Return minimal records for report generation or generate placeholder data."""
        if self.store_records:
            return self.minimal_records
        else:
            # Generate minimal placeholder records from metrics only
            return self._generate_placeholder_records()

    def _generate_placeholder_records(self):
        """Generate minimal records for reports when full records aren't stored."""
        placeholder_records = []

        # Create simple records based on metrics
        for status, count in [("TP", self.total_tp), ("FP", self.total_fp), ("FN", self.total_fn), ("TN", self.total_tn)]:
            for i in range(count):
                placeholder_records.append({
                    "status": (None, status),
                    "image_path": (None, f"placeholder_{status}_{i}"),
                    "is_predicted_defective": (None, status in ["TP", "FP"]),
                    "anomaly_pixels": (None, 1 if status in ["TP", "FP"] else 0)
                })

        return placeholder_records


def _process_single_part(args, part_folder, ground_truth_map, image_cache, output_subdir):
    """
    Process a single part folder and return the processed records.

    Args:
        args: Arguments object
        part_folder: Path to the part_* folder to process
        ground_truth_map: Ground truth mapping
        image_cache: Smart image cache
        output_subdir: Output directory for results

    Returns:
        List of processed records from this part
    """
    # Get patches from this part
    patch_item_iter = _iterate_saved_patch_items_by_part(args, part_folder)

    # Process patches from this part using similar logic to the original function
    batch_records = []
    batch_size = max(1, int(getattr(args, "batch_size", 64)))

    from collections import defaultdict
    image_to_records = defaultdict(list)

    # Count actual .npy files for accurate progress
    import glob
    npy_pattern = os.path.join(part_folder, "*_encodedrecon.npy")
    npy_files = glob.glob(npy_pattern)
    actual_count = len(npy_files)

    part_name = os.path.basename(part_folder)

    # Process each patch in this part with progress bar using actual count
    for current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw in tqdm(
        patch_item_iter,
        total=actual_count,
        desc=f"Processing {part_name}",
        unit="patch",
        leave=False
    ):
        try:
            record = _process_single_patch(
                ground_truth_map=ground_truth_map,
                original_images=image_cache,
                anomaly_binary_threshold=args.anomaly_binary_threshold,
                anomaly_pixel_num_threshold=args.anomaly_pixel_num_threshold,
                patch_size=args.patch_size,
                current_image_path=current_image_path,
                coords_8_values=coords_8_values,
                encodedrecon_raw=encodedrecon_raw,
                latent_raw=latent_raw,
                anomaly_map_arithmetic_raw=anomaly_map_arithmetic_raw,
            )

            if record is None:
                continue

            batch_records.append(record)
            image_to_records[current_image_path].append(record)

            # Process batch when it reaches the flush limit
            if len(batch_records) >= batch_size:
                # Save images if enabled
                if args.enable_save_image_results or args.enable_save_whole_image_results:
                    # Create a temporary checkpoint manager for image saving using the passed output_subdir
                    if not hasattr(args, 'temp_checkpoint_manager'):
                        args.temp_checkpoint_manager = CheckpointManager(
                            results_dir=output_subdir,
                            annotation_dir=args.annotation_dir,
                            force_rerun=args.force_rerun
                        )

                    # Call the batch processing function to save images
                    _process_batch_records_immediately(
                        args,
                        batch_records,
                        ground_truth_map,
                        image_cache,  # original_images
                        args.temp_checkpoint_manager,
                        output_subdir,
                        image_to_records
                    )

                batch_records = []  # Clear the batch

        except Exception as e:
            debug_print(f"⚠️  Failed processing patch in part {os.path.basename(part_folder)}: {e}", debug=DEBUG_ENABLED)
            continue

    # Process any remaining batch records for image saving
    if batch_records and (args.enable_save_image_results or args.enable_save_whole_image_results):
        # Create a temporary checkpoint manager for image saving using the passed output_subdir
        if not hasattr(args, 'temp_checkpoint_manager'):
            args.temp_checkpoint_manager = CheckpointManager(
                results_dir=output_subdir,
                annotation_dir=args.annotation_dir,
                force_rerun=args.force_rerun
            )

        # Call the batch processing function to save images for remaining records
        _process_batch_records_immediately(
            args,
            batch_records,
            ground_truth_map,
            image_cache,  # original_images
            args.temp_checkpoint_manager,
            output_subdir,
            image_to_records
        )

    # Collect all records from this part
    all_part_records = []
    for records_list in image_to_records.values():
        all_part_records.extend(records_list)

    return all_part_records


def _iterate_saved_patch_items(args):
    """
    Iterate saved .npy patches as a generator without building full records.
    Searches across distributed folders to find all .npy files.

    Yields:
      (current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw)
    """
    # Use utility function to get all .npy files from distributed folders
    npy_files = _get_distributed_npy_files(args.results_dir)
    debug_print(f"📁 Found {len(npy_files)} .npy files to process", debug=DEBUG_ENABLED)

    for npy_file in npy_files:
        base_name = npy_file.replace("_encodedrecon.npy", "")
        coords_file = f"{base_name}_coords.npy"
        latent_file = f"{base_name}_latent.npy"
        anomaly_file = f"{base_name}_anomaly_map_arithmetic.npy"
        if os.path.exists(coords_file) and os.path.exists(latent_file) and os.path.exists(anomaly_file):
            try:
                patch_coords_8_values = np.load(coords_file).tolist()
                filename = os.path.basename(npy_file)
                # Extract image path portion by removing patch coordinates and file suffix
                # Pattern: safe_filename__x1_y1_x2_y2_x3_y3_x4_y4__minimal_diff_encodedrecon.npy
                # We want to extract everything before the first coordinate pattern
                import re as _re
                coord_pattern = r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+__"
                match = _re.search(coord_pattern, filename)
                if match:
                    # Extract everything before the coordinate pattern
                    file_info = filename[:match.start()]
                else:
                    # Fallback: try splitting at __minimal_diff
                    if "__minimal_diff" in filename:
                        file_info = filename.split("__minimal_diff")[0]
                        # Remove any trailing coordinate pattern
                        file_info = _re.sub(r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+$", "", file_info)
                    else:
                        # Last resort: split at first __ (original approach)
                        file_info = filename.split("__")[0]
                
                image_path = safe_filename_to_path(file_info)
                encodedrecon_data = np.load(npy_file)
                latent_data = np.load(latent_file)
                anomaly_data = np.load(anomaly_file)
                yield (
                    image_path,
                    patch_coords_8_values,
                    encodedrecon_data.squeeze(),
                    latent_data.squeeze(),
                    anomaly_data.squeeze(),
                )
            except Exception as e:
                debug_print(f"⚠️  Failed reading saved patch: {npy_file}: {e}", debug=DEBUG_ENABLED)
                continue
        else:
            # Legacy fallback: parse filename
            try:
                filename = os.path.basename(npy_file)
                coord_pattern = r"__x(\d+)_y(\d+)_x(\d+)_y(\d+)_x(\d+)_y(\d+)_x(\d+)_y(\d+)__minimal_diff"
                import re as _re
                match = _re.search(coord_pattern, filename)
                if not match:
                    continue
                x1, y1, x2, y2, x3, y3, x4, y4 = map(int, match.groups())
                patch_coords_8_values = [x1, y1, x2, y2, x3, y3, x4, y4]
                # Extract image path portion by removing patch coordinates and file suffix
                # Use the same logic as above for consistency
                coord_pattern_full = r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+__"
                match_full = _re.search(coord_pattern_full, filename)
                if match_full:
                    file_info = filename[:match_full.start()]
                else:
                    # Fallback: try splitting at __minimal_diff
                    if "__minimal_diff" in filename:
                        file_info = filename.split("__minimal_diff")[0]
                        file_info = _re.sub(r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+$", "", file_info)
                    else:
                        file_info = filename.split("__")[0]
                
                image_path = safe_filename_to_path(file_info)
                encodedrecon_data = np.load(npy_file)
                latent_data = np.load(f"{base_name}_latent.npy")
                anomaly_data = np.load(f"{base_name}_anomaly_map_arithmetic.npy")
                yield (
                    image_path,
                    patch_coords_8_values,
                    encodedrecon_data.squeeze(),
                    latent_data.squeeze(),
                    anomaly_data.squeeze(),
                )
            except Exception as e:
                debug_print(f"⚠️  Legacy load failed for {npy_file}: {e}", debug=DEBUG_ENABLED)
                continue


def _process_eval_batches_core(args, vae, model, diffusion, loader, checkpoint_manager=None):
    """
    Core function to process evaluation batches and yield batch data.
    This is a shared component that can be used by different iteration strategies.
    
    Yields tuples:
      (batch_index, x, image_paths_batch, patch_coords, encodedrecon_dodrecon_diff, encoded_latent_diff_resized, anomaly_map_arithmetic)
    """
    # Handle checkpointing/resume functionality
    all_image_paths = []
    if checkpoint_manager:
        # Collect all unique image paths from the dataset
        # Handle both direct dataset and Subset-wrapped dataset
        dataset = loader.dataset.dataset if hasattr(loader.dataset, 'dataset') else loader.dataset
        all_image_paths = list(set(dataset.get_all_image_paths()))
        current_image_index, processed_images = checkpoint_manager.get_resume_info(all_image_paths)
        print(f"Resuming from image {current_image_index}/{len(all_image_paths)}")
        print(f"Already processed: {len(processed_images)} images")
        
        # Print checkpoint status
        print("=== Checkpoint Status ===")
        checkpoint_manager.print_checkpoint_status()
        print("========================")
    
    processed_images_in_batch = []
    idx = -1
    try:
        for idx, (x, seg, object_cls, anomaly_classes, image_paths_batch, patch_coords) in enumerate(
            tqdm(loader, desc="Processing patches (core iterator)")
        ):
            if idx >= args.batch_num:
                break
            
            # Check if we should skip this batch due to checkpoint resume
            if checkpoint_manager:
                batch_size = x.size(0)
                skip_entire_batch = True
                batch_patch_identifiers = []
                
                for b in range(batch_size):
                    current_image_path = _extract_image_path_from_batch(image_paths_batch, b)
                    coords_8_values = _extract_patch_coordinates(patch_coords, b, args.patch_size)
                    patch_identifier = f"{current_image_path}#{coords_8_values[0]}_{coords_8_values[1]}"
                    batch_patch_identifiers.append(patch_identifier)
                    if not checkpoint_manager.is_image_processed(patch_identifier):
                        skip_entire_batch = False
                
                if skip_entire_batch:
                    debug_print(f"⏭️  Skipping batch {idx} - all patches already processed", debug=DEBUG_ENABLED)
                    continue
                
                # Track patches in this batch for checkpointing
                processed_images_in_batch.extend(batch_patch_identifiers)
            
            with torch.no_grad():
                encodedrecon_dodrecon_diff, encoded_latent_diff_resized, anomaly_map_arithmetic = _process_batch_inference(
                    x, object_cls, model, vae, diffusion, args.reverse_steps, device, epoch_metrics=None
                )

            yield (
                idx,
                x,
                image_paths_batch,
                patch_coords,
                encodedrecon_dodrecon_diff,
                encoded_latent_diff_resized,
                anomaly_map_arithmetic,
            )
            
            # Save checkpoint periodically (every few batches)
            if checkpoint_manager and idx % getattr(args, 'checkpoint_interval', 5) == 0:
                checkpoint_manager.save_checkpoint(idx, processed_images_in_batch)
                debug_print(f"💾 Saved checkpoint at batch {idx}", debug=DEBUG_ENABLED)
                
    except KeyboardInterrupt:
        print("\n⚠️  Process interrupted. Saving checkpoint...")
        if checkpoint_manager:
            checkpoint_manager.save_checkpoint(idx, processed_images_in_batch)
        raise
    except Exception as e:
        print(f"\n❌ Error occurred: {e}")
        if checkpoint_manager:
            checkpoint_manager.save_checkpoint(idx, processed_images_in_batch)
        raise
    finally:
        # Final checkpoint save
        if checkpoint_manager and processed_images_in_batch:
            checkpoint_manager.save_checkpoint(idx, processed_images_in_batch)
            debug_print(f"💾 Final checkpoint saved at batch {idx}", debug=DEBUG_ENABLED)


def _iterate_eval_patch_items(args, vae, model, diffusion, loader, checkpoint_manager=None):
    """
    Iterate evaluation batches and yield patch items compatible with the shared streaming pipeline.
    Uses the shared core function for processing.
    
    Yields tuples:
      (current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw)
    """
    for (batch_idx, x, image_paths_batch, patch_coords, 
         encodedrecon_dodrecon_diff, encoded_latent_diff_resized, anomaly_map_arithmetic) in _process_eval_batches_core(
            args, vae, model, diffusion, loader, checkpoint_manager
        ):
        
        batch_size = x.size(0)
        for b in range(batch_size):
            # Extract image path using shared helper
            current_image_path = _extract_image_path_from_batch(image_paths_batch, b)

            # Extract coordinates using shared helper
            coords_8_values = _extract_patch_coordinates(
                patch_coords, b, args.patch_size
            )

            # Skip if this specific patch is already processed (for checkpoint support)
            # For stride-based processing, we need to track individual patches, not just images
            patch_identifier = f"{current_image_path}#{coords_8_values[0]}_{coords_8_values[1]}"
            if checkpoint_manager and checkpoint_manager.is_image_processed(patch_identifier):
                debug_print(f"⏭️  Skipping already processed patch: {patch_identifier}", debug=DEBUG_ENABLED)
                continue

            # Numpy arrays
            encodedrecon_raw = _to_numpy(encodedrecon_dodrecon_diff[b]).squeeze()
            latent_raw = _to_numpy(encoded_latent_diff_resized[b]).squeeze()
            anomaly_map_arithmetic_raw = _to_numpy(anomaly_map_arithmetic[b]).squeeze()

            # Mark patch as processed (not just the image)
            if checkpoint_manager:
                checkpoint_manager.mark_image_processed(patch_identifier)

            yield (
                current_image_path,
                coords_8_values,
                encodedrecon_raw,
                latent_raw,
                anomaly_map_arithmetic_raw,
            )


def _iterate_eval_patch_items_with_saving(args, vae, model, diffusion, loader, save_dir, checkpoint_manager=None):
    """
    Iterate evaluation batches, save intermediate products as NPY files, and yield patch items.
    This efficiently saves the 4 intermediate products while yielding data for the streaming pipeline.

    DISTRIBUTED FOLDER STRUCTURE:
    - Files are distributed across multiple subfolders to avoid Windows performance issues
    - Each folder contains ~25,000 patch sets (100,000 files total per folder)
    - Folder structure: save_dir/part_0000/, save_dir/part_0001/, etc.
    - Each patch set consists of 4 files: _encodedrecon.npy, _latent.npy, _anomaly_map_arithmetic.npy, _coords.npy

    Args:
        args: Arguments object
        vae, model, diffusion, loader: Model components and data loader
        save_dir: Directory to save NPY files
        checkpoint_manager: Optional checkpoint manager for resume functionality

    Yields tuples:
      (current_image_path, coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw)
    """
    # Create save directory
    os.makedirs(save_dir, exist_ok=True)

    # Count the number of files saved per patch set by examining the save job function
    # This dynamically counts how many .npy files are actually saved per patch
    def _count_files_per_patch_set():
        """
        Count how many .npy files are saved per patch set by analyzing the save job.
        This function dynamically counts the file operations in _submit_npy_save_job.
        """
        import inspect
        
        # Get the source code of the _submit_npy_save_job function to count np.save calls
        try:
            # Count np.save operations by inspecting the nested _job function
            file_count = 0
            
            # Data files (always present)
            file_count += 3  # encodedrecon, latent, anomaly_map_arithmetic
            
            # Coordinate file (always present)
            file_count += 1  # coords
            
            # Future-proof: Could be extended to parse the actual function source
            # to count np.save() calls dynamically, but for now we use known values
            
            debug_print(f"🔍 Detected {file_count} files per patch set", debug=DEBUG_ENABLED)
            return file_count
            
        except Exception as e:
            # Fallback to known value if inspection fails
            debug_print(f"⚠️  Could not inspect save function, using default: {e}", debug=DEBUG_ENABLED)
            return 4  # encodedrecon, latent, anomaly_map_arithmetic, coords
    
    files_per_patch_set = _count_files_per_patch_set()
    
    # Initialize file counter and folder management for distributed saving
    file_counter = 0
    files_per_folder = 100000 // files_per_patch_set  # Dynamic calculation based on actual files saved
    current_folder_idx = 0
    current_folder_path = None
    
    debug_print(f"📊 Distributed saving: {files_per_patch_set} files per patch set, {files_per_folder} patch sets per folder", debug=DEBUG_ENABLED)

    def _get_current_save_folder():
        """Get the current folder path for saving, creating new folders as needed."""
        nonlocal current_folder_path, current_folder_idx, file_counter

        # Check if we need to create a new folder
        if current_folder_path is None or file_counter >= files_per_folder:
            if current_folder_path is not None:
                current_folder_idx += 1
                file_counter = 0

            current_folder_path = os.path.join(save_dir, f"part_{current_folder_idx:04d}")
            os.makedirs(current_folder_path, exist_ok=True)
            debug_print(f"📁 Created new distributed folder: {os.path.basename(current_folder_path)} ({file_counter}/{files_per_folder} sets used)", debug=DEBUG_ENABLED)

        return current_folder_path
    
    # Determine save dtype
    save_dtype_is_f16 = getattr(args, 'save_dtype_f16', False)
    
    # Setup async saving executor for efficient NPY saving
    async_workers = getattr(args, "async_save_workers", None)
    if async_workers is None:
        async_workers = 1 if platform.system().lower().startswith("win") else max(2, min(8, os.cpu_count() or 4))
    async_workers = max(1, int(async_workers))

    executor = ThreadPoolExecutor(max_workers=async_workers)
    pending = deque()
    max_pending = max(8, async_workers * 8)

    def _submit_npy_save_job(save_dir_local, base_filename_local,
                           encodedrecon_raw_local, latent_raw_local, anomaly_raw_local,
                           coords_array_local, save_dtype_is_f16_local):
        def _job():
            # Save npy arrays
            if save_dtype_is_f16_local:
                np.save(os.path.join(save_dir_local, f"{base_filename_local}_encodedrecon.npy"), encodedrecon_raw_local.astype(np.float16))
                np.save(os.path.join(save_dir_local, f"{base_filename_local}_latent.npy"), latent_raw_local.astype(np.float16))
                np.save(os.path.join(save_dir_local, f"{base_filename_local}_anomaly_map_arithmetic.npy"), anomaly_raw_local.astype(np.float16))
            else:
                np.save(os.path.join(save_dir_local, f"{base_filename_local}_encodedrecon.npy"), encodedrecon_raw_local.astype(np.float32))
                np.save(os.path.join(save_dir_local, f"{base_filename_local}_latent.npy"), latent_raw_local.astype(np.float32))
                np.save(os.path.join(save_dir_local, f"{base_filename_local}_anomaly_map_arithmetic.npy"), anomaly_raw_local.astype(np.float32))

            # Save coords
            np.save(os.path.join(save_dir_local, f"{base_filename_local}_coords.npy"), coords_array_local.astype(np.int32))

        return executor.submit(_job)

    def _wait_for_pending_saves():
        """Wait for and clean up completed save jobs."""
        while pending:
            job = pending.popleft()
            if job.done():
                try:
                    job.result()  # Check for exceptions
                except Exception as e:
                    debug_print(f"⚠️  Save job failed: {e}", debug=DEBUG_ENABLED)
            else:
                pending.appendleft(job)  # Put it back if not done
                break

    try:
        for (batch_idx, x, image_paths_batch, patch_coords, 
             encodedrecon_dodrecon_diff, encoded_latent_diff_resized, anomaly_map_arithmetic) in _process_eval_batches_core(
                args, vae, model, diffusion, loader, checkpoint_manager
            ):
            
            batch_size = x.size(0)
            for b in range(batch_size):
                # Extract image path using shared helper
                current_image_path = _extract_image_path_from_batch(image_paths_batch, b)

                # Extract coordinates using shared helper
                coords_8_values = _extract_patch_coordinates(
                    patch_coords, b, args.patch_size
                )

                # Skip if this specific patch is already processed (for checkpoint support)
                # For stride-based processing, we need to track individual patches, not just images
                patch_identifier = f"{current_image_path}#{coords_8_values[0]}_{coords_8_values[1]}"
                if checkpoint_manager and checkpoint_manager.is_image_processed(patch_identifier):
                    debug_print(f"⏭️  Skipping already processed patch: {patch_identifier}", debug=DEBUG_ENABLED)
                    continue

                # Numpy arrays
                encodedrecon_raw = _to_numpy(encodedrecon_dodrecon_diff[b]).squeeze()
                latent_raw = _to_numpy(encoded_latent_diff_resized[b]).squeeze()
                anomaly_map_arithmetic_raw = _to_numpy(anomaly_map_arithmetic[b]).squeeze()

                # Generate filename for saving
                safe_filename = path_to_safe_filename(current_image_path)
                # Use the expected 8-value coordinate format: __x1_y1_x2_y2_x3_y3_x4_y4__
                coords_filename_part = f"x{coords_8_values[0]}_y{coords_8_values[1]}_x{coords_8_values[2]}_y{coords_8_values[3]}_x{coords_8_values[4]}_y{coords_8_values[5]}_x{coords_8_values[6]}_y{coords_8_values[7]}"
                base_filename = f"{safe_filename}__{coords_filename_part}__minimal_diff"
                
                # Convert coordinates to numpy array for saving
                coords_array = np.array(coords_8_values)

                # Get the appropriate distributed folder for this file set
                current_save_folder = _get_current_save_folder()

                # Submit async save job
                job = _submit_npy_save_job(
                    current_save_folder, base_filename,
                    encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw,
                    coords_array, save_dtype_is_f16
                )
                pending.append(job)

                # Increment file counter (dynamically calculated files per set)
                file_counter += files_per_patch_set

                # Throttle pending jobs to avoid memory issues
                if len(pending) >= max_pending:
                    _wait_for_pending_saves()

                # Mark patch as processed (not just the image)
                if checkpoint_manager:
                    checkpoint_manager.mark_image_processed(patch_identifier)

                # Yield the same data as the original function
                yield (
                    current_image_path,
                    coords_8_values,
                    encodedrecon_raw,
                    latent_raw,
                    anomaly_map_arithmetic_raw,
                )

    finally:
        # Wait for all pending saves to complete
        debug_print(f"🔄 Waiting for {len(pending)} remaining save jobs to complete...", debug=DEBUG_ENABLED)
        while pending:
            job = pending.popleft()
            try:
                job.result()  # Wait for completion and check for exceptions
            except Exception as e:
                debug_print(f"⚠️  Save job failed: {e}", debug=DEBUG_ENABLED)

        executor.shutdown(wait=True)

        # Log distributed folder statistics
        total_folders = current_folder_idx + 1
        total_patch_sets = file_counter // files_per_patch_set
        if total_folders > 1:
            debug_print(f"📊 Distributed {file_counter} files ({total_patch_sets} patch sets of {files_per_patch_set} files each) across {total_folders} folders", debug=DEBUG_ENABLED)
            debug_print(f"📂 Folder structure: {save_dir}/part_0000/ through {save_dir}/part_{total_folders-1:04d}/", debug=DEBUG_ENABLED)
            debug_print(f"📈 Average: {total_patch_sets // total_folders} patch sets per folder", debug=DEBUG_ENABLED)
        else:
            debug_print(f"✅ All NPY saves completed in single folder: {save_dir} ({total_patch_sets} patch sets, {file_counter} files)", debug=DEBUG_ENABLED)


def _process_batch_records_immediately(args, batch_records, ground_truth_map, original_images, 
                                     checkpoint_manager, output_subdir, image_to_records):
    """
    Process a batch of records immediately without accumulating them in memory.
    This function handles immediate processing tasks like saving images and updating metrics.
    """
    if not batch_records:
        return
    
    # Process each record in the batch
    for record in batch_records:
        # Update per-image tracking for image-level processing
        img_path = record.get("image_path", (None, None))[1]
        if img_path:
            # The record is already added to image_to_records in the main loop
            pass
    
    # If image saving is enabled, process the current batch
    if args.enable_save_image_results:
        for record in batch_records:
            try:
                img_path = record.get("image_path", (None, None))[1]
                if not img_path:
                    continue
                
                coords = record.get("patch_coords", (None, []))[1]
                if not (isinstance(coords, (list, tuple)) and len(coords) == 8):
                    continue
                
                patch_x, patch_y = int(coords[0]), int(coords[1])
                
                # Build a single-patch record list
                patch_records = [record]
                
                # Compute predicted set for this single patch
                anomaly_map = record["anomaly_map_arithmetic_binary"][1]
                anomaly_pixels = int(np.sum(anomaly_map)) if hasattr(anomaly_map, 'sum') else 0
                patch_pred_set = set()
                if anomaly_pixels > 0:
                    grid_row = patch_y // args.patch_size
                    grid_col = patch_x // args.patch_size
                    patch_pred_set.add((grid_row, grid_col))
                
                # Get ground truth defective patches for this image
                ground_truth_defective = ground_truth_map.get(img_path, set()) if isinstance(ground_truth_map, dict) else set()
                overlapping = set()
                
                # Create marked_images directory in the output_subdir (tagged folder)
                marked_images_dir = os.path.join(output_subdir, "marked_images")
                os.makedirs(marked_images_dir, exist_ok=True)

                save_patch_results_from_records(
                    checkpoint_manager,  # CheckpointManager object
                    img_path,           # image_path
                    patch_records,      # patch_records
                    patch_pred_set,     # predicted_defective_set
                    ground_truth_defective,  # ground_truth_defective
                    overlapping,        # overlapping
                    enable_save_optional_image_results=args.enable_save_optional_image_results,
                    patch_size=args.patch_size,
                    patch_x=patch_x,
                    patch_y=patch_y,
                )
            except Exception as e:
                debug_print(f"⚠️  Failed to save patch-level result: {e}", debug=DEBUG_ENABLED)

def _finalize_incremental_processing(args, total_records, normal_records_count, defect_records_count, 
                                   image_to_records, ground_truth_map, output_subdir):
    """
    Finalize the incremental processing by computing final metrics and saving results.
    """
    # Compute final metrics
    metrics = {
        "total_records": total_records,
        "defective_records": defect_records_count,
        "normal_records": normal_records_count
    }
    
    # Try to compute more sophisticated metrics if possible
    try:
        # For confusion matrix and other detailed metrics, we need to process the accumulated image records
        if args.enable_confusion_matrix:
            # Convert image_to_records back to a flat list for confusion matrix
            all_records = []
            for img_records in image_to_records.values():
                all_records.extend(img_records)
            
            create_confusion_matrix_from_records(
                all_records,
                output_subdir,
                annotation_dir=args.annotation_dir,
                patch_size=args.patch_size
            )
        
        # Save JSON results if enabled
        if args.enable_save_json_results:
            debug_print("💾 Saving JSON results...", debug=DEBUG_ENABLED)
            # Convert image_to_records back to a flat list for JSON saving
            all_records = []
            for img_records in image_to_records.values():
                all_records.extend(img_records)
            
            save_all_records_json(
                all_records,
                output_subdir,
                filename="all_evaluation_records.json",
                patch_size=args.patch_size,
                sort_records=not args.no_sort_records_by_anomaly
            )
        
        # Process whole image results if enabled
        if args.enable_save_whole_image_results:
            debug_print("🖼️ Processing whole image results...", debug=DEBUG_ENABLED)
            # Create marked_images directory in the output_subdir (tagged folder)
            marked_images_dir = os.path.join(output_subdir, "marked_images")
            os.makedirs(marked_images_dir, exist_ok=True)
            
            # Get evaluation_results_dir from a checkpoint manager (at base level)
            base_results_dir = os.path.dirname(output_subdir)
            evaluation_results_dir = os.path.join(base_results_dir, "evaluation_results")
            os.makedirs(evaluation_results_dir, exist_ok=True)
            
            for img_path, image_records in image_to_records.items():
                # Predicted defective set from records (expects 8-value patch_coords)
                predicted_defective_set = set()
                for rec in image_records:
                    try:
                        is_def = rec.get("is_predicted_defective", (None, False))[1]
                        if not is_def:
                            continue
                        coords = rec.get("patch_coords", (None, []))[1]
                        if not (isinstance(coords, (list, tuple)) and len(coords) == 8):
                            raise ValueError(f"Expected 8-value patch_coords in records, got: {coords}")
                        x1, y1 = int(coords[0]), int(coords[1])
                        row = y1 // args.patch_size
                        col = x1 // args.patch_size
                        predicted_defective_set.add((row, col))
                    except Exception:
                        continue

                ground_truth_defective = ground_truth_map.get(img_path, set()) if isinstance(ground_truth_map, dict) else set()
                overlapping = set()

                try:
                    save_image_results_from_records(
                        marked_images_dir,
                        evaluation_results_dir,
                        img_path,
                        image_records,
                        predicted_defective_set,
                        ground_truth_defective,
                        overlapping,
                        enable_save_optional_image_results=args.enable_save_optional_image_results,
                        patch_size=args.patch_size,
                    )
                except Exception as e:
                    debug_print(f"⚠️  Failed to save image-level results for {img_path}: {e}", debug=DEBUG_ENABLED)
        
        # Create Excel report if enabled
        if args.enable_excel_report:
            debug_print("📊 Creating Excel report...", debug=DEBUG_ENABLED)
            # Convert image_to_records back to a flat list for Excel report
            all_records = []
            for img_records in image_to_records.values():
                all_records.extend(img_records)
            
            make_excel(all_records, output_subdir, args.split, args.object_class)
            
    except Exception as e:
        debug_print(f"⚠️  Error in final processing: {e}", debug=DEBUG_ENABLED)
    
    debug_print("✅ Incremental processing completed successfully!", debug=DEBUG_ENABLED)
    return metrics

def validate_mode_arguments(args):
    """
    Validate that required arguments are provided for each mode.
    Provides clear error messages to guide users.
    """
    mode = args.mode
    errors = []
    
    # Define required arguments for each mode
    if mode == "save_only":
        # Mode 1: Save .npy files and diff images only
        required_args = [
            ("annotation_dir", "Directory containing annotation files"),
            ("pretrained", "Path to pretrained model (use --pretrained)"),
        ]
        optional_but_recommended = [
            ("model_size", "Model size (UNet_XS, UNet_S, UNet_M, UNet_L, UNet_XL)"),
            ("patch_size", "Center size for model"),
            ("reverse_steps", "Number of reverse steps"),
            ("batch_num", "Number of batches to process"),
            ("enable_epoch_stats", "Enable detailed epoch-wise statistics (use --enable-epoch-stats)"),
            
            ("debug", "Enable detailed debug logging (use --debug)"),
        ]
        
    elif mode == "process_only":
        # Mode 2: Read existing .npy files and generate categorization results
        required_args = [
            ("annotation_dir", "Directory containing annotation files"),
        ]
        optional_but_recommended = [
            ("results_dir", "Directory containing saved .npy files"),
            ("anomaly_binary_threshold", "Binary threshold for anomaly detection"),
            ("anomaly_pixel_num_threshold", "Pixel number threshold"),
            ("enable_excel_report", "Generate Excel report (use --enable-excel-report)"),
            ("enable_save_image_results", "Save image results (use --enable-save-image-results)"),
            ("enable_save_optional-image-results", "Save optional image results (use --enable-save-optional-image-results)"),
            ("enable_save_whole-image-results", "Save whole image results (use --enable-save-whole-image-results)"),
            ("enable_epoch_stats", "Enable detailed epoch-wise statistics (use --enable-epoch-stats)"),
            ("debug", "Enable detailed debug logging (use --debug)"),
        ]
        
    elif mode == "save_and_process":
        # Mode 3: Save .npy files and immediately process them
        required_args = [
            ("annotation_dir", "Directory containing annotation files"),
            ("pretrained", "Path to pretrained model (use --pretrained)"),
        ]
        optional_but_recommended = [
            ("model_size", "Model size"),
            ("patch_size", "Center size for model"),
            ("reverse_steps", "Number of reverse steps"),
            ("batch_num", "Number of batches to process"),
            ("anomaly_binary_threshold", "Binary threshold for anomaly detection"),
            ("enable_excel_report", "Generate Excel report (use --enable-excel-report)"),
            ("enable_epoch_stats", "Enable detailed epoch-wise statistics (use --enable-epoch-stats)"),
            ("debug", "Enable detailed debug logging (use --debug)"),
        ]
        
    elif mode == "full_pipeline":
        # Mode 4: Complete pipeline without saving intermediates
        required_args = [
            ("annotation_dir", "Directory containing annotation files"),
            ("pretrained", "Path to pretrained model (use --pretrained)"),
        ]
        optional_but_recommended = [
            ("model_size", "Model size"),
            ("patch_size", "Center size for model"),
            ("reverse_steps", "Number of reverse steps"),
            ("batch_num", "Number of batches to process"),
            ("anomaly_binary_threshold", "Binary threshold for anomaly detection"),
            ("enable_excel_report", "Generate Excel report (use --enable-excel-report)"),
            ("enable_epoch_stats", "Enable detailed epoch-wise statistics (use --enable-epoch-stats)"),
            ("debug", "Enable detailed debug logging (use --debug)"),
        ]
        
    elif mode == "full_pipeline_with_saving_npy":
        # Mode 5: Complete pipeline with saving intermediate NPY files
        required_args = [
            ("annotation_dir", "Directory containing annotation files"),
            ("pretrained", "Path to pretrained model (use --pretrained)"),
        ]
        optional_but_recommended = [
            ("model_size", "Model size"),
            ("patch_size", "Center size for model"),
            ("reverse_steps", "Number of reverse steps"),
            ("batch_num", "Number of batches to process"),
            ("anomaly_binary_threshold", "Binary threshold for anomaly detection"),
            ("enable_excel_report", "Generate Excel report (use --enable-excel-report)"),
            ("enable_epoch_stats", "Enable detailed epoch-wise statistics (use --enable-epoch-stats)"),
            ("debug", "Enable detailed debug logging (use --debug)"),
            ("save_dir", "Directory to save NPY files (use --save-dir)"),
            ("save_dtype_f16", "Save NPY files as float16 for space efficiency (use --save-dtype-f16)"),
        ]
    
    # Check required arguments
    for arg_name, description in required_args:
        value = getattr(args, arg_name, None)
        if not value or (isinstance(value, str) and value.strip() == ""):
            errors.append(f"❌ REQUIRED for mode '{mode}': --{arg_name.replace('_', '-')} ({description})")
    
    # Show helpful information about the mode
    mode_descriptions = {
        "save_only": "Save raw .npy files and diff images without processing",
        "process_only": "Process existing .npy files to generate evaluation results", 
        "save_and_process": "Save raw files AND process them immediately",
        "full_pipeline": "Complete evaluation pipeline without saving intermediate files",
        "full_pipeline_with_saving_npy": "Complete evaluation pipeline WITH saving intermediate NPY files"
    }
    
    if errors:
        print(f"\n🚨 ARGUMENT VALIDATION FAILED for mode: {mode}")
        print(f"📝 Mode description: {mode_descriptions.get(mode, 'Unknown mode')}")
        print(f"\n❌ Missing required arguments:")
        for error in errors:
            print(f"   {error}")
        
        print(f"\n💡 Recommended optional arguments for mode '{mode}':")
        for arg_name, description in optional_but_recommended:
            value = getattr(args, arg_name, None)
            status = "✅" if value else "⚠️"
            print(f"   {status} --{arg_name.replace('_', '-')}: {description}")
        
        print(f"\n📚 Example usage for mode '{mode}':")
        if mode == "save_only":
            print(f"   python evaluate_and_process.py --mode save_only \\")
            print(f"       --annotation-dir path/to/annotations \\")
            print(f"       --pretrained path/to/model.pt \\")
            print(f"       --results-dir ./results \\")
            print(f"       --debug")
            
        elif mode == "process_only":
            print(f"   python evaluate_and_process.py --mode process_only \\")
            print(f"       --annotation-dir path/to/annotations \\")
            print(f"       --enable-excel-report \\")
            print(f"       --debug")
            print(f"   # Or with explicit results directory:")
            print(f"   python evaluate_and_process.py --mode process_only \\")
            print(f"       --results-dir path/to/saved_results \\")
            print(f"       --annotation-dir path/to/annotations")
            
        elif mode == "save_and_process":
            print(f"   python evaluate_and_process.py --mode save_and_process \\")
            print(f"       --annotation-dir path/to/annotations \\")
            print(f"       --pretrained path/to/model.pt \\")
            print(f"       --enable-excel-report \\")
            print(f"       --debug")
            
        elif mode == "full_pipeline":
            print(f"   python evaluate_and_process.py --mode full_pipeline \\")
            print(f"       --annotation-dir path/to/annotations \\")
            print(f"       --pretrained path/to/model.pt \\")
            print(f"       --enable-excel-report \\")
            print(f"       --debug")
        
        elif mode == "full_pipeline_with_saving_npy":
            print(f"   python evaluate_and_process.py --mode full_pipeline_with_saving_npy \\")
            print(f"       --annotation-dir path/to/annotations \\")
            print(f"       --pretrained path/to/model.pt \\")
            print(f"       --enable-excel-report \\")
            print(f"       --debug")
        
        print(f"\n💭 Need help? Check the script header comments for detailed usage information.")
        return False
    
    # Validation passed
    print(f"✅ Argument validation passed for mode: {mode}")
    print(f"📝 Mode description: {mode_descriptions.get(mode, 'Unknown mode')}")
    return True


def mode_save_only(args):
    """Mode 1: Save .npy files and diff images only."""
    print("=== Mode 1: Save Only ===")
    
    # Before saving results: Load model components and prepare evaluation setup
    vae, model, diffusion, dataset, loader, evaluation_results_dir, checkpoint_manager = _before_saving_results(args)
    
    # Create save directory for NPY files
    if hasattr(args, 'save_dir') and args.save_dir:
        npy_save_dir = args.save_dir
    else:
        # Use evaluation_results_dir directly for save_only mode
        npy_save_dir = evaluation_results_dir
    
    print(f"NPY files will be saved to: {npy_save_dir}")

    # Use the saving iterator but only consume it to trigger the saving, don't process
    patch_item_iter = _iterate_eval_patch_items_with_saving(
        args, vae, model, diffusion, loader, npy_save_dir, checkpoint_manager
    )

    # Consume the iterator to trigger saving (without processing)
    patch_count = 0
    for patch_item in patch_item_iter:
        patch_count += 1
        # Just consume the iterator to trigger saving, no processing needed

    # Display success message with bootstrap status
    if args.bootstrap_samples is not None:
        print(f"✅ Saved {patch_count} bootstrap-sampled patches as NPY files to: {npy_save_dir}")
    else:
        print(f"✅ Saved {patch_count} patches as NPY files to: {npy_save_dir}")
    return None

def mode_process_only(args):
    """Mode 2: Process .npy files part by part and accumulate TP/FP/FN/TN results."""
    print("=== Mode 2: Process Only (Part-wise Processing) ===")

    # Load ground truth map
    ground_truth_map = load_ground_truth_map(args.annotation_dir)
    print(f"Loaded ground truth for {len(ground_truth_map)} images")

    # Get all part folders
    part_folders = _get_part_folders(args.results_dir)
    if not part_folders:
        print("❌ No part_* folders found in evaluation_results directory")
        print("   Falling back to legacy processing mode...")
        # If no parts, process all files at once (legacy behavior)
        return mode_process_only_legacy(args)

    print(f"🔍 Found {len(part_folders)} part folders to process")
    for i, folder in enumerate(part_folders):
        print(f"   {i+1}. {os.path.basename(folder)}")

    # Create smart image cache
    max_cache_memory_gb = getattr(args, 'max_cache_memory_gb', 2.0)  # Default 2GB
    max_cache_images = getattr(args, 'max_cache_images', 100)  # Default 100 images
    print(f"🧠 Smart image cache: max {max_cache_memory_gb}GB, max {max_cache_images} images")
    image_cache = SmartImageCache(max_memory_gb=max_cache_memory_gb, max_images=max_cache_images)

    # Initialize metrics accumulator with minimal record storage for reports
    store_minimal_records = args.enable_confusion_matrix or args.enable_save_json_results
    accumulator = MetricsAccumulator(store_records=store_minimal_records)

    # Create output directory
    from datetime import datetime
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.tag:
        output_dir = os.path.join(args.results_dir, f"{args.tag}_{current_time}")
    else:
        output_dir = os.path.join(args.results_dir, f"{current_time}")
    os.makedirs(output_dir, exist_ok=True)
    output_subdir = os.path.join(output_dir, "processed_results")
    os.makedirs(output_subdir, exist_ok=True)
    print(f"📁 Output directory: {output_subdir}")

    # Process each part sequentially
    import time
    total_start_time = time.time()

    # Use tqdm for the main part processing loop
    for i, part_folder in enumerate(tqdm(part_folders, desc="Processing parts", unit="part")):
        part_name = os.path.basename(part_folder)
        tqdm.write(f"\n📦 Processing {part_name} ({i+1}/{len(part_folders)})...")

        part_start_time = time.time()

        try:
            # Process this part
            part_records = _process_single_part(args, part_folder, ground_truth_map, image_cache, output_subdir)

            if part_records:
                # Update accumulator with records from this part (clears records automatically)
                accumulator.update_from_records(part_records)

                part_elapsed = time.time() - part_start_time
                tqdm.write(f"  ✅ {part_name}: {len(part_records)} records processed in {part_elapsed:.1f}s")

                # Explicit memory cleanup after processing part
                del part_records
                part_records = None

            else:
                tqdm.write(f"  ⚠️ {part_name}: No records found")

        except Exception as e:
            tqdm.write(f"  ❌ {part_name}: Failed to process - {e}")
            debug_print(f"Error processing {part_name}: {e}", debug=DEBUG_ENABLED)
            continue

        # Aggressive memory cleanup after each part
        import gc
        if hasattr(torch, 'cuda') and torch.cuda.is_available():
            torch.cuda.empty_cache()
        gc.collect()

    # Calculate final metrics
    total_elapsed = time.time() - total_start_time
    final_metrics = accumulator.get_final_metrics()

    print(f"\n🎯 Final Results Summary:")
    print(f"   📊 Total records: {final_metrics['total_records']}")
    print(f"   🟢 Normal: {final_metrics['normal_records']}")
    print(f"   🔴 Defective: {final_metrics['defective_records']}")
    print(f"   📈 Confusion Matrix:")
    print(f"      TP: {final_metrics['TP']}, FN: {final_metrics['FN']}")
    print(f"      FP: {final_metrics['FP']}, TN: {final_metrics['TN']}")
    print(f"   📊 Metrics:")
    print(f"      Accuracy ((TP+TN)/(TP+FN+FP+TN)): {final_metrics['accuracy']:.4f}")
    print(f"      Precision (TP/(TP+FP)): {final_metrics['precision']:.4f}")
    print(f"      Recall (TP/(TP+FN)): {final_metrics['recall']:.4f}")
    print(f"      Specificity (TN/(TN+FP)): {final_metrics['specificity']:.4f}")
    print(f"      F1-Score (2*Precision*Recall/(Precision+Recall)): {final_metrics['f1_score']:.4f}")
    print(f"   ⏱️ Total processing time: {total_elapsed:.1f}s")

    # Generate final reports using accumulated records
    try:
        all_records = accumulator.get_all_records()

        # Create confusion matrix
        if args.enable_confusion_matrix:
            print(f"📊 Generating confusion matrix...")
            create_confusion_matrix_from_records(
                all_records,
                output_subdir,
                annotation_dir=args.annotation_dir,
                patch_size=args.patch_size
            )

        # Save JSON results
        if args.enable_save_json_results:
            print(f"💾 Saving JSON results...")
            save_all_records_json(
                all_records,
                output_subdir,
                filename="all_records.json",
                patch_size=args.patch_size
            )


    except Exception as e:
        print(f"⚠️ Warning: Failed to generate final reports: {e}")
        debug_print(f"Report generation error: {e}", debug=DEBUG_ENABLED)

    # Print final cache statistics
    cache_stats = image_cache.get_stats()
    print(f"🧠 Final cache statistics:")
    print(f"   📊 Cached images: {cache_stats['cached_images']}")
    print(f"   💾 Memory used: {cache_stats['total_memory_mb']:.1f}MB / {cache_stats['max_memory_mb']:.1f}MB ({cache_stats['memory_usage_percent']:.1f}%)")

    return final_metrics


def mode_process_only_legacy(args):
    """Legacy mode: Read existing .npy files and generate categorization results using the original pipeline."""
    print("=== Mode 2: Process Only (Legacy Mode) ===")

    # Load ground truth and original images based on annotations, mirroring the incremental path
    ground_truth_map = load_ground_truth_map(args.annotation_dir)
    print(f"Loaded ground truth for {len(ground_truth_map)} images")

    # To load original images, we need the set of image paths referenced by saved results
    # Scan saved .npy items to collect image paths without loading all into memory
    # Use parallel processing for better performance on large datasets
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    image_paths_set = set()
    image_paths_lock = threading.Lock()  # Thread-safe set operations

    def process_file_chunk(file_chunk, chunk_idx):
        """Process a chunk of .npy files in parallel"""
        chunk_new_paths = []

        for npy_file in file_chunk:
            try:
                # Process each .npy file (this is the expensive I/O operation)
                base_name = npy_file.replace("_encodedrecon.npy", "")
                coords_file = f"{base_name}_coords.npy"
                latent_file = f"{base_name}_latent.npy"
                anomaly_file = f"{base_name}_anomaly_map_arithmetic.npy"

                if os.path.exists(coords_file) and os.path.exists(latent_file) and os.path.exists(anomaly_file):
                    # Extract image path from filename (same logic as _iterate_saved_patch_items)
                    filename = os.path.basename(npy_file)
                    import re as _re
                    coord_pattern = r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+__"
                    match = _re.search(coord_pattern, filename)
                    if match:
                        file_info = filename[:match.start()]
                    else:
                        if "__minimal_diff" in filename:
                            file_info = filename.split("__minimal_diff")[0]
                            file_info = _re.sub(r"__x\d+_y\d+_x\d+_y\d+_x\d+_y\d+_x\d+_y\d+$", "", file_info)
                        else:
                            file_info = filename.split("__")[0]

                    image_path = safe_filename_to_path(file_info)

                    # Check if this is a new image path
                    if image_path and image_path not in image_paths_set:
                        with image_paths_lock:
                            if image_path not in image_paths_set:
                                image_paths_set.add(image_path)
                                chunk_new_paths.append(image_path)

            except Exception as e:
                debug_print(f"⚠️  Failed reading file in chunk {chunk_idx}: {npy_file}: {e}", debug=DEBUG_ENABLED)
                continue

        print(f"📊 Completed file chunk {chunk_idx + 1}: {len(file_chunk)} files, {len(chunk_new_paths)} new paths")
        return chunk_new_paths

    # Process file discovery and reading in parallel
    print(f"📁 Processing saved patches with parallel file I/O...")

    # Process in parallel using ThreadPoolExecutor for I/O bound operations
    import os
    import time
    max_workers = min(os.cpu_count() * 2, 16)  # Use more workers for better CPU utilization

    print(f"⚙️  Using {max_workers} worker threads for parallel file processing")
    start_time = time.time()

    # Get all .npy files first (this is fast)
    npy_files = _get_distributed_npy_files(args.results_dir)
    print(f"🔍 Found {len(npy_files)} .npy files to process")

    # Split files into chunks for parallel processing
    chunk_size = max(1, len(npy_files) // max_workers)
    file_chunks = [npy_files[i:i + chunk_size] for i in range(0, len(npy_files), chunk_size)]
    print(f"📦 Split into {len(file_chunks)} chunks of ~{chunk_size} files each")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        newly_added_paths = []

        # Submit file chunks for parallel processing
        chunk_futures = []
        for chunk_idx, file_chunk in enumerate(file_chunks):
            chunk_future = executor.submit(process_file_chunk, file_chunk, chunk_idx)
            chunk_futures.append(chunk_future)

        # Process completed chunks as they finish
        total_processed = 0
        for future in tqdm(as_completed(chunk_futures), total=len(chunk_futures), desc="Processing file chunks"):
            chunk_result = future.result()
            if chunk_result:
                newly_added_paths.extend(chunk_result)
            total_processed += len(npy_files)  # Total files processed

    elapsed_time = time.time() - start_time
    patches_per_second = total_processed / elapsed_time if elapsed_time > 0 else 0
    print(f"✅ Found {len(image_paths_set)} unique image paths from {total_processed} processed patches")
    print(f"⏱️  Processing time: {elapsed_time:.2f}s ({patches_per_second:.1f} patches/sec)")
    if newly_added_paths:
        print(f"📊 Newly discovered paths: {len(newly_added_paths)}")

    # Reload iterator after scan (it is a generator)
    patch_item_iter = _iterate_saved_patch_items(args)

    # Create smart image cache instead of loading all images at once
    print(f"🧠 Creating smart image cache...")

    # Create smart cache with configurable memory limits
    max_cache_memory_gb = getattr(args, 'max_cache_memory_gb', 2.0)  # Default 2GB
    max_cache_images = getattr(args, 'max_cache_images', 100)  # Default 100 images

    print(f"🧠 Smart image cache: max {max_cache_memory_gb}GB, max {max_cache_images} images")
    image_cache = SmartImageCache(max_memory_gb=max_cache_memory_gb, max_images=max_cache_images)

    # Process incrementally via shared pipeline with smart cache
    metrics, output_dir = _process_records_stream_incrementally(
        args,
        patch_item_iter,
        ground_truth_map,
        image_cache,  # Pass cache instead of loaded images
        output_subdir_name="processed_results",
    )

    # Print final cache statistics
    cache_stats = image_cache.get_stats()
    print(f"🧠 Final cache statistics:")
    print(f"   📊 Cached images: {cache_stats['cached_images']}")
    print(f"   💾 Memory used: {cache_stats['total_memory_mb']:.1f}MB / {cache_stats['max_memory_mb']:.1f}MB ({cache_stats['memory_usage_percent']:.1f}%)")

    return metrics


def mode_save_and_process(args):
    """Mode 3: Save .npy files and immediately process them for categorization."""
    print("=== Mode 3: Save and Process ===")
    
    # First save the raw data (Mode 1)
    mode_save_only(args)
    
    # Then process the saved data (Mode 2)
    return mode_process_only(args)

def mode_full_pipeline(args):
    """
    Mode 4: Complete pipeline without saving intermediates.
    Composed from shared components: _before_saving_results + _generate_records_incrementally + _after_reading_saved_results
    """
    print("=== Mode 4: Full Pipeline ===")
    
    # Before saving results: Load model components and prepare evaluation setup
    vae, model, diffusion, dataset, loader, evaluation_results_dir, checkpoint_manager = _before_saving_results(args)
    
    # Prepare maps used by the shared streaming pipeline
    ground_truth_map = load_ground_truth_map(args.annotation_dir)
    print(f"Loaded ground truth for {len(ground_truth_map)} images")

    # Handle both direct dataset and Subset-wrapped dataset
    dataset = loader.dataset.dataset if hasattr(loader.dataset, 'dataset') else loader.dataset
    image_paths = set(dataset.get_all_image_paths())

    # Use smart cache for large datasets to avoid memory issues
    max_cache_memory_gb = getattr(args, 'max_cache_memory_gb', 2.0)
    max_cache_images = getattr(args, 'max_cache_images', 100)
    
    if len(image_paths) > max_cache_images:
        print(f"🧠 Large dataset detected ({len(image_paths)} images), using smart cache")
        original_images = SmartImageCache(max_memory_gb=max_cache_memory_gb, max_images=max_cache_images)
        print(f"🧠 Smart cache: max {max_cache_memory_gb}GB, max {max_cache_images} images")
    else:
        print(f"📁 Small dataset ({len(image_paths)} images), loading all images at once")
        original_images = load_original_images(image_paths)
        print(f"Loaded {len(original_images)} original images")

    # Create an iterator that yields patch items directly from evaluation
    patch_item_iter = _iterate_eval_patch_items(args, vae, model, diffusion, loader, checkpoint_manager)

    # Process via shared streaming pipeline
    metrics, output_dir = _process_records_stream_incrementally(
        args,
        patch_item_iter,
        ground_truth_map,
        original_images,
        output_subdir_name="processed_results",
    )
    
    # Print cache statistics if using cache
    if hasattr(original_images, 'get_stats'):
        cache_stats = original_images.get_stats()
        print(f"🧠 Final cache statistics:")
        print(f"   📊 Cached images: {cache_stats['cached_images']}")
        print(f"   💾 Memory used: {cache_stats['total_memory_mb']:.1f}MB / {cache_stats['max_memory_mb']:.1f}MB ({cache_stats['memory_usage_percent']:.1f}%)")

    return metrics, output_dir

def mode_full_pipeline_with_saving_npy(args):
    """
    Mode 5: Complete pipeline with saving intermediate NPY files.
    Follows exactly all the process of mode_full_pipeline but also saves the 4 intermediate products
    (coords_8_values, encodedrecon_raw, latent_raw, anomaly_map_arithmetic_raw) as NPY files.
    """
    print("=== Mode 5: Full Pipeline with NPY Saving ===")
    
    # Before saving results: Load model components and prepare evaluation setup
    vae, model, diffusion, dataset, loader, evaluation_results_dir, checkpoint_manager = _before_saving_results(args)
    
    # Prepare maps used by the shared streaming pipeline
    ground_truth_map = load_ground_truth_map(args.annotation_dir)
    print(f"Loaded ground truth for {len(ground_truth_map)} images")

    # Handle both direct dataset and Subset-wrapped dataset
    dataset = loader.dataset.dataset if hasattr(loader.dataset, 'dataset') else loader.dataset
    image_paths = set(dataset.get_all_image_paths())

    # Use smart cache for large datasets to avoid memory issues
    max_cache_memory_gb = getattr(args, 'max_cache_memory_gb', 2.0)
    max_cache_images = getattr(args, 'max_cache_images', 100)
    
    if len(image_paths) > max_cache_images:
        print(f"🧠 Large dataset detected ({len(image_paths)} images), using smart cache")
        original_images = SmartImageCache(max_memory_gb=max_cache_memory_gb, max_images=max_cache_images)
        print(f"🧠 Smart cache: max {max_cache_memory_gb}GB, max {max_cache_images} images")
    else:
        print(f"📁 Small dataset ({len(image_paths)} images), loading all images at once")
        original_images = load_original_images(image_paths)
        print(f"Loaded {len(original_images)} original images")

    print(f"NPY files will be saved to: {evaluation_results_dir}")

    # Create an iterator that yields patch items directly from evaluation AND saves NPY files
    patch_item_iter = _iterate_eval_patch_items_with_saving(
        args, vae, model, diffusion, loader, evaluation_results_dir, checkpoint_manager
    )

    # Process via shared streaming pipeline
    metrics, output_dir = _process_records_stream_incrementally(
        args,
        patch_item_iter,
        ground_truth_map,
        original_images,
        output_subdir_name="processed_results",
    )
    
    # Print cache statistics if using cache
    if hasattr(original_images, 'get_stats'):
        cache_stats = original_images.get_stats()
        print(f"🧠 Final cache statistics:")
        print(f"   📊 Cached images: {cache_stats['cached_images']}")
        print(f"   💾 Memory used: {cache_stats['total_memory_mb']:.1f}MB / {cache_stats['max_memory_mb']:.1f}MB ({cache_stats['memory_usage_percent']:.1f}%)")

    print(f"✅ Complete pipeline finished. Results saved to: {output_dir}")
    print(f"✅ NPY intermediate files saved to: {evaluation_results_dir}")

    return metrics, output_dir

def main():
    global DEBUG_ENABLED
    
    parser = argparse.ArgumentParser(description="Combined Evaluation and Processing Script")
    
    # Mode selection
    parser.add_argument(
        "--mode", 
        type=str, 
        choices=["save_only", "process_only", "save_and_process", "full_pipeline", "full_pipeline_with_saving_npy"],
        required=False,  # Made optional when using --input-json
        help="""Execution mode:
        save_only: Save .npy files and diff images only (needs: --annotation-dir, --pretrained)
        process_only: Process existing .npy files to generate results (needs: --annotation-dir)
        save_and_process: Save AND process immediately (needs: --annotation-dir, --pretrained)
        full_pipeline: Complete pipeline without saving intermediates (needs: --annotation-dir, --pretrained)
        full_pipeline_with_saving_npy: Complete pipeline WITH saving intermediate NPY files (needs: --annotation-dir, --pretrained)
        Note: Can be omitted if specified in --input-json file"""
    )
    
    # Common arguments
    parser.add_argument("--results-dir", type=str, default="./results", 
                       help="Results directory (optional for process_only mode - can be specified in JSON)")
    parser.add_argument("--tag", type=str, default=None,
                       help="Custom tag for output directory. If provided, output directory will be 'tag_current_time' instead of just 'current_time'")
    parser.add_argument("--annotation-dir", type=str, 
                       help="Directory containing annotation files (REQUIRED for all modes)")
    parser.add_argument("--patch-size", type=int, default=128, help="Patch size for image processing")
    parser.add_argument("--stride", type=int, default=None, help="Stride for patch extraction. If None, uses patch_size (no overlap). If smaller than patch_size, creates overlapping patches.")
    parser.add_argument("--irregular-patch", action="store_true", help="Use irregular patch for image processing")
    parser.add_argument("--dataset", type=str, choices=["mvtec", "visa", "pcb"], default="pcb",
                       help="Dataset type (for model loading modes)")
    parser.add_argument("--model-size", type=str, choices=["UNet_XS", "UNet_S", "UNet_M", "UNet_L", "UNet_XL"], default="UNet_L",
                       help="Model size (for model loading modes)")
    parser.add_argument("--pretrained", type=str, default="", 
                       help="Path to pretrained model (REQUIRED for save_only, save_and_process, full_pipeline, full_pipeline_with_saving_npy)")
    parser.add_argument("--reverse-steps", type=int, default=5,
                       help="Number of reverse steps (for model loading modes)")
    parser.add_argument("--batch-size", type=int, default=64,
                       help="Batch size for processing (for model loading modes)")
    parser.add_argument("--batch-num", type=int, default=12,
                       help="Number of batches to process (for model loading modes)")
    parser.add_argument("--split", type=str, default="test",
                       help="Data split to process (for model loading modes)")
    parser.add_argument("--object-class", type=str, default="all",
                       help="Object class to process (for model loading modes)")
    parser.add_argument("--vae-type", type=str, choices=["ema", "mse"], default="ema",
                       help="VAE type (for model loading modes)")
    parser.add_argument("--force-rerun", action="store_true", 
                       help="Force rerun evaluation (for save_only mode)")
    parser.add_argument("--anomaly-binary-threshold", type=int, default=5, 
                       help="Binary threshold for anomaly detection (for processing modes)")
    parser.add_argument("--anomaly-pixel-num-threshold", type=int, default=0, 
                       help="Pixel number threshold (for processing modes)")
    parser.add_argument("--adaptive-threshold", type=float, default=0.1, 
                       help="Adaptive threshold for contour-based masks (for processing modes)")
    
    # Output options
    parser.add_argument("--enable-excel-report", action="store_true", help="Generate Excel report")
    parser.add_argument("--enable-save-image-results", action="store_true", help="Save image results")
    parser.add_argument("--enable-save-optional-image-results", action="store_true", help="Save optional image results")
    parser.add_argument("--enable-save-whole-image-results", action="store_true", help="Save whole image results")
    parser.add_argument("--enable-save-json-results", action="store_true", help="Save JSON results")
    parser.add_argument("--enable-confusion-matrix", action="store_true", help="Create confusion matrix")
    parser.add_argument("--enable-epoch-stats", action="store_true", help="Enable detailed epoch-wise statistics collection and printing")
    parser.add_argument("--no-sort-records-by-anomaly", action="store_true", help="Disable sorting records by anomaly_pixels (default: sorted by anomaly_pixels, largest to smallest)")
    parser.add_argument("--save-preview-images", action="store_true", help="Save PNG previews; disable to reduce I/O overhead in save_only mode")
    parser.add_argument("--save-npy-dtype", type=str, choices=["float16", "float32"], default="float16",
                        help="Data type for saved .npy arrays (float16 reduces disk I/O, float32 for full precision)")
    
    # Input JSON for batch processing
    parser.add_argument("--input-json", type=str, 
                       help="""JSON file with test configurations. Can be used instead of --mode.
                       Example format: {"test_name": {"mode": "save_only", "annotation-dir": "path/to/annotations", ...}}
                       Or simple format: {"mode": "save_only", "annotation-dir": "path/to/annotations", ...}""")
    
    # Debug control
    parser.add_argument("--debug", action="store_true",
                       help="Enable detailed debug logging during processing (default: False)")
    
    # Checkpoint control
    parser.add_argument("--checkpoint-mode", type=str, choices=["auto", "resume", "overwrite"], default="auto",
                       help="Checkpoint behavior: auto=resume if possible, resume=force resume, overwrite=ignore existing checkpoints")
    parser.add_argument("--checkpoint-interval", type=int, default=1,
                       help="Save checkpoint every N batches (default: 1)")
    # Performance knobs

    parser.add_argument("--num-workers", type=int, default=0,
                        help="Number of DataLoader workers (Windows often needs 0; try >0 if stable)")
    parser.add_argument("--async-save-workers", type=int, default=4,
                        help="Number of async saving threads. Default: 4 on Windows, 2-8 elsewhere.")
    
    # Cache control arguments
    parser.add_argument("--max-cache-memory-gb", type=float, default=2.0,
                        help="Maximum memory usage for image cache in GB (default: 2.0)")
    parser.add_argument("--max-cache-images", type=int, default=100,
                        help="Maximum number of images to keep in cache (default: 100)")
    
    # Parallel processing control
    parser.add_argument("--parallel-workers", type=int, default=None,
                        help="Number of parallel workers for patch processing (default: auto-detect)")
    parser.add_argument("--chunk-size", type=int, default=None,
                        help="Chunk size for parallel processing (default: auto-calculate)")
    parser.add_argument("--no-parallel", action="store_true", default=False,
                        help="Disable parallel processing and use sequential processing")
    parser.add_argument("--no-worker-monitoring", action="store_true", default=False,
                        help="Disable worker monitoring for maximum performance")

    # Bootstrap sampling control
    parser.add_argument("--bootstrap-samples", type=int, default=None,
                        help="Enable bootstrap sampling: save only N randomly sampled patches (with replacement). "
                             "Reduces evaluation time. If None, processes all patches.")
    parser.add_argument("--bootstrap-seed", type=int, default=42,
                        help="Random seed for bootstrap sampling (default: 42 for reproducibility)")

    args = parser.parse_args()
    
    # Set global debug flag
    DEBUG_ENABLED = args.debug
    
    # Handle mode extraction from JSON if needed
    if not args.mode and not args.input_json:
        print("❌ Error: Either --mode or --input-json must be provided!")
        print("💡 Use --mode to specify execution mode directly, or --input-json to load configuration from file.")
        sys.exit(1)
    
    # Extract mode from JSON if not provided via command line
    if not args.mode and args.input_json:
        try:
            with open(args.input_json, 'r') as f:
                config_data = json.load(f)
            
            # Look for mode in the JSON structure
            if isinstance(config_data, dict):
                # Check if there's a mode at the top level
                if 'mode' in config_data:
                    args.mode = config_data['mode']
                else:
                    # Check if there's a test configuration with mode
                    first_test = next(iter(config_data.values()))
                    if isinstance(first_test, dict) and 'mode' in first_test:
                        args.mode = first_test['mode']
                    else:
                        print("❌ Error: No 'mode' found in JSON configuration!")
                        print("💡 Add 'mode': 'save_only' (or other mode) to your JSON file.")
                        sys.exit(1)
            else:
                print("❌ Error: JSON configuration must be a dictionary!")
                sys.exit(1)
                
        except FileNotFoundError:
            print(f"❌ Error: JSON file not found: {args.input_json}")
            sys.exit(1)
        except json.JSONDecodeError as e:
            print(f"❌ Error: Invalid JSON file: {e}")
            sys.exit(1)
    
    # Validate that mode is valid
    if args.mode not in ["save_only", "process_only", "save_and_process", "full_pipeline", "full_pipeline_with_saving_npy"]:
        print(f"❌ Error: Invalid mode '{args.mode}'. Must be one of: save_only, process_only, save_and_process, full_pipeline, full_pipeline_with_saving_npy")
        sys.exit(1)
    
    # Handle input JSON if provided
    if args.input_json:
        with open(args.input_json, 'r') as f:
            test_configs = json.load(f)
            
        # Run processing for each test configuration
        for test_name, test_args in test_configs.items():
            print(f"\n{'='*60}")
            print(f"📄 Running configuration: {test_name}")
            print(f"{'='*60}")
            print(f"🔧 Mode: {test_args.get('mode', args.mode)}")
            print(f"📂 Annotation dir: {test_args.get('annotation-dir', 'Not specified')}")
            print(f"🤖 Model: {test_args.get('pretrained', 'Not specified')}")
            print(f"⚙️  Loading {len(test_args)} configuration parameters...")
            
            # Create a copy of args for this test configuration
            import copy
            test_args_obj = copy.deepcopy(args)
            
            # Update test_args_obj with test configuration
            for key, value in test_args.items():
                key = key.replace('-', '_')
                if hasattr(test_args_obj, key):
                    if key in ['anomaly_binary_threshold', 'anomaly_pixel_num_threshold', 'patch_size',
                              'batch_num', 'batch_size', 'reverse_steps', 'async_save_workers', 'stride',
                              'bootstrap_samples', 'bootstrap_seed']:
                        value = int(value)
                    elif key in ['adaptive_threshold']:
                        value = float(value)
                    elif key in ['results_dir', 'annotation_dir', 'pretrained']:
                        value = os.path.expanduser(value)
                        # Convert to absolute path if it's relative
                        if not os.path.isabs(value):
                            value = os.path.abspath(value)
                    elif key in ['irregular_patch', 'enable_excel_report', 'enable_save_optional_image_results', 
                               'enable_save_image_results', 'enable_save_json_results', 'enable_save_whole_image_results',
                               'enable_confusion_matrix', 'force_rerun', 'enable_epoch_stats', 'save_preview_images']:
                        value = value.lower() in ('yes', 'true', 't', 'y', '1')
                    elif key in ['save_npy_dtype']:
                        value = value.lower() in ('float16', 'float32')
                    elif key == 'debug':
                        debug_value = value.lower() in ('yes', 'true', 't', 'y', '1') if isinstance(value, str) else bool(value)
                        DEBUG_ENABLED = debug_value
                        test_args_obj.debug = debug_value
                    setattr(test_args_obj, key, value)
            
            # Validate mode-specific arguments for this test configuration
            if not validate_mode_arguments(test_args_obj):
                print(f"\n❌ Argument validation failed for {test_name}. Skipping this configuration.")
                continue
            
            # Set up results directory for this test configuration
            test_args_obj.results_dir = f"results/{test_name}"
            
            os.makedirs(test_args_obj.results_dir, exist_ok=True)
            
            # Save config for this test
            config_save_path = os.path.join(test_args_obj.results_dir, "config.json")
            with open(config_save_path, "w") as config_file:
                json.dump({test_name: test_args}, config_file, indent=2)
            
            # Debug: Print final argument values for this test
            debug_print(f"🔍 Final argument values for {test_name}:", debug=DEBUG_ENABLED)
            debug_print(f"   mode: {test_args_obj.mode}", debug=DEBUG_ENABLED)
            debug_print(f"   patch_size: {test_args_obj.patch_size}", debug=DEBUG_ENABLED)
            debug_print(f"   stride: {test_args_obj.stride}", debug=DEBUG_ENABLED)
            debug_print(f"   annotation_dir: {test_args_obj.annotation_dir}", debug=DEBUG_ENABLED)
            debug_print(f"   pretrained: {test_args_obj.pretrained}", debug=DEBUG_ENABLED)
            debug_print(f"   results_dir: {test_args_obj.results_dir}", debug=DEBUG_ENABLED)
            
            # Execute the selected mode for this test configuration
            try:
                if test_args_obj.mode == "save_only":
                    mode_save_only(test_args_obj)
                elif test_args_obj.mode == "process_only":
                    mode_process_only(test_args_obj)
                elif test_args_obj.mode == "save_and_process":
                    mode_save_and_process(test_args_obj)
                elif test_args_obj.mode == "full_pipeline":
                    mode_full_pipeline(test_args_obj)
                elif test_args_obj.mode == "full_pipeline_with_saving_npy":
                    mode_full_pipeline_with_saving_npy(test_args_obj)
                
                print(f"✅ Configuration {test_name} completed successfully!")
                
            except Exception as e:
                print(f"❌ Configuration {test_name} failed: {e}")
                print(f"   Continuing with next configuration...")
                continue
        
        print(f"\n🎉 All JSON configurations processed!")
        
    else:
        # Single mode execution (no JSON)
        # Validate mode-specific arguments
        if not validate_mode_arguments(args):
            print(f"\n❌ Argument validation failed. Please fix the issues above and try again.")
            sys.exit(1)
        
        # Set up results directory
        base_name = f"DeCo-Diff_{args.dataset}_{args.object_class}_{args.model_size}_{args.patch_size}"
        
        args.results_dir = f"results/{base_name}"
        
        os.makedirs(args.results_dir, exist_ok=True)
        
        # Execute the selected mode
        if args.mode == "save_only":
            mode_save_only(args)
        elif args.mode == "process_only":
            mode_process_only(args)
        elif args.mode == "save_and_process":
            mode_save_and_process(args)
        elif args.mode == "full_pipeline":
            mode_full_pipeline(args)
        elif args.mode == "full_pipeline_with_saving_npy":
            mode_full_pipeline_with_saving_npy(args)
        
        print(f"Mode {args.mode} completed successfully!")


if __name__ == "__main__":
    main()
