# %%
from __future__ import annotations
import os
os.environ['NO_ALBUMENTATIONS_UPDATE'] = '1'

import warnings
warnings.filterwarnings(
    "ignore",
    message="A new version of Albumentations is available.*",
    category=UserWarning
)
from datetime import datetime
import torch
from skimage.transform import resize
from diffusion import create_diffusion
from diffusers.models.autoencoders.autoencoder_kl import AutoencoderKL
from models import UNET_models
import argparse
import numpy as np
import torch.nn.functional as F

from glob import glob

from torch.utils.data import DataLoader, Dataset
from torchvision import transforms
from MVTECDataLoader import MVTECDataset
from VISADataLoader import VISADataset
from PCBDataLoader import PCBDataset
from scipy.ndimage import gaussian_filter

from anomalib import metrics
from sklearn.metrics import average_precision_score
from numpy import ndarray
import pandas as pd
from skimage import measure
from sklearn.metrics import auc

import sys
from typing import List
import matplotlib.pyplot as plt
from collections import OrderedDict, defaultdict
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from typing import Sequence

from io import BytesIO
from pathlib import Path

from typing import Any, Tuple, cast

from torchmetrics.functional.image import (
    learned_perceptual_image_patch_similarity as _lpips,
    structural_similarity_index_measure as _ssim,
)
from sklearn.metrics import roc_curve
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay
import json
import cv2
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score

import re
import platform
from utils import path_to_safe_filename

torch.set_grad_enabled(False)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

if device == torch.device("cpu"):
    print("GPU not found. Using CPU instead.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_LATENT_SCALE = 0.18215

Kinded = Tuple[str, Any]  # (kind, value)
Record = OrderedDict[str, Kinded]

# ---------------------------------------------------------------------------
# Base Classes for Reducing Boilerplate
# ---------------------------------------------------------------------------

class BaseEvaluator:
    """Base class for evaluation operations to reduce boilerplate code."""
    
    def __init__(self, args, diffusion, model, vae, device=None):
        self.args = args
        self.diffusion = diffusion
        self.model = model
        self.vae = vae
        self.device = device or torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.transform = self._get_transform()
        self.results_dir = args.results_dir
        
    def _get_transform(self):
        """Get standard transform for all datasets."""
        return transforms.Compose([
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5], inplace=True)
        ])
    
    def _get_dataloader(self, dataset, batch_size=None, shuffle=False):
        """Get standard dataloader configuration."""
        batch_size = batch_size or getattr(self.args, 'batch_size', 64)
        
        # Memory optimization: Reduce batch size for large datasets
        if hasattr(self.args, 'memory_optimization') and self.args.memory_optimization:
            batch_size = min(batch_size, 1)  # Force batch size of 1 for memory optimization
            print(f"Memory optimization enabled: using batch_size={batch_size}")
        
        # Windows-friendly DataLoader configuration
        # Use 0 workers on Windows to avoid multiprocessing issues
        
        # Use command line argument if provided, otherwise use platform-specific defaults
        if hasattr(self.args, 'num_workers') and self.args.num_workers is not None:
            num_workers = self.args.num_workers
        elif platform.system() == 'Windows':
            num_workers = 0  # Single-threaded on Windows to avoid hanging
        else:
            num_workers = 2 if hasattr(self.args, 'memory_optimization') and self.args.memory_optimization else 4
        
        return DataLoader(
            dataset, 
            batch_size=batch_size, 
            shuffle=shuffle, 
            num_workers=num_workers, 
            drop_last=False,
            persistent_workers=False if num_workers == 0 else True
        )
    
    def _create_output_dirs(self):
        """Create standard output directories."""
        dirs = {
            'marked_images': os.path.join(self.results_dir, "marked_images"),
            'evaluation_results': os.path.join(self.results_dir, "evaluation_results"),
            'accuracy_plots': os.path.join(self.results_dir, "accuracy_vs_param")
        }
        for dir_path in dirs.values():
            os.makedirs(dir_path, exist_ok=True)
        return dirs
    
    def _save_image(self, img, path, description=""):
        """Save image to path."""
        import os
        from PIL import Image
        import numpy as np
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(path), exist_ok=True)
        
        # Convert numpy array to PIL Image if needed
        if isinstance(img, np.ndarray):
            # Handle different array formats
            if img.dtype == np.uint8:
                pil_img = Image.fromarray(img)
            else:
                # Convert float arrays to uint8
                if img.max() <= 1.0:
                    img = (img * 255).astype(np.uint8)
                else:
                    img = img.astype(np.uint8)
                pil_img = Image.fromarray(img)
        elif hasattr(img, 'save'):
            # Already a PIL Image
            pil_img = img
        else:
            raise ValueError(f"Unsupported image type: {type(img)}")
        
        pil_img.save(path)
        print(f"Saved {description}: {path}")
    
    def _save_json(self, data, path, description=""):
        """Save JSON with standard error handling."""
        try:
            with open(path, 'w') as f:
                json.dump(data, f, indent=2)
            if description:
                print(f"Saved {description}: {path}")
        except Exception as e:
            print(f"Error saving {description} to {path}: {e}")

class ImageProcessor:
    """Utility class for common image processing operations."""
    
    @staticmethod
    def create_anomaly_map_image(anomaly_map, predicted_defective_set, ground_truth_defective, overlapping, is_binary=True, patch_size=128, add_grid=True, grid_color=(255, 255, 255), grid_thickness=1):
        """Create anomaly map visualization with optional grid overlay and patch prediction rectangles."""
        if is_binary:
            # Binary map: 0 or 1 - create custom red colormap
            anomaly_map_img = (anomaly_map * 255).astype(np.uint8)
            # Create custom colormap: 0 -> transparent (black), 255 -> pure red
            h, w = anomaly_map_img.shape
            anomaly_map_colored_bgr = np.zeros((h, w, 3), dtype=np.uint8)
            # Set red channel based on anomaly values (0 = transparent/black, 255 = pure red)
            anomaly_map_colored_bgr[:, :, 2] = anomaly_map_img  # Red channel (BGR format)
        else:
            anomaly_map_img = (anomaly_map * 255).astype(np.uint8)
            anomaly_map_colored_bgr = cv2.applyColorMap(anomaly_map_img, cv2.COLORMAP_HOT)

        # Add grid overlay if requested
        if add_grid:
            h, w = anomaly_map.shape
            
            # Draw vertical grid lines
            for x in range(patch_size, w, patch_size):
                cv2.line(anomaly_map_colored_bgr, (x, 0), (x, h), grid_color, grid_thickness)
            
            # Draw horizontal grid lines
            for y in range(patch_size, h, patch_size):
                cv2.line(anomaly_map_colored_bgr, (0, y), (w, y), grid_color, grid_thickness)
            
            # Draw border around the entire image
            cv2.rectangle(anomaly_map_colored_bgr, (0, 0), (w-1, h-1), grid_color, grid_thickness)
        
        # Convert BGR to RGB for proper display
        anomaly_map_colored = cv2.cvtColor(anomaly_map_colored_bgr, cv2.COLOR_BGR2RGB)
        
        # Add patch prediction rectangles if provided
        if predicted_defective_set is not None and ground_truth_defective is not None and overlapping is not None:
            anomaly_map_colored = draw_patch_rectangles_on_image(
                anomaly_map_colored, 
                predicted_defective_set, 
                ground_truth_defective, 
                overlapping, 
                patch_size=patch_size, 
                grid_thickness=grid_thickness
            )
        
        return anomaly_map_colored
    
    @staticmethod
    def create_anomaly_overlay(original_img, anomaly_map, alpha=0.6, is_binary=True):
        """Create overlay of anomaly map on original image."""
        if is_binary:
            # Binary overlay using pure red
            overlay = original_img.copy()
            mask = anomaly_map > 0
            # Create pure red overlay for anomaly regions
            h, w = anomaly_map.shape
            anomaly_colored = np.zeros((h, w, 3), dtype=np.uint8)
            # Set red channel for anomaly regions (pure red)
            anomaly_colored[mask, 0] = 255  # Red channel (RGB format)
            # Apply the red anomaly regions to the overlay
            overlay[mask] = anomaly_colored[mask]
        else:
            # Continuous overlay using HOT colormap
            # Convert [0, 1] range to [0, 255] range for colormap
            anomaly_map_uint8 = (anomaly_map * 255).astype(np.uint8)
            
            # Apply HOT colormap to the map (returns BGR)
            anomaly_colored_bgr = cv2.applyColorMap(anomaly_map_uint8, cv2.COLORMAP_HOT)
            # Convert BGR to RGB for proper overlay
            anomaly_colored = cv2.cvtColor(anomaly_colored_bgr, cv2.COLOR_BGR2RGB)
            # Create overlay using alpha blending
            overlay = cv2.addWeighted(original_img, 1-alpha, anomaly_colored, alpha, 0)
        
        return overlay.astype(np.uint8)
    


class EvaluationMetrics:
    """Utility class for computing and storing evaluation metrics with memory-efficient approach."""
    
    def __init__(self):
        # Use running statistics instead of storing all values
        self.epoch_stats = {
            'encodedrecon_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []},
            'latent_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []},
            'anomaly_map_arithmetic_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []},
            'anomaly_map_geometric_values': {'count': 0, 'sum': 0.0, 'sum_sq': 0.0, 'min': float('inf'), 'max': float('-inf'), 'values': []}
        }
        # Store histogram bins for distribution analysis
        self.hist_bins = np.arange(0, 0.4, 0.01)
        self.histograms = {
            'encodedrecon_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64),
            'latent_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64),
            'anomaly_map_arithmetic_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64),
            'anomaly_map_geometric_values': np.zeros(len(self.hist_bins) - 1, dtype=np.int64)
        }
    
    def add_batch_stats(self, encodedrecon_diff, latent_diff, anomaly_map_arithmetic, anomaly_map_geometric):
        """Add batch statistics to epoch collection using running statistics."""
        # Process encodedrecon_diff
        encodedrecon_flat = encodedrecon_diff.flatten().cpu().numpy()
        self._update_running_stats('encodedrecon_values', encodedrecon_flat)
        self._update_histogram('encodedrecon_values', encodedrecon_flat)
        
        # Process latent_diff
        latent_flat = latent_diff.flatten().cpu().numpy()
        self._update_running_stats('latent_values', latent_flat)
        self._update_histogram('latent_values', latent_flat)

        # Process anomaly_map_arithmetic
        anomaly_map_arithmetic_flat = anomaly_map_arithmetic.flatten().cpu().numpy()
        self._update_running_stats('anomaly_map_arithmetic_values', anomaly_map_arithmetic_flat)
        self._update_histogram('anomaly_map_arithmetic_values', anomaly_map_arithmetic_flat)
        
        # Process anomaly_map_geometric
        anomaly_map_geometric_flat = anomaly_map_geometric.flatten().cpu().numpy()
        self._update_running_stats('anomaly_map_geometric_values', anomaly_map_geometric_flat)
        self._update_histogram('anomaly_map_geometric_values', anomaly_map_geometric_flat)

    
    def _update_running_stats(self, key, values):
        """Update running statistics for a given key."""
        stats = self.epoch_stats[key]
        n = len(values)
        
        if n == 0:
            return
            
        # Update count
        stats['count'] += n
        
        # Update sum and sum of squares
        stats['sum'] += np.sum(values)
        stats['sum_sq'] += np.sum(values ** 2)
        
        # Update min and max
        stats['min'] = min(stats['min'], np.min(values))
        stats['max'] = max(stats['max'], np.max(values))
        
        stats['values'].extend(values.tolist())
        ## Store values for quantile calculation (memory efficient: only store every 10th value for large datasets)
        #if stats['count'] <= 10000:  # Store all values if dataset is small
        #    stats['values'].extend(values.tolist())
        #else:
        #    # For large datasets, sample every 10th value to keep memory usage reasonable
        #    if 'sampling_counter' not in stats:
        #        stats['sampling_counter'] = 0
        #    stats['sampling_counter'] += 1
        #    if stats['sampling_counter'] % 10 == 0:
        #        stats['values'].extend(values[::10].tolist())
    
    def _update_histogram(self, key, values):
        """Update histogram for a given key."""
        hist, _ = np.histogram(values, bins=self.hist_bins, range=(0, 3))
        self.histograms[key] += hist
    
    def print_epoch_stats(self):
        """Print epoch-wise statistics."""
        print(f'\n=== EPOCH-WISE STATISTICS ===')
        
        for name, stats in self.epoch_stats.items():
            if stats['count'] > 0:
                # Calculate statistics from running values
                mean = stats['sum'] / stats['count']
                variance = (stats['sum_sq'] / stats['count']) - (mean ** 2)
                std = np.sqrt(max(0, variance))  # Ensure non-negative
                
                print(f'{name} epoch stats:')
                print(f'  min: {stats["min"]:.6f}, max: {stats["max"]:.6f}')
                print(f'  mean: {mean:.6f}, std: {std:.6f}')
                
                # Calculate quantiles if we have stored values
                if len(stats['values']) > 0:
                    values_array = np.array(stats['values'])
                    q1 = np.percentile(values_array, 25)
                    median = np.percentile(values_array, 50)
                    q3 = np.percentile(values_array, 75)
                    print(f'  Q1: {q1:.6f}, median: {median:.6f}, Q3: {q3:.6f}')
                
                # Print histogram
                hist = self.histograms[name]
                print(f'{name} epoch distribution:')
                for i, (count, edge) in enumerate(zip(hist, self.hist_bins[:-1])):
                    if stats['count'] > 0:
                        percentage = (count / stats['count']) * 100
                        print(f'  [{edge:.2f}-{edge+0.01:.2f}): {count:6d} ({percentage:5.1f}%)')

# ---------------------------------------------------------------------------
# New Dataset Class for Irregular Images
# ---------------------------------------------------------------------------

class AnnotatedImageDataset(Dataset):
    """Dataset for handling images with JSON annotations for defective regions."""
    
    def __init__(
        self,
        annotation_dir: str,
        patch_size: int = 128,
        transform=None,
        object_class: str = "pcb",
    ):
        """
        Args:
            annotation_dir: Directory containing JSON annotation files
            patch_size: Size of patches to extract (default: 128)
            transform: Optional transform to apply to patches
            object_class: Object class for classification
        """
        self.annotation_dir = annotation_dir
        self.patch_size = patch_size
        self.transform = transform
        self.object_class = object_class
        
        # Find all annotation files
        annotation_files = glob(os.path.join(annotation_dir, "*__annotations.json"))
        
        self.images = []
        self.patches = []
        self.patch_coords = []  # (x, y) coordinates of each patch in original image
        self.image_paths = []
        self.anomaly_classes = []
        self.is_defective = []  # Boolean indicating if patch is defective
        
        for annotation_file in annotation_files:
            with open(annotation_file, 'r') as f:
                annotation = json.load(f)
            
            image_path = annotation["image_path"]
            if not os.path.exists(image_path):
                print(f"Warning: Image {image_path} not found, skipping...")
                continue
            
            # Load original image
            img = np.array(PILImage.open(image_path).convert('RGB'))
            original_height, original_width = img.shape[:2]
            
            # Get defective patch coordinates
            defective_patches = set()
            for patch_coord in annotation["defective_patches"]:
                grid_row, grid_col = patch_coord
                defective_patches.add((grid_row, grid_col))
            
            # Generate all possible patches including edge patches
            grid_rows = (original_height + patch_size - 1) // patch_size  # Ceiling division
            grid_cols = (original_width + patch_size - 1) // patch_size   # Ceiling division
            
            for grid_row in range(grid_rows):
                for grid_col in range(grid_cols):
                    # Calculate pixel coordinates
                    x = grid_col * patch_size
                    y = grid_row * patch_size
                    
                    # Calculate actual patch dimensions (may be smaller at edges)
                    patch_width = min(patch_size, original_width - x)
                    patch_height = min(patch_size, original_height - y)
                    
                    # Extract patch
                    patch = img[y:y + patch_height, x:x + patch_width]
                    
                    # Pad patch to full size if necessary
                    if patch.shape[:2] != (patch_size, patch_size):
                        padded_patch = np.zeros((patch_size, patch_size, 3), dtype=patch.dtype)
                        padded_patch[:patch_height, :patch_width] = patch
                        patch = padded_patch
                    
                    # Determine if this patch is defective
                    is_defective = (grid_row, grid_col) in defective_patches
                    
                    self.images.append(img)  # Store original image
                    self.patches.append(patch)
                    self.patch_coords.append((x, y))
                    self.image_paths.append(image_path)
                    self.anomaly_classes.append("defect" if is_defective else "normal")
                    self.is_defective.append(is_defective)
        
        print(f"Created {len(self.patches)} patches from {len(set(self.image_paths))} images")
        print(f"Defective patches: {sum(self.is_defective)}, Normal patches: {len(self.is_defective) - sum(self.is_defective)}")
    
    def __len__(self):
        return len(self.patches)
    
    def __getitem__(self, index):
        patch = self.patches[index].astype(np.float32) / 255.0
        original_img = self.images[index]
        x, y = self.patch_coords[index]
        image_path = self.image_paths[index]
        anomaly_class = self.anomaly_classes[index]
        is_defective = self.is_defective[index]
        
        # Create segmentation mask (all zeros for normal patches, all ones for defective)
        seg = np.ones((self.patch_size, self.patch_size), dtype=np.float32) if is_defective else np.zeros((self.patch_size, self.patch_size), dtype=np.float32)
        
        # Apply transform if provided
        if self.transform:
            patch = self.transform(patch)
        else:
            patch = torch.from_numpy(patch.transpose(2, 0, 1))
            patch = (patch - 0.5) / 0.5
        
        # Convert coordinates to tensors for proper batching
        coords_tensor = torch.tensor([x, y], dtype=torch.int32)
        
        # Don't return original_img in the batch to avoid tensor size mismatch
        # We'll handle original image loading separately in the processing function
        return patch, seg, 0, image_path, anomaly_class, coords_tensor


class IrregularImageDataset(Dataset):
    """Dataset for handling irregular-sized images by splitting them into patches."""
    
    def __init__(
        self,
        data_dir: str,
        patch_size: int = 128,
        stride: int = 64,  # Overlap between patches
        transform=None,
        object_class: str = "pcb",
        anomaly_class: str = "all",
        split_csv_path: str | None = None,
    ):
        """
        Args:
            data_dir: Directory containing the original images
            patch_size: Size of patches to extract (default: 128)
            stride: Stride between patches (default: 64 for overlap)
            transform: Optional transform to apply to patches
            object_class: Object class for classification
            anomaly_class: Anomaly class filter
            split_csv_path: Path to CSV file with image information
        """
        self.data_dir = data_dir
        self.patch_size = patch_size
        self.stride = stride
        self.transform = transform
        self.object_class = object_class
        self.anomaly_class = anomaly_class
        
        # Load image information from CSV
        if split_csv_path and os.path.exists(split_csv_path):
            df = pd.read_csv(split_csv_path)
            if anomaly_class != "all":
                df = df.query(f'category=="{anomaly_class}"')
        else:
            # If no CSV, scan directory for images
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']
            image_files = []
            for ext in image_extensions:
                image_files.extend(glob(os.path.join(data_dir, f"*{ext}")))
                image_files.extend(glob(os.path.join(data_dir, f"*{ext.upper()}")))
            
            # Create a simple dataframe
            df = pd.DataFrame({
                'image': [os.path.basename(f) for f in image_files],
                'category': ['unknown'] * len(image_files),
                'object': [object_class] * len(image_files)
            })
        
        self.images = []
        self.patches = []
        self.patch_coords = []  # (x, y) coordinates of each patch in original image
        self.image_paths = []
        self.anomaly_classes = []
        
        object_cls_dict = {"pcb": 0}
        
        for _, row in df.iterrows():
            image_path = os.path.join(data_dir, str(row['image']))
            if not os.path.exists(image_path):
                continue

            # Load original image
            img = np.array(PILImage.open(image_path).convert('RGB'))
            original_height, original_width = img.shape[:2]
            
            # Generate patches
            patches, coords = self._extract_patches(img)
            
            for patch, (x, y) in zip(patches, coords):
                self.images.append(img)  # Store original image
                self.patches.append(patch)
                self.patch_coords.append((x, y))
                self.image_paths.append(image_path)
                self.anomaly_classes.append(row['category'])
        
        print(f"Created {len(self.patches)} patches from {len(set(self.image_paths))} images")
    
    def _extract_patches(self, img):
        """Extract patches from an image with overlap, including edge patches."""
        height, width = img.shape[:2]
        patches = []
        coords = []
        
        # Calculate the range of patch positions
        y_positions = list(range(0, height, self.stride))
        x_positions = list(range(0, width, self.stride))
        
        # Ensure we include edge patches
        if y_positions and height - y_positions[-1] < self.patch_size:
            # Add a position that will create an edge patch
            y_positions.append(max(0, height - self.patch_size))
        if x_positions and width - x_positions[-1] < self.patch_size:
            # Add a position that will create an edge patch
            x_positions.append(max(0, width - self.patch_size))
        
        for y in y_positions:
            for x in x_positions:
                # Calculate actual patch dimensions
                patch_width = min(self.patch_size, width - x)
                patch_height = min(self.patch_size, height - y)
                
                # Extract patch
                patch = img[y:y + patch_height, x:x + patch_width]
                
                # Pad patch to full size if necessary
                if patch.shape[:2] != (self.patch_size, self.patch_size):
                    padded_patch = np.zeros((self.patch_size, self.patch_size, 3), dtype=patch.dtype)
                    padded_patch[:patch_height, :patch_width] = patch
                    patch = padded_patch
                
                patches.append(patch)
                coords.append((x, y))
        
        # Handle edge cases if image is smaller than patch_size
        if height < self.patch_size or width < self.patch_size:
            # Pad the image to patch_size
            padded_img = np.zeros((max(height, self.patch_size), max(width, self.patch_size), 3), dtype=img.dtype)
            padded_img[:height, :width] = img
            patch = padded_img[:self.patch_size, :self.patch_size]
            patches = [patch]
            coords = [(0, 0)]
        
        return patches, coords
    
    def __len__(self):
        return len(self.patches)
    
    def __getitem__(self, index):
        patch = self.patches[index].astype(np.float32) / 255.0
        original_img = self.images[index]
        x, y = self.patch_coords[index]
        image_path = self.image_paths[index]
        anomaly_class = self.anomaly_classes[index]
        
        # Create dummy segmentation (all zeros for normal patches)
        seg = np.zeros((self.patch_size, self.patch_size), dtype=np.float32)
        
        # Apply transform if provided
        if self.transform:
            patch = self.transform(patch)
        else:
            patch = torch.from_numpy(patch.transpose(2, 0, 1))
            patch = (patch - 0.5) / 0.5
        
        # Convert coordinates to tensors for proper batching
        coords_tensor = torch.tensor([x, y], dtype=torch.int32)
        
        return patch, seg, 0, image_path, anomaly_class, coords_tensor, original_img

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def add_metric_fields(rec: Record, *, device=torch.device("cpu")) -> None:

    def to4d(x):
        if isinstance(x, np.ndarray):
            x = torch.from_numpy(x)
            if x.dtype != torch.float32:
                x = x.float()
            if x.ndim == 3 and x.shape[-1] == 3:  # HWC ➜ CHW
                x = x.permute(2, 0, 1)
            if x.ndim == 2:
                x = x.unsqueeze(0)
            if x.ndim == 3:
                x = x.unsqueeze(0)
            return x.to(device).clamp(-1, 1)
        return None

    a = to4d(rec["encoded_recon"][1])
    b = to4d(rec["dod_recon"][1])
    
    # Add null checks before computing metrics
    if a is not None and b is not None:
        rec["lpips"] = ("metric", _lpips(a, b, net_type="alex").item())
        ssim_result = _ssim(a, b)
        rec["ssim"] = ("metric", ssim_result[0].item() if isinstance(ssim_result, tuple) else ssim_result.item())
        rec["mse"] = ("metric", F.mse_loss(a, b).item())
    else:
        # Set default values if conversion fails
        rec["lpips"] = ("metric", 0.0)
        rec["ssim"] = ("metric", 0.0)
        rec["mse"] = ("metric", 0.0)


def make_record(**kwargs) -> Record:
    """Return an **ordered** dict whose values are (kind, value) pairs."""
    return OrderedDict(kwargs)


def _compute_diff_mean(a: torch.Tensor, b: torch.Tensor, diff_scale: float = 1.0) -> torch.Tensor:
    """Return the mean channel‑wise difference *scaled* by ``diff_scale``."""
    return (a - b).mean(dim=1, keepdim=True) / diff_scale

def _compute_abs_diff_mean(a: torch.Tensor, b: torch.Tensor, diff_scale: float = 1.0) -> torch.Tensor:
    """Return the mean channel‑wise absolute difference *scaled* by ``diff_scale``."""
    return torch.abs(a - b).mean(dim=1, keepdim=True) / diff_scale

def _compute_abs_diff_max(a: torch.Tensor, b: torch.Tensor) -> torch.Tensor:
    """Return the maximum channel‑wise absolute difference."""
    return torch.abs(a - b).max(dim=1, keepdim=True).values

def _binary_mask(diff: torch.Tensor, threshold: int = 5) -> torch.Tensor:
    """Return a binary mask in ``{0, 1}`` based on *absolute* diff magnitude."""
    return (diff.abs() > (threshold / 255.0)).float()


def _get_largest_connected_component_pixels(anomaly_binary: torch.Tensor) -> int:
    """
    Calculate the number of pixels in the largest connected component of white pixels.
    
    Args:
        anomaly_binary: Binary tensor with shape (H, W) or (1, H, W) where 1 indicates white pixels
        
    Returns:
        Number of pixels in the largest connected component
    """
    import cv2
    import numpy as np
    
    # Convert to numpy and ensure 2D shape
    if anomaly_binary.dim() == 3 and anomaly_binary.shape[0] == 1:
        binary_np = anomaly_binary.squeeze(0).cpu().numpy()
    else:
        binary_np = anomaly_binary.cpu().numpy()
    
    # Ensure binary values (0 or 1)
    binary_np = (binary_np > 0).astype(np.uint8)
    
    # Find connected components
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary_np, connectivity=8)
    
    if num_labels <= 1:  # Only background (label 0) or no components
        return 0
    
    # Find the largest component (excluding background which is label 0)
    largest_component_size = 0
    for i in range(1, num_labels):  # Skip background (i=0)
        component_size = stats[i, cv2.CC_STAT_AREA]
        if component_size > largest_component_size:
            largest_component_size = component_size
    
    return largest_component_size


def _create_contour_based_binary_mask(anomaly_map: torch.Tensor, adaptive_threshold: float = 0.1, anomaly_binary_threshold: int = 5) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
    """
    Create three different binary masks based on contours with adaptive selection based on distribution.
    
    Args:
        anomaly_map: Anomaly map tensor with shape (batch_size, 1, H, W) with values in [0, 1]
        adaptive_threshold: Threshold for adaptive contour selection (default: 0.1)
                          - Lower values select more contours
                          - Higher values select fewer contours
        anomaly_binary_threshold: Threshold value for binary conversion (0-255, default: 5)
                                 - Lower values create more white pixels
                                 - Higher values create fewer white pixels
        
    Returns:
        Tuple of three binary tensors with same shape as input where selected contour pixels are 1, others are 0:
        - style1: Top contours by sum (most significant)
        - style2: Statistical outliers (mean + threshold * std)
        - style3: Contours contributing significant portion of total
    """
    import cv2
    import numpy as np
    
    # Handle batch processing for shape (batch_size, 1, H, W)
    if anomaly_map.dim() == 4:
        batch_size = anomaly_map.shape[0]
        binary_masks_style1 = []
        binary_masks_style2 = []
        binary_masks_style3 = []
        
        for b in range(batch_size):
            # Extract single image from batch
            single_map = anomaly_map[b, 0]  # Shape: (H, W)
            single_binary_style1, single_binary_style2, single_binary_style3 = _create_contour_based_binary_mask_single(single_map, adaptive_threshold, anomaly_binary_threshold)
            binary_masks_style1.append(single_binary_style1)
            binary_masks_style2.append(single_binary_style2)
            binary_masks_style3.append(single_binary_style3)
        
        # Stack back into batch
        style1 = torch.stack(binary_masks_style1, dim=0).unsqueeze(1)  # Shape: (batch_size, 1, H, W)
        style2 = torch.stack(binary_masks_style2, dim=0).unsqueeze(1)  # Shape: (batch_size, 1, H, W)
        style3 = torch.stack(binary_masks_style3, dim=0).unsqueeze(1)  # Shape: (batch_size, 1, H, W)
        return style1, style2, style3
    else:
        # Handle single image case
        return _create_contour_based_binary_mask_single(anomaly_map, adaptive_threshold, anomaly_binary_threshold)


def _create_contour_based_binary_mask_single(anomaly_map: torch.Tensor, adaptive_threshold: float = 0.1, anomaly_binary_threshold: int = 5) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
    """
    Create three different binary masks for a single image based on different contour selection styles.
    
    Args:
        anomaly_map: Anomaly map tensor with shape (H, W) with values in [0, 1]
        adaptive_threshold: Threshold for adaptive contour selection (default: 0.1)
        anomaly_binary_threshold: Threshold value for binary conversion (0-255, default: 5)
                                 - Lower values create more white pixels
                                 - Higher values create fewer white pixels
        
    Returns:
        Tuple of three binary tensors with shape (H, W) where selected contour pixels are 1, others are 0:
        - style1: Top contours by sum (most significant)
        - style2: Statistical outliers (mean + threshold * std)
        - style3: Contours contributing significant portion of total
    """
    import cv2
    import numpy as np
    
    # Convert to numpy
    map_np = anomaly_map.cpu().numpy()
    
    # Ensure the map is 2D
    if map_np.ndim != 2:
        print(f"Warning: Expected 2D array, got shape {map_np.shape}")
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask
    
    # Handle negative values and ensure proper range
    map_np = np.clip(map_np, 0, 1)  # Clip to [0, 1] range
    
    # Convert to uint8 for contour detection (0-255 range)
    map_uint8 = (map_np * 255).astype(np.uint8)
    
    # Apply morphological operations to reduce noise
    kernel = np.ones((3, 3), np.uint8)
    map_uint8 = cv2.morphologyEx(map_uint8, cv2.MORPH_CLOSE, kernel)  # Close small holes
    map_uint8 = cv2.morphologyEx(map_uint8, cv2.MORPH_OPEN, kernel)   # Remove small noise
    
    # Convert to binary image for contour detection
    ## Use Otsu's method for automatic thresholding
    #_, binary_map = cv2.threshold(map_uint8, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    # Use anomaly_binary_threshold instead of Otsu's method
    _, binary_map = cv2.threshold(map_uint8, anomaly_binary_threshold, 255, cv2.THRESH_BINARY)
    
    # Check if the image is all zeros (no contours possible)
    if np.all(binary_map == 0):
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask
    
    try:
        # Find contours on the binary image
        contours, _ = cv2.findContours(binary_map, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    except cv2.error as e:
        print(f"OpenCV error in findContours: {e}")
        print(f"Binary map shape: {binary_map.shape}, dtype: {binary_map.dtype}")
        print(f"Binary map min: {binary_map.min()}, max: {binary_map.max()}")
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask
    
    if not contours:
        # No contours found, return all zeros
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask
    
    # Calculate contour statistics (sum of pixel values within each contour)
    contour_stats = []
    
    # Filter parameters to reduce noise
    min_contour_area = 10.0  # Minimum contour area to consider
    min_white_pixels = 5     # Minimum white pixels to consider
    
    for i, contour in enumerate(contours):
        # Create a mask for this contour
        contour_mask = np.zeros_like(binary_map)
        cv2.fillPoly(contour_mask, [contour], (255,))
        
        # Calculate sum of pixel values within this contour
        contour_white_pixels = np.sum(contour_mask > 0)
        contour_sum = np.sum(map_np * (contour_mask > 0))
        contour_area = cv2.contourArea(contour)
        
        # Filter out noise: skip contours that are too small
        if contour_area < min_contour_area or contour_white_pixels < min_white_pixels:
            #print(f"DEBUG: Skipping contour {i} - area: {contour_area}, white_pixels: {contour_white_pixels}")
            continue
            
        ##print(f"DEBUG: Contour {i} has {contour_white_pixels} white pixels")
        ##print(f"DEBUG: Contour {i} has {contour_sum} sum")
        ##print(f"DEBUG: Contour {i} has {contour_area} area")
        ##print(f"DEBUG: Contour {i} has {contour_sum / contour_area} sum/area")
        contour_stats.append({
            'index': i,
            'contour': contour,
            'sum': contour_sum,
            'area': contour_area,
            'white_pixels': contour_white_pixels
        })
    #print(f'DEBUG: length of contour_stats (after filtering): {len(contour_stats)}')
    
    # Sort by sum (descending)
    contour_stats.sort(key=lambda x: x['sum'], reverse=True)
    
    if not contour_stats:
        zero_mask = torch.zeros_like(anomaly_map)
        return zero_mask, zero_mask, zero_mask
    
    # Extract sums for adaptive selection
    sums = np.array([stat['sum'] for stat in contour_stats])
    
    # Calculate statistics for adaptive selection
    total_sum = np.sum(sums)
    max_sum = np.max(sums)
    mean_sum = np.mean(sums)
    std_sum = np.std(sums)
    
    #print(f"DEBUG: Total sum: {total_sum}, Max sum: {max_sum}, Mean sum: {mean_sum}, Std sum: {std_sum}")
    
    # Remove statistical outliers (contours with sums > mean + 2*std)
    outlier_threshold = mean_sum + 2 * std_sum
    filtered_contour_stats = [stat for stat in contour_stats if stat['sum'] <= outlier_threshold]
    
    if not filtered_contour_stats:
        # If all contours were outliers, keep the top 3
        filtered_contour_stats = contour_stats[:3]
    
    #print(f"DEBUG: Contours after outlier removal: {len(filtered_contour_stats)}")
    
    # Use filtered contours for selection
    contour_stats = filtered_contour_stats
    
    # Create binary masks for each style
    binary_mask_style1 = np.zeros_like(binary_map)
    binary_mask_style2 = np.zeros_like(binary_map)
    binary_mask_style3 = np.zeros_like(binary_map)
    
    for contour_info in contour_stats:
        cv2.fillPoly(binary_mask_style1, [contour_info['contour']], (255,))
    
    for contour_info in contour_stats:
        cv2.fillPoly(binary_mask_style2, [contour_info['contour']], (255,))
    
    for contour_info in contour_stats:
        cv2.fillPoly(binary_mask_style3, [contour_info['contour']], (255,))
    
    # Convert back to tensors and normalize to [0, 1]
    binary_tensor_style1 = torch.from_numpy(binary_mask_style1).float() / 255.0
    binary_tensor_style2 = torch.from_numpy(binary_mask_style2).float() / 255.0
    binary_tensor_style3 = torch.from_numpy(binary_mask_style3).float() / 255.0
    
    # Return all three styles
    return (binary_tensor_style1.to(anomaly_map.device), 
            binary_tensor_style2.to(anomaly_map.device), 
            binary_tensor_style3.to(anomaly_map.device))


def _to_numpy(
    t: torch.Tensor,
) -> np.ndarray:  # keep Images API compatibility
    """Detach, move to CPU and convert to ``numpy`` if ``t`` is a tensor."""
    return t.detach().cpu().numpy() if isinstance(t, torch.Tensor) else t


def _tensor_to_xlimage(arr, size: int) -> XLImage:
    if isinstance(arr, torch.Tensor):
        arr = arr.detach().cpu().numpy()
    arr = np.squeeze(arr)
    if arr.ndim == 2:
        arr = arr[..., None]
    if arr.ndim == 3 and arr.shape[0] in (1, 3, 4):
        arr = np.transpose(arr, (1, 2, 0))
    c = arr.shape[2]
    if c == 1:
        arr = np.repeat(arr, 3, axis=2)
    elif c == 4:
        # Split the image into 4 quadrants if c == 4 (e.g., 4-channel image)
        # We'll arrange the 4 channels as 2x2 grid: [0|1]
        #                                             [2|3]
        # h, w = arr.shape[0], h2, w2 = h // 2, w // 2
        # If the image is not square, just split in half along each axis
        # Each channel is a grayscale image, so we tile them
        q0 = arr[..., 0]
        q1 = arr[..., 1]
        q2 = arr[..., 2]
        q3 = arr[..., 3]
        # Stack as 2x2 grid
        top = np.concatenate([q0, q1], axis=1)
        bottom = np.concatenate([q2, q3], axis=1)
        arr = np.stack(
            [np.concatenate([top, bottom], axis=0)] * 3, axis=2
        )  # make 3 channels
    elif c != 3:
        raise ValueError(f"unsupported channels: {c}")
    if np.any(arr < 0):
        arr = ((np.clip(arr, -1, 1) + 1) / 2 * 255).astype(np.uint8)
    else:
        arr = (arr * 255).astype(np.uint8)

    buf = BytesIO()
    PILImage.fromarray(arr, mode="RGB").save(buf, format="PNG")
    buf.seek(0)
    img = XLImage(buf)
    img.width = img.height = size
    return img


def _write_row(ws, row_idx: int, rec: dict, size: int):
    scalars, embeds = [], []
    for col_idx, key in enumerate(rec.keys(), 1):
        kind, val = rec[key]
        if kind == "image":
            embeds.append((col_idx, _tensor_to_xlimage(val, size)))
            scalars.append("")
            ws.column_dimensions[get_column_letter(col_idx)].width = size // 8
        else:
            # Convert tuples and other non-serializable types to strings
            if isinstance(val, (tuple, list)):
                scalars.append(str(val))
            elif isinstance(val, (np.ndarray, torch.Tensor)):
                scalars.append(str(val.tolist() if hasattr(val, 'tolist') else val))
            else:
                scalars.append(val)
    ws.append(scalars)
    ws.row_dimensions[row_idx].height = size * 0.75
    for col_idx, img in embeds:
        ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")


def make_excel(
    records: List[Record],
    image_size: int,
    save_dir: str | Path = "report",
    save_filename: str | None = datetime.now().strftime("%y%m%d_%H%M%S"),
    max_rows_per_file: int = 50,
) -> List[Path]:
    """Create Excel report with all evaluation records and images.
    
    Args:
        records: List of evaluation records
        image_size: Size of images to embed
        save_dir: Directory to save Excel files
        save_filename: Base filename for the Excel files
        max_rows_per_file: Maximum number of rows per Excel file (default: 100)
    
    Returns:
        List of paths to created Excel files
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    # Header comes from the first record's keys (order preserved)
    header = list(records[0].keys())
    
    # Calculate how many files we need
    total_rows = len(records)
    num_files = (total_rows + max_rows_per_file - 1) // max_rows_per_file  # Ceiling division
    
    # Check if the last file would have fewer than max_rows_per_file rows
    # If so, reduce the number of files by 1 and append remaining rows to the previous file
    if num_files > 1:
        remaining_rows = total_rows % max_rows_per_file
        if remaining_rows > 0 and remaining_rows < max_rows_per_file:
            num_files -= 1
    
    created_files = []
    
    for file_index in range(num_files):
        # Calculate start and end indices for this file
        start_idx = file_index * max_rows_per_file
        
        # For the last file, include all remaining rows
        if file_index == num_files - 1:
            end_idx = total_rows
        else:
            end_idx = (file_index + 1) * max_rows_per_file
        
        # Get records for this file
        file_records = records[start_idx:end_idx]
        
        # Create workbook for this file
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            # Set worksheet title
            if num_files == 1:
                ws.title = "Report"
            else:
                ws.title = f"Report_Part{file_index + 1}"
            
            # Add header
            ws.append(header)
            
            # Add data rows
            for r, rec in enumerate(file_records, start=2):
                _write_row(ws, r, rec, image_size)
            
            # Set column widths
            for c in range(1, len(header) + 1):
                ws.column_dimensions[get_column_letter(c)].width = 18
        
        # Generate filename
        if num_files == 1:
            out_filename = f"report_{save_filename}.xlsx"
        else:
            out_filename = f"report_{save_filename}_part{file_index + 1:02d}.xlsx"
        
        out_path = save_dir / out_filename
        wb.save(out_path)
        created_files.append(out_path)
        
        print(f"Report part {file_index + 1}/{num_files} saved to {out_path} ({len(file_records)} rows)")
    
    if num_files > 1:
        print(f"\nTotal: Created {num_files} Excel files with {total_rows} total rows")
    
    return created_files


def draw_patch_rectangles_on_image(base_img, predicted_defective_set, ground_truth_defective, overlapping, patch_size=128, grid_thickness=1):
    """
    Draw patch rectangles (TP/FP/FN) on top of an image.
    Args:
        base_img: The image to draw on (np.uint8, HxWx3)
        predicted_defective: Set of (grid_row, grid_col) for predicted defective patches
        ground_truth_defective: Set of (grid_row, grid_col) for ground truth defective patches
        overlapping: Set of (grid_row, grid_col) where prediction and ground truth overlap
        patch_size: Size of patches
        grid_thickness: Thickness of the rectangle lines (default: 1)
    Returns:
        Image with rectangles drawn:
        - Yellow rectangles around predicted defective regions
        - Red rectangles around ground truth defective regions
        - Green rectangles where prediction and ground truth overlap
    """
    img = base_img.copy()
    # Draw predicted defective regions (yellow)
    for grid_row, grid_col in predicted_defective_set:
        x = grid_col * patch_size
        y = grid_row * patch_size
        cv2.rectangle(img, (x, y), (x + patch_size, y + patch_size), (255, 255, 0), grid_thickness)
    # Draw ground truth defective regions (red)
    for grid_row, grid_col in ground_truth_defective:
        x = grid_col * patch_size
        y = grid_row * patch_size
        cv2.rectangle(img, (x, y), (x + patch_size, y + patch_size), (255, 0, 0), grid_thickness)
    # Draw overlapping regions (green)
    for grid_row, grid_col in overlapping:
        x = grid_col * patch_size
        y = grid_row * patch_size
        cv2.rectangle(img, (x, y), (x + patch_size, y + patch_size), (0, 255, 0), grid_thickness)
    return img



# ---------------------------------------------------------------------------
# Functions
# ---------------------------------------------------------------------------

def evaluation(args):
    if os.path.exists("./models/config.json"):
        vae = cast(AutoencoderKL, AutoencoderKL.from_pretrained("./models", local_files_only=True)).to(
            device
        )
    else:
        vae = cast(AutoencoderKL, AutoencoderKL.from_pretrained(f"stabilityai/sd-vae-ft-{args.vae_type}")).to(
            device
        )
    vae.eval()
    try:
        if args.pretrained != "":
            # Handle relative paths by checking if they exist relative to current directory
            # or parent directory (project root)
            ckpt = args.pretrained
            if not os.path.exists(ckpt):
                # Try relative to parent directory (project root)
                parent_ckpt = os.path.join("..", ckpt)
                if os.path.exists(parent_ckpt):
                    ckpt = parent_ckpt
                else:
                    raise FileNotFoundError(f"Checkpoint not found: {args.pretrained}")
        else:
            path = f"./DeCo-Diff_{args.dataset}_{args.object_class}_{args.model_size}_{args.center_size}"
            try:
                ckpt = sorted(glob(f"{path}/last.pt"))[-1]
            except (IndexError, FileNotFoundError):
                ckpt = sorted(glob(f"{path}/*/last.pt"))[-1]
    except (IndexError, FileNotFoundError, OSError) as e:
        raise Exception(f"Please provide the model's pretrained path using --pretrained. Error: {e}")

    latent_size = int(args.center_size) // 8
    model = UNET_models[args.model_size](latent_size=latent_size)

    state_dict = torch.load(ckpt)["model"]
    print(model.load_state_dict(state_dict))
    model.eval()  # important!
    model.cuda()
    print("model loaded")

    print("==" * 30)
    print("Starting Evaluation...")
    print("==" * 30)
    diffusion = create_diffusion(
        f"ddim{args.reverse_steps}",
        predict_deviation=True,
        sigma_small=False,
        predict_xstart=False,
        diffusion_steps=1000,
    )

    evaluation_annotated_images(args, diffusion, model, vae)

def evaluation_annotated_images(args, diffusion, model, vae):
    """Evaluate images with JSON annotations for defective regions."""
    # Create base evaluator
    evaluator = BaseEvaluator(args, diffusion, model, vae)
    
    # Create dataset for annotated images
    dataset = AnnotatedImageDataset(
        annotation_dir=args.annotation_dir,
        patch_size=args.patch_size,
        transform=evaluator.transform,
        object_class=args.object_class,
    )
    
    loader = evaluator._get_dataloader(dataset, batch_size=8)
    
    # Create checkpoint manager
    checkpoint_manager = CheckpointManager(args.results_dir, args.annotation_dir, args.force_rerun)
    
    # Print checkpoint status
    print("=== Checkpoint Status ===")
    checkpoint_manager.print_checkpoint_status()
    print("========================")
    
    # Process patches with checkpoint functionality
    records, optional_records = process_split_irregular_with_checkpoint(
        loader,
        args.split,
        diffusion,
        model,
        vae,
        args.reverse_steps,
        args.patch_size,
        args.batch_num,
        device,
        args.anomaly_binary_threshold,
        args.anomaly_pixel_num_threshold,
        0.1,  # adaptive_threshold
        args.enable_epoch_stats,  # Pass only the boolean flag
        args.enable_excel_report,
        args.enable_save_optional_image_results,
        checkpoint_manager,  # Pass checkpoint manager
        args.patch_size
    )
    
    # Create Excel report (if enabled)
    if hasattr(args, 'enable_excel_report') and args.enable_excel_report:
        print("Creating Excel report...")
        # Use optional_records for Excel if available, otherwise use regular records
        excel_records = optional_records if optional_records else records
        excel_path = make_excel(
            records=excel_records,
            image_size=args.patch_size,
            save_dir=checkpoint_manager.evaluation_results_dir,
            save_filename=f"report_{datetime.now().strftime('%y%m%d_%H%M%S')}"
        )
        print(f"Excel report saved to: {excel_path}")
    else:
        print("Skipping Excel report generation (disabled)")
    
    print("==" * 30)
    # Compute confusion matrix and accuracy
    compute_confusion_matrix_and_accuracy(args.annotation_dir, checkpoint_manager.evaluation_results_dir)

def collect_records_for_params(
    *,
    param_name: str,
    param_values,
    split: str,
    object_class: str,
    rootdir: str,
    transform,
    anomaly_class: str,
    image_size: int,
    center_size: int,
    center_crop: bool,
    process_split_fn,
    diffusion,
    model,
    vae,
    reverse_steps,
    batch_num,
    device=None,
    dataset_class,
    split_csv_path: str | None = None,
):
    common_args = dict(
        mode=split,
        object_class=object_class,
        rootdir=rootdir,
        transform=transform,
        anomaly_class=anomaly_class,
        image_size=image_size,
        center_size=center_size,
        center_crop=center_crop,
        split_csv_path=split_csv_path,
    )
    all_records = []
    for val in param_values:
        print(f"Processing {param_name} = {val}")
        kwargs = common_args.copy()
        kwargs[param_name] = val
        dataset = dataset_class(**kwargs)
        # Windows-friendly DataLoader configuration
        import platform
        if platform.system() == 'Windows':
            num_workers = 0  # Single-threaded on Windows to avoid hanging
        else:
            num_workers = 4
        
        loader = DataLoader(
            dataset, batch_size=8, shuffle=False, num_workers=num_workers, drop_last=False,
            persistent_workers=False if num_workers == 0 else True
        )
        records = process_split_fn(
            loader,
            split,
            diffusion,
            model,
            vae,
            reverse_steps,
            center_size,
            batch_num,
            device,
        )

        kwargs["scratch"] = True
        dataset_defect = dataset_class(**kwargs)
        loader_defect = DataLoader(
            dataset_defect, batch_size=8, shuffle=False, num_workers=num_workers, drop_last=False,
            persistent_workers=False if num_workers == 0 else True
        )
        records_defect = process_split_fn(
            loader_defect,
            split,
            diffusion,
            model,
            vae,
            reverse_steps,
            center_size,
            batch_num,
            device,
        )
        all_records.append((records, records_defect))
    return all_records


def compute_y_true_y_score(all_records):
    """
    For each param value, compute y_true and y_score arrays from records.
    Returns: list of (y_true, y_score) tuples, one for each param value.
    """
    y_true_score_list = []
    for i, (records, records_defect) in enumerate(all_records):
        y_true = []
        y_score = []
        for rec in records:
            y_true.append(0)
            mask = (
                rec["anomaly_map_arithmetic_binary"][1]
                if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
                else rec["anomaly_map_arithmetic_binary"]
            )
            num_white = np.sum(mask == 1)
            y_score.append(num_white)
        for rec in records_defect:
            y_true.append(1)
            mask = (
                rec["anomaly_map_arithmetic_binary"][1]
                if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
                else rec["anomaly_map_arithmetic_binary"]
            )
            num_white = np.sum(mask == 1)
            y_score.append(num_white)
        y_score = np.array(y_score)
        y_true = np.array(y_true)
        y_true_score_list.append((y_true, y_score))
    return y_true_score_list


def compute_metrics_from_y_true_y_score(y_true_score_list):
    """
    For each (y_true, y_score), compute accuracy, threshold, and ROC stats.
    Returns: accuracies, thresholds, and a dict of lists for each ROC metric.
    """
    accuracies = []
    fpr_list = []
    tpr_list = []
    thresholds_list = []
    best_thresholds = []
    best_idxs = []
    aucs = []
    y_trues = []
    y_preds = []
    y_scores = []

    for i, (y_true, y_score) in enumerate(y_true_score_list):
        fpr, tpr, thresholds_, best_threshold, best_idx, auc_score = compute_roc_stats(
            y_true, y_score
        )
        y_pred = (y_score >= best_threshold).astype(int)
        y_preds.append(y_pred)
        accuracy = np.mean(y_pred == y_true)
        accuracies.append(accuracy)
        fpr_list.append(fpr)
        tpr_list.append(tpr)
        # thresholds_list.append(thresholds_)
        best_thresholds.append(best_threshold)
        best_idxs.append(best_idx)
        aucs.append(auc_score)
        y_trues.append(y_true)
        y_scores.append(y_score)
        print(f"Accuracy {accuracy:.4f} (threshold={best_threshold})")

    roc_stats = {
        "fpr": fpr_list,
        "tpr": tpr_list,
        # "thresholds": thresholds_list,
        "best_threshold": best_thresholds,
        "best_idx": best_idxs,
        "auc": aucs,
        "y_true": y_trues,
        "y_pred": y_preds,
        "y_score": y_scores,
        "accuracies": accuracies,
    }
    return roc_stats


def plot_accuracy_results(
    param_values,
    accuracies,
    param_name: str,
    save_dir="accuracy_vs_param",
    save_filename=datetime.now().strftime("%y%m%d_%H%M%S"),
    title=None,
    xlabel=None,
    ylabel="Accuracy",
    grid=True,
    marker="o",
    **plot_kwargs,
):
    """
    Plot accuracy results with customizable parameters.

    Args:
        param_values: List of parameter values
        accuracies: List of corresponding accuracies
        param_name: Name of the parameter being varied
        save_dir: Directory to save the plot
        save_filename: Filename for saving the plot
        title: Custom title for the plot
        xlabel: Custom x-axis label
        ylabel: Custom y-axis label
        grid: Whether to show grid
        marker: Marker style for the plot
        **plot_kwargs: Additional plotting parameters
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    plt.figure()
    plt.plot(param_values, accuracies, marker=marker, **plot_kwargs)
    plt.xlabel(xlabel or param_name)
    plt.ylabel(ylabel)
    plt.title(title or f"Accuracy vs {param_name.capitalize()} (synthetic defect)")
    plt.ylim(0.5, 1.0)
    plt.grid(grid)
    out_path = os.path.join(save_dir, f"accuracy_vs_{param_name}_{save_filename}.png")
    plt.savefig(out_path)
    print(f"Accuracy vs {param_name} saved to {out_path}")
    plt.close()


def save_perturbation_results(
    param_name: str,
    roc_stats: dict,
    param_values: list,
    save_dir: str,
):
    """
    Save perturbation experiment data (roc_stats and param_values) to a specified folder in JSON format.

    Args:
        param_name: Name of the perturbation parameter
        roc_stats: Dictionary containing ROC statistics
        param_values: List of parameter values
        save_dir: Directory to save the results
    """
    save_dir_path = Path(save_dir).expanduser()
    save_dir_path.mkdir(parents=True, exist_ok=True)

    # Convert numpy arrays to lists for JSON serialization
    def convert_for_json(obj):
        if hasattr(obj, 'tolist'):  # numpy arrays
            return obj.tolist()
        elif isinstance(obj, list):
            return [convert_for_json(item) for item in obj]
        elif isinstance(obj, dict):
            return {key: convert_for_json(value) for key, value in obj.items()}
        else:
            return obj

    # Convert roc_stats to JSON-serializable format
    roc_stats_json = convert_for_json(roc_stats)
    
    # Convert param_values to JSON-serializable format
    param_values_json = convert_for_json(param_values)

    # Save both roc_stats and param_values in a single JSON file
    results_data = {
        "param_name": param_name,
        "param_values": param_values_json,
        "roc_stats": roc_stats_json
    }
    
    json_path = os.path.join(save_dir, f"{param_name}_results.json")
    with open(json_path, "w") as f:
        json.dump(results_data, f, indent=2)

    print(f"Perturbation results saved to {json_path}")


def compute_roc_stats(y_true, y_score):
    """
    Compute ROC statistics and best threshold using Youden's J statistic.
    Returns: fpr, tpr, thresholds, best_threshold, best_idx, auc_score
    """
    fpr, tpr, thresholds = roc_curve(y_true, y_score)
    youden_j = tpr - fpr
    best_idx = np.argmax(youden_j)
    best_threshold = thresholds[best_idx]
    auc_score = auc(fpr, tpr)
    return fpr, tpr, thresholds, best_threshold, best_idx, auc_score


def compute_confusion_matrix_and_accuracy(annotation_dir, evaluation_results_dir):
    import glob
    from collections import Counter
    import numpy as np
    import os
    import json

    # Check if evaluation results directory exists
    if not os.path.exists(evaluation_results_dir):
        print(f"Warning: Evaluation results directory {evaluation_results_dir} does not exist")
        return

    # Find all evaluation result files
    eval_files = glob.glob(os.path.join(evaluation_results_dir, '*__evaluation.json'))
    
    if not eval_files:
        print(f"Warning: No evaluation result files found in {evaluation_results_dir}")
        return
    all_TP = all_FP = all_FN = all_TN = 0
    for eval_file in eval_files:
        try:
            with open(eval_file, 'r') as f:
                eval_data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Warning: Corrupted JSON file {eval_file}: {e}")
            print(f"Skipping this file and continuing...")
            continue
        except Exception as e:
            # handle other exceptions
            pass
        image_path = eval_data['image_path']
        # Handle new format with patch_analysis
        if 'patch_analysis' in eval_data:
            # New format: extract defective patches from patch_analysis
            predicted = set()
            for patch in eval_data['patch_analysis']:
                if patch['status'] == "TP" or patch['status'] == "FP":
                    predicted.add((patch['grid_row'], patch['grid_col']))
        else:
            # Old format: direct defective_patches list
            predicted = set(tuple(x) for x in eval_data['defective_patches'])
        grid_size = eval_data['grid_size']
        annotation_file = os.path.join(annotation_dir, f"{path_to_safe_filename(image_path)}__annotations.json")
        if not os.path.exists(annotation_file):
            print(f"Warning: No annotation for {image_path}")
            continue
        try:
            with open(annotation_file, 'r') as f:
                anno_data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"Warning: Corrupted annotation file {annotation_file}: {e}")
            print(f"Skipping this file and continuing...")
            continue
        except Exception as e:
            print(f"Warning: Error reading annotation file {annotation_file}: {e}")
            print(f"Skipping this file and continuing...")
            continue
        gt = set(tuple(x) for x in anno_data['defective_patches'])
        # Get all possible grid cells including edge cells
        img = PILImage.open(image_path)
        h, w = img.height, img.width
        n_rows = (h + grid_size - 1) // grid_size  # Ceiling division
        n_cols = (w + grid_size - 1) // grid_size   # Ceiling division
        all_cells = set((r, c) for r in range(n_rows) for c in range(n_cols))
        for cell in all_cells:
            pred = cell in predicted
            truth = cell in gt
            if pred and truth:
                all_TP += 1
            elif pred and not truth:
                all_FP += 1
            elif not pred and truth:
                all_FN += 1
            else:
                all_TN += 1
    total = all_TP + all_FP + all_FN + all_TN
    accuracy = (all_TP + all_TN) / total if total > 0 else 0
    
    # Calculate additional metrics
    precision = all_TP / (all_TP + all_FP) if (all_TP + all_FP) > 0 else 0
    recall = all_TP / (all_TP + all_FN) if (all_TP + all_FN) > 0 else 0
    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    
    print("Confusion Matrix (patch-level):")
    print(f"TP: {all_TP}, FP: {all_FP}, FN: {all_FN}, TN: {all_TN}")
    print(f"Accuracy: {accuracy:.4f}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall: {recall:.4f}")
    print(f"F1-Score: {f1_score:.4f}")
    
    # Create confusion matrix visualization
    cm = np.array([[all_TP, all_FN], [all_FP, all_TN]])
    
    plt.figure(figsize=(8, 6))
    plt.imshow(cm, interpolation='nearest', cmap='Blues')
    plt.title('Confusion Matrix (Patch-level)', fontsize=16, fontweight='bold')
    plt.colorbar()
    
    # Add text annotations
    thresh = cm.max() / 2.
    for i in range(2):
        for j in range(2):
            plt.text(j, i, format(cm[i, j], 'd'),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black",
                    fontsize=14, fontweight='bold')
    
    # Set labels
    tick_marks = np.arange(2)
    plt.xticks(tick_marks, ['Defective', 'Normal'], fontsize=12)
    plt.yticks(tick_marks, ['Defective', 'Normal'], fontsize=12)
    plt.ylabel('True Label', fontsize=12)
    plt.xlabel('Predicted Label', fontsize=12)
    
    # Add metrics text
    metrics_text = f'Accuracy: {accuracy:.4f}\nPrecision: {precision:.4f}\nRecall: {recall:.4f}\nF1-Score: {f1_score:.4f}'
    plt.figtext(0.02, 0.02, metrics_text, fontsize=10, bbox=dict(boxstyle="round,pad=0.3", facecolor="lightgray"))
    
    plt.tight_layout()
    
    # Save the confusion matrix plot
    cm_plot_path = os.path.join(evaluation_results_dir, "confusion_matrix.png")
    plt.savefig(cm_plot_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"Confusion matrix plot saved to: {cm_plot_path}")
    
    # Save detailed results to file
    result = {
        "TP": all_TP,
        "FP": all_FP,
        "FN": all_FN,
        "TN": all_TN,
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1_score": f1_score,
        "total_patches": total
    }
    with open(os.path.join(evaluation_results_dir, "confusion_matrix.json"), "w") as f:
        json.dump(result, f, indent=2)

# ---------------------------------------------------------------------------
# Checkpoint/Resume System
# ---------------------------------------------------------------------------

class CheckpointManager:
    """Manages checkpoint/resume functionality for evaluation."""
    
    def __init__(self, results_dir: str, annotation_dir: str | None = None, force_rerun: bool = False):
        self.results_dir = results_dir
        self.annotation_dir = annotation_dir
        self.force_rerun = force_rerun
        
        # Extract base name without timestamp for consistent checkpoint location
        base_name = self._extract_base_name(results_dir)
        self.base_checkpoint_dir = os.path.join(os.path.dirname(results_dir), f"{base_name}_checkpoints")
        
        # Create checkpoint directory
        os.makedirs(self.base_checkpoint_dir, exist_ok=True)
        
        self.checkpoint_file = os.path.join(self.base_checkpoint_dir, "evaluation_checkpoint.json")

        # Create results directories
        self.evaluation_results_dir = os.path.join(results_dir, "evaluation_results")
        self.marked_images_dir = os.path.join(results_dir, "marked_images")
        os.makedirs(self.evaluation_results_dir, exist_ok=True)
        os.makedirs(self.marked_images_dir, exist_ok=True)
        
        # Clear checkpoint files if force rerun is enabled
        if self.force_rerun:
            self.clear_checkpoint_files()
            print("Force rerun enabled: Cleared existing checkpoint files")
        
        # Cache for processed images to avoid frequent file I/O
        self._processed_images_cache = None
        self._cache_timestamp = 0
    
    def _extract_base_name(self, results_dir: str) -> str:
        """Extract base name from timestamped directory."""
        # Handle patterns like "test_name_250708_143022" -> "test_name"
        dir_name = os.path.basename(results_dir)
        
        # Try to find the last underscore followed by timestamp pattern
        import re
        # Pattern for timestamp: YYMMDD_HHMMSS
        timestamp_pattern = r'_\d{6}_\d{6}$'
        match = re.search(timestamp_pattern, dir_name)
        
        if match:
            # Remove the timestamp part
            base_name = dir_name[:match.start()]
            return base_name
        else:
            # If no timestamp pattern found, use the original name
            return dir_name
    
    def find_latest_checkpoint(self) -> str:
        """Find the latest checkpoint file from the specific checkpoint directory."""
        if os.path.exists(self.checkpoint_file):
            return self.checkpoint_file
        
        # Look for checkpoint in the specific directory based on base name
        # e.g., if results_dir is "test_250711_162955", look for "test_checkpoints"
        base_name = self._extract_base_name(self.results_dir)
        specific_checkpoint_dir = os.path.join(os.path.dirname(self.results_dir), f"{base_name}_checkpoints")
        specific_checkpoint_file = os.path.join(specific_checkpoint_dir, "evaluation_checkpoint.json")
        
        if os.path.exists(specific_checkpoint_file):
            print(f"Found existing checkpoint: {specific_checkpoint_file}")
            return specific_checkpoint_file
        
        return self.checkpoint_file
    
    def get_checkpoint_data(self) -> dict:
        """Load checkpoint data if it exists."""
        checkpoint_file = self.find_latest_checkpoint()
        if os.path.exists(checkpoint_file):
            try:
                with open(checkpoint_file, 'r') as f:
                    return json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                pass
        return {"current_image_index": 0, "processed_images": []}
    
    def _get_processed_images_with_cache(self) -> set:
        """Get processed images with caching to avoid frequent file I/O."""
        # Check if cache is valid (file hasn't been modified since last cache)
        checkpoint_file = self.find_latest_checkpoint()
        if os.path.exists(checkpoint_file):
            current_timestamp = os.path.getmtime(checkpoint_file)
            if (self._processed_images_cache is not None and 
                self._cache_timestamp == current_timestamp):
                return self._processed_images_cache
        
        # Load from file and update cache
        checkpoint_data = self.get_checkpoint_data()
        processed_images = set(checkpoint_data.get("processed_images", []))
        
        # Update cache
        self._processed_images_cache = processed_images
        self._cache_timestamp = current_timestamp if os.path.exists(checkpoint_file) else 0
        
        return processed_images
    
    def save_checkpoint(self, current_image_index: int, processed_images: list):
        """Save current progress to checkpoint file."""
        # Get existing checkpoint data to preserve processed_images
        existing_data = self.get_checkpoint_data()
        
        # Merge processed images from both sources (optimized)
        existing_processed = set(existing_data.get("processed_images", []))
        new_processed = set(processed_images)
        
        # Only merge if there are actually new images to add
        if new_processed - existing_processed:
            merged_processed = list(existing_processed.union(new_processed))
        else:
            merged_processed = list(existing_processed)
        
        # Sort for consistent, predictable order (only when writing to file)
        merged_processed.sort()
        
        checkpoint_data = {
            "current_image_index": current_image_index,
            "processed_images": merged_processed,
            "timestamp": datetime.now().isoformat()
        }
        
        with open(self.checkpoint_file, 'w') as f:
            json.dump(checkpoint_data, f, indent=2)
        
        # Invalidate cache since we just wrote to the file
        self._processed_images_cache = None
    
    def get_processed_images(self) -> set:
        """Get set of already processed images from checkpoint file."""
        return self._get_processed_images_with_cache()
    
    def mark_image_processed(self, image_path: str):
        """Mark an image as processed by updating the checkpoint file."""
        # Use cache to avoid reading file again
        processed_images = self._get_processed_images_with_cache()
        
        # Only update if image is not already processed
        if image_path not in processed_images:
            processed_images.add(image_path)
            
            # Get current checkpoint data
            checkpoint_data = self.get_checkpoint_data()
            # Sort for consistent, predictable order (only when writing to file)
            checkpoint_data["processed_images"] = sorted(processed_images)
            checkpoint_data["timestamp"] = datetime.now().isoformat()
            
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f, indent=2)
            
            # Update cache
            self._processed_images_cache = processed_images
            self._cache_timestamp = os.path.getmtime(self.checkpoint_file)
    
    def is_image_processed(self, image_path: str) -> bool:
        """Check if an image has been fully processed."""
        # If force rerun is enabled, no image is considered processed
        if self.force_rerun:
            return False
        return image_path in self.get_processed_images()
    
    def get_resume_info(self, all_image_paths: list) -> tuple[int, list]:
        """Get resume information for evaluation."""
        # If force rerun is enabled, start from the beginning
        if self.force_rerun:
            print("Force rerun enabled: Starting from the beginning")
            return 0, []
        
        checkpoint_data = self.get_checkpoint_data()
        processed_images = self.get_processed_images()
        
        # Find the first unprocessed image
        current_index = 0
        for i, image_path in enumerate(all_image_paths):
            if image_path not in processed_images:
                current_index = i
                break
        else:
            # All images processed
            current_index = len(all_image_paths)
        
        return current_index, sorted(processed_images)
    
    def cleanup_checkpoint(self):
        """Clean up checkpoint files after successful completion."""
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
        
        # Also clean up the checkpoint directory if it's empty
        try:
            if os.path.exists(self.base_checkpoint_dir) and not os.listdir(self.base_checkpoint_dir):
                os.rmdir(self.base_checkpoint_dir)
        except OSError:
            pass  # Directory not empty or already removed

    def clear_checkpoint_files(self):
        """Clear all checkpoint files to force a fresh start."""
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
            print(f"Removed checkpoint file: {self.checkpoint_file}")
        
        # Clear cache
        self._processed_images_cache = None
        self._cache_timestamp = 0
    
    def batch_mark_images_processed(self, image_paths: list):
        """Mark multiple images as processed in a single operation for better performance."""
        if not image_paths:
            return
        
        # Use cache to avoid reading file again
        processed_images = self._get_processed_images_with_cache()
        
        # Find new images to add
        new_images = [img for img in image_paths if img not in processed_images]
        
        if new_images:
            # Add new images
            processed_images.update(new_images)
            
            # Get current checkpoint data
            checkpoint_data = self.get_checkpoint_data()
            # Sort for consistent, predictable order (only when writing to file)
            checkpoint_data["processed_images"] = sorted(processed_images)
            checkpoint_data["timestamp"] = datetime.now().isoformat()
            
            with open(self.checkpoint_file, 'w') as f:
                json.dump(checkpoint_data, f, indent=2)
            
            # Update cache
            self._processed_images_cache = processed_images
            self._cache_timestamp = os.path.getmtime(self.checkpoint_file)
    
    def print_checkpoint_status(self):
        """Print current checkpoint status for debugging."""
        checkpoint_file = self.find_latest_checkpoint()
        
        print(f"Checkpoint directory: {self.base_checkpoint_dir}")
        print(f"Current checkpoint file: {checkpoint_file}")
        
        if os.path.exists(checkpoint_file):
            try:
                with open(checkpoint_file, 'r') as f:
                    data = json.load(f)
                    print(f"Checkpoint timestamp: {data.get('timestamp', 'unknown')}")
                    print(f"Current image index: {data.get('current_image_index', 0)}")
                    print(f"Processed images count: {len(data.get('processed_images', []))}")
                    if data.get('processed_images'):
                        print(f"Last processed: {data['processed_images'][-1]}")
            except Exception as e:
                print(f"Error reading checkpoint: {e}")
        else:
            print("No checkpoint file found")


def save_image_results_from_records(checkpoint_manager: CheckpointManager, image_path: str, 
                                  image_records: list, 
                                  predicted_defective_set: set, ground_truth_defective: set, overlapping: set,
                                  enable_save_optional_image_results: bool = False, patch_size: int = 256):
    """Save all results for a single image immediately using records."""
    safe_name = path_to_safe_filename(image_path)
    
    # Load original image
    original_img = np.array(PILImage.open(image_path).convert('RGB'))
    h, w, _ = original_img.shape
    
    # Save marked image (always saved)
    marked_img = draw_patch_rectangles_on_image(
        original_img, predicted_defective_set, ground_truth_defective, overlapping, patch_size=patch_size, grid_thickness=1
    )
    marked_path = os.path.join(checkpoint_manager.marked_images_dir, f"{safe_name}__marked.png")
    PILImage.fromarray(marked_img).save(marked_path)
    
    # Initialize anomaly maps based on flag
    anomaly_maps = {
        'required': {
            'arithmetic': np.zeros((h, w), dtype=np.float32),
            'arithmetic_binary': np.zeros((h, w), dtype=np.float32),
            'geometric': np.zeros((h, w), dtype=np.float32),
            'geometric_binary': np.zeros((h, w), dtype=np.float32),
        }
    }
    
    # Initialize optional maps only when flag is enabled
    if enable_save_optional_image_results:
        anomaly_maps['optional'] = {
            'arithmetic_binary_style1': np.zeros((h, w), dtype=np.float32),
            'arithmetic_binary_style2': np.zeros((h, w), dtype=np.float32),
            'arithmetic_binary_style3': np.zeros((h, w), dtype=np.float32),
            'geometric_binary_style1': np.zeros((h, w), dtype=np.float32),
            'geometric_binary_style2': np.zeros((h, w), dtype=np.float32),
            'geometric_binary_style3': np.zeros((h, w), dtype=np.float32),
        }
    
    for record in image_records:
        # Extract coordinates from record
        x_coord, y_coord = record["patch_coords"][1]
        
        # Calculate actual patch dimensions for this position
        patch_width = min(patch_size, w - x_coord)
        patch_height = min(patch_size, h - y_coord)
        
        # Extract required anomaly maps from record
        patch_arithmetic = record["anomaly_map_arithmetic"][1]
        patch_arithmetic_binary = record["anomaly_map_arithmetic_binary"][1]
        patch_geometric = record["anomaly_map_geometric"][1]
        patch_geometric_binary = record["anomaly_map_geometric_binary"][1]
        
        # Extract patch regions
        patch_regions = {
            'arithmetic': patch_arithmetic.squeeze()[:patch_height, :patch_width],
            'arithmetic_binary': patch_arithmetic_binary.squeeze()[:patch_height, :patch_width],
            'geometric': patch_geometric.squeeze()[:patch_height, :patch_width],
            'geometric_binary': patch_geometric_binary.squeeze()[:patch_height, :patch_width],
        }
        
        # Extract optional patch regions if flag is enabled
        if enable_save_optional_image_results:
            # Use default binary maps as fallback for style fields
            patch_regions.update({
                'arithmetic_binary_style1': record.get("anomaly_map_arithmetic_binary_style1", [None, patch_arithmetic_binary])[1].squeeze()[:patch_height, :patch_width],
                'arithmetic_binary_style2': record.get("anomaly_map_arithmetic_binary_style2", [None, patch_arithmetic_binary])[1].squeeze()[:patch_height, :patch_width],
                'arithmetic_binary_style3': record.get("anomaly_map_arithmetic_binary_style3", [None, patch_arithmetic_binary])[1].squeeze()[:patch_height, :patch_width],
                'geometric_binary_style1': record.get("anomaly_map_geometric_binary_style1", [None, patch_geometric_binary])[1].squeeze()[:patch_height, :patch_width],
                'geometric_binary_style2': record.get("anomaly_map_geometric_binary_style2", [None, patch_geometric_binary])[1].squeeze()[:patch_height, :patch_width],
                'geometric_binary_style3': record.get("anomaly_map_geometric_binary_style3", [None, patch_geometric_binary])[1].squeeze()[:patch_height, :patch_width],
            })
        
        # Assign required regions to anomaly maps
        for map_name, region in patch_regions.items():
            if map_name in anomaly_maps['required']:
                anomaly_maps['required'][map_name][y_coord:y_coord+patch_height, x_coord:x_coord+patch_width] = region
            elif enable_save_optional_image_results and map_name in anomaly_maps['optional']:
                anomaly_maps['optional'][map_name][y_coord:y_coord+patch_height, x_coord:x_coord+patch_width] = region
    
    # Define image configurations
    image_configs = {
        'required': [
            (anomaly_maps['required']['arithmetic'], "am_ar", False),
            (anomaly_maps['required']['arithmetic_binary'], "am_ar_bin", True),
            (anomaly_maps['required']['geometric'], "am_ge", False),
            (anomaly_maps['required']['geometric_binary'], "am_ge_bin", True),
        ]
    }
    
    # Add optional configurations if flag is enabled
    if enable_save_optional_image_results:
        image_configs['optional'] = [
            (anomaly_maps['optional']['arithmetic_binary_style1'], "am_ar_bin_st1", True),
            (anomaly_maps['optional']['arithmetic_binary_style2'], "am_ar_bin_st2", True),
            (anomaly_maps['optional']['arithmetic_binary_style3'], "am_ar_bin_st3", True),
            (anomaly_maps['optional']['geometric_binary_style1'], "am_ge_bin_st1", True),
            (anomaly_maps['optional']['geometric_binary_style2'], "am_ge_bin_st2", True),
            (anomaly_maps['optional']['geometric_binary_style3'], "am_ge_bin_st3", True),
        ]
    
    # Save all configured images
    for _config_type, configs in image_configs.items():
        for anomaly_map, suffix, is_binary in configs:
            # Save anomaly map image
            anomaly_map_img = ImageProcessor.create_anomaly_map_image(
                anomaly_map, patch_size=patch_size, add_grid=True, 
                predicted_defective_set=predicted_defective_set, ground_truth_defective=ground_truth_defective, 
                overlapping=overlapping, is_binary=is_binary
            )
            anomaly_map_path = os.path.join(checkpoint_manager.marked_images_dir, f"{safe_name}__{suffix}.png")
            PILImage.fromarray(anomaly_map_img).save(anomaly_map_path)
            
            # Save overlay image
            overlay_img = ImageProcessor.create_anomaly_overlay(original_img, anomaly_map, alpha=0.8, is_binary=is_binary)
            overlay_path = os.path.join(checkpoint_manager.marked_images_dir, f"{safe_name}__ao_{suffix}.png")
            PILImage.fromarray(overlay_img).save(overlay_path)
            
            # Save marked overlay image
            marked_overlay_img = draw_patch_rectangles_on_image(overlay_img, predicted_defective_set, ground_truth_defective, overlapping, patch_size=patch_size, grid_thickness=1)
            marked_overlay_path = os.path.join(checkpoint_manager.marked_images_dir, f"{safe_name}__mo_{suffix}.png")
            PILImage.fromarray(marked_overlay_img).save(marked_overlay_path)
    
    # Save evaluation results
    patch_analysis = []
    for record in image_records:
        x, y = record["patch_coords"][1]
        anomaly_map = record["anomaly_map_arithmetic_binary"][1]
        anomaly_pixels = int(np.sum(anomaly_map))
        grid_row = y // patch_size
        grid_col = x // patch_size
        patch_analysis.append({
            "grid_row": grid_row,
            "grid_col": grid_col,
            "anomaly_max": record["anomaly_max"][1],
            "anomaly_pixels":record["anomaly_pixels"][1],
            "status": record["status"][1]
        })
    
    result_filename = f"{safe_name}__evaluation.json"
    result_path = os.path.join(checkpoint_manager.evaluation_results_dir, result_filename)
    evaluation_result = {
        "image_path": image_path,
        "patch_analysis": patch_analysis,
        "grid_size": patch_size
    }
    with open(result_path, 'w') as f:
        json.dump(evaluation_result, f, indent=2)
    
    #print(f"Saved all results for image: {image_path}")


def process_split_irregular_with_checkpoint(
    dataloader,
    split: str,
    diffusion,
    model,
    vae,
    reverse_steps: int,
    center_size: int,
    batch_num: int,
    device: torch.device = torch.device("cpu"),
    anomaly_binary_threshold: int = 0,
    anomaly_pixel_num_threshold: int = 0,
    adaptive_threshold: float = 0.1,
    enable_epoch_stats: bool = False,
    enable_excel_report: bool = False,
    enable_save_optional_image_results: bool = False,
    checkpoint_manager: CheckpointManager | None = None,
    patch_size: int = 256
) -> tuple[list[Record], list[Record]]:
    """Run evaluation with checkpoint/resume functionality."""
    # Initialize evaluation metrics
    if enable_epoch_stats:
        metrics = EvaluationMetrics()

    results: list[Record] = []
    optional_results: list[Record] = []  # Store optional results separately

    # Get all unique image paths for checkpoint management
    all_image_paths = list(set([item[3] for item in dataloader.dataset]))
    all_image_paths.sort()  # Ensure consistent ordering

    # Build a mapping from image_path to number of patches for that image
    from collections import defaultdict
    image_patch_counts = defaultdict(int)
    for item in dataloader.dataset:
        image_patch_counts[item[3]] += 1

    # Get resume information
    if checkpoint_manager:
        current_image_index, processed_images = checkpoint_manager.get_resume_info(all_image_paths)
        print(f"Resuming from image {current_image_index}/{len(all_image_paths)}")
        print(f"Already processed: {len(processed_images)} images")
    else:
        current_image_index = 0
        processed_images = []

    # Memory optimization: Clear cache periodically
    if torch.cuda.is_available():
        torch.cuda.empty_cache()

    idx = -1  # Ensure idx is always defined for exception handling
    try:
        for idx, (x, seg, object_cls, image_paths, anomaly_classes, patch_coords) in enumerate(
            tqdm(dataloader, desc=f"{split} split")
        ):
            if idx >= batch_num:
                break

            # Build ground_truth_map for all unique image_paths in this batch
            ground_truth_map = {}
            for unique_image_path in set(image_paths):
                annotation_filename = f"{path_to_safe_filename(unique_image_path)}__annotations.json"
                if checkpoint_manager and checkpoint_manager.annotation_dir:
                    annotation_path = os.path.join(checkpoint_manager.annotation_dir, annotation_filename)
                else:
                    annotation_path = None
                if annotation_path and os.path.exists(annotation_path):
                    with open(annotation_path, 'r') as f:
                        annotation = json.load(f)
                        ground_truth_map[unique_image_path] = set(tuple(x) for x in annotation.get("defective_patches", []))
                else:
                    ground_truth_map[unique_image_path] = set()

            with torch.no_grad():
                # -----------------------------------------------------------------
                # Forward pass through VAE encoder (to latent space)
                # -----------------------------------------------------------------
                x = x.to(device)
                object_cls = object_cls.to(device)

                encoded = vae.encode(x).latent_dist.mean * _LATENT_SCALE

                # -----------------------------------------------------------------
                # Reverse DDIM sampling conditioned on encoder latents
                # -----------------------------------------------------------------
                model_kwargs = {"context": object_cls.unsqueeze(1), "mask": None}
                
                latent_samples_list = []
                for samples in diffusion.ddim_deviation_sample_loop_progressive(
                    model,
                    shape=encoded.shape,
                    noise=encoded,
                    clip_denoised=False,
                    start_t=reverse_steps,
                    model_kwargs=model_kwargs,
                    progress=False,
                    device=device,
                    eta=0.0,
                ):
                    latent_samples_list.append(samples["sample"])
                latent_samples_final = latent_samples_list[-1]

                image_samples_list = []
                for latent_samples in latent_samples_list:
                    image_samples_list.append(
                        vae.decode(latent_samples / _LATENT_SCALE).sample
                    )

                # -----------------------------------------------------------------
                # Reconstructions & other intermediate images
                # -----------------------------------------------------------------
                image_samples = vae.decode(latent_samples_final / _LATENT_SCALE).sample
                x0 = vae.decode(encoded / _LATENT_SCALE).sample

                # -----------------------------------------------------------------
                # Difference / binary maps
                # -----------------------------------------------------------------
                orig_dodrecon_diff_raw = _compute_abs_diff_max(x, image_samples)
                orig_dodrecon_diff = torch.clamp(orig_dodrecon_diff_raw, 0.0, 0.05) * 20
                orig_dodrecon_binary = _binary_mask(orig_dodrecon_diff, anomaly_binary_threshold)

                orig_encodedrecon_diff_raw = _compute_abs_diff_max(x, x0)
                orig_encodedrecon_diff = torch.clamp(orig_encodedrecon_diff_raw, 0.0, 0.05) * 20
                orig_encodedrecon_binary = _binary_mask(orig_encodedrecon_diff, anomaly_binary_threshold)
                
                encodedrecon_dodrecon_diff_raw = _compute_abs_diff_max(x0, image_samples)
                encodedrecon_dodrecon_diff = torch.clamp(encodedrecon_dodrecon_diff_raw, 0.0, 0.05) * 20

                encoded_latent_diff_raw = _compute_abs_diff_mean(latent_samples_final,encoded)
                encoded_latent_diff = torch.clamp(encoded_latent_diff_raw, 0.0, 0.05) * 20

                # Resize encoded_latent_diff to match the spatial dimensions of encodedrecon_dodrecon_diff
                # For irregular images, we want to use the patch size (128) as the target size
                patch_size = x.shape[-1]  # Should be 128 for patches
                encoded_latent_diff_resized = F.interpolate(
                    encoded_latent_diff,
                    size=(patch_size, patch_size),
                    mode="bilinear",
                    align_corners=False,
                )

                # -----------------------------------------------------------------
                # Composite anomaly maps
                # -----------------------------------------------------------------
                anomaly_map_arithmetic = 0.5 * (
                    encodedrecon_dodrecon_diff + encoded_latent_diff_resized
                )
                # Use contour-based binary masks for all three styles
                #anomaly_map_arithmetic_binary_style1, anomaly_map_arithmetic_binary_style2, anomaly_map_arithmetic_binary_style3 = _create_contour_based_binary_mask(anomaly_map_arithmetic, adaptive_threshold=adaptive_threshold, anomaly_binary_threshold=anomaly_binary_threshold)
                # Use style1 as default for backward compatibility
                anomaly_map_arithmetic_binary = _binary_mask(anomaly_map_arithmetic, anomaly_binary_threshold) #anomaly_map_arithmetic_binary_style1
                
                anomaly_map_geometric = (
                    encodedrecon_dodrecon_diff * encoded_latent_diff_resized
                )
                # Use contour-based binary masks for all three styles
                #anomaly_map_geometric_binary_style1, anomaly_map_geometric_binary_style2, anomaly_map_geometric_binary_style3 = _create_contour_based_binary_mask(anomaly_map_geometric, adaptive_threshold=adaptive_threshold, anomaly_binary_threshold=anomaly_binary_threshold)
                # Use style1 as default for backward compatibility
                anomaly_map_geometric_binary = _binary_mask(anomaly_map_geometric, anomaly_binary_threshold) #anomaly_map_geometric_binary_style1

                # Collect epoch-wise statistics
                if enable_epoch_stats:
                    metrics.add_batch_stats(encodedrecon_dodrecon_diff_raw, encoded_latent_diff_raw, anomaly_map_arithmetic, anomaly_map_geometric)
            
            # ---------------------------------------------------------------------
            # Per‑sample aggregation
            # ---------------------------------------------------------------------
            batch_size = x.size(0)
            
            for b in range(batch_size):
                # Determine if this patch is defective
                anomaly_max = int(round(anomaly_map_arithmetic[b].max().item() * 255))
                anomaly_binary = anomaly_map_arithmetic_binary[b]
                anomaly_pixels = torch.sum(anomaly_binary).item()
                is_predicted_defective = anomaly_pixels > anomaly_pixel_num_threshold
                
                # Store patch result for original image marking
                # patch_coords[b] is now a tensor [x, y]
                patch_coord_tensor = patch_coords[b]
                if isinstance(patch_coord_tensor, torch.Tensor):
                    x_coord = int(patch_coord_tensor[0].item())
                    y_coord = int(patch_coord_tensor[1].item())
                elif isinstance(patch_coord_tensor, (list, tuple)) and len(patch_coord_tensor) == 2:
                    x_coord, y_coord = patch_coord_tensor
                else:
                    # Fallback
                    print(f"Warning: unexpected patch_coord format, using default coordinates")
                    x_coord, y_coord = 0, 0
                
                # Always set ground_truth_defective before using it
                ground_truth_defective = ground_truth_map.get(image_paths[b], set())
                status = "TP" if is_predicted_defective and (x_coord, y_coord) in ground_truth_defective else \
                         "FP" if is_predicted_defective else \
                         "FN" if (x_coord, y_coord) in ground_truth_defective else "TN"
                
                # Create required record (always included)
                required_rec = make_record(
                    split=("meta", split),
                    image_path=("meta", image_paths[b]),
                    anomaly_class=("meta", anomaly_classes[b]),
                    patch_coords=("meta", (x_coord, y_coord)),
                    anomaly_max=("meta", anomaly_max),
                    anomaly_pixels=("meta", anomaly_pixels),
                    status=("meta", status),
                    orig=("image", _to_numpy(x[b])),
                    dod_recon=("image", _to_numpy(image_samples[b])),
                    encoded_recon=("image", _to_numpy(x0[b])),
                    anomaly_map_arithmetic=("image", _to_numpy(anomaly_map_arithmetic[b])),
                    anomaly_map_geometric=("image", _to_numpy(anomaly_map_geometric[b])),
                    anomaly_map_arithmetic_binary=(
                        "image",
                        _to_numpy(anomaly_map_arithmetic_binary[b]),
                    ),
                    anomaly_map_geometric_binary=(
                        "image",
                        _to_numpy(anomaly_map_geometric_binary[b]),
                    ),
                    encoded=("image", _to_numpy(encoded[b])),
                )

                results.append(required_rec)

                # Create optional record (only if flags are enabled)
                if enable_excel_report or enable_save_optional_image_results:
                    optional_rec = make_record(
                        split=("meta", split),
                        image_path=("meta", image_paths[b]),
                        anomaly_class=("meta", anomaly_classes[b]),
                        patch_coords=("meta", (x_coord, y_coord)),
                        status=("meta", status),
                        orig=("image", _to_numpy(x[b])),
                        dod_recon=("image", _to_numpy(image_samples[b])),
                        encoded_recon=("image", _to_numpy(x0[b])),
                        orig_dodrecon_diff=("image", _to_numpy(orig_dodrecon_diff[b])),
                        orig_dodrecon_binary=("image", _to_numpy(orig_dodrecon_binary[b])),
                        orig_encodedrecon_diff=("image", _to_numpy(orig_encodedrecon_diff[b])),
                        orig_encodedrecon_binary=(
                            "image",
                            _to_numpy(orig_encodedrecon_binary[b]),
                        ),
                        encodedrecon_dodrecon_diff=(
                            "image",
                            _to_numpy(encodedrecon_dodrecon_diff[b]),
                        ),
                        encoded_latent_diff=("image", _to_numpy(encoded_latent_diff[b])),
                        anomaly_map_arithmetic=("image", _to_numpy(anomaly_map_arithmetic[b])),
                        anomaly_map_arithmetic_binary=(
                            "image",
                            _to_numpy(anomaly_map_arithmetic_binary[b]),
                        ),
                        anomaly_map_geometric=("image", _to_numpy(anomaly_map_geometric[b])),
                        anomaly_map_geometric_binary=(
                            "image",
                            _to_numpy(anomaly_map_geometric_binary[b]),
                        ),
                        encoded=("image", _to_numpy(encoded[b])),
                        # Include all contour style anomaly maps in the optional record
                        #anomaly_map_arithmetic_binary_style1=("image", _to_numpy(anomaly_map_arithmetic_binary_style1[b])),
                        #anomaly_map_arithmetic_binary_style2=("image", _to_numpy(anomaly_map_arithmetic_binary_style2[b])),
                        #anomaly_map_arithmetic_binary_style3=("image", _to_numpy(anomaly_map_arithmetic_binary_style3[b])),
                        #anomaly_map_geometric_binary_style1=("image", _to_numpy(anomaly_map_geometric_binary_style1[b])),
                        #anomaly_map_geometric_binary_style2=("image", _to_numpy(anomaly_map_geometric_binary_style2[b])),
                        #anomaly_map_geometric_binary_style3=("image", _to_numpy(anomaly_map_geometric_binary_style3[b])),
                    )

                    add_metric_fields(optional_rec, device=device)
                    optional_results.append(optional_rec)
            
            # Memory optimization: Clear cache every 10 batches
            if idx % 10 == 0 and idx > 0:
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                import gc
                gc.collect()
                
                # Save checkpoint periodically
                if checkpoint_manager:
                    checkpoint_manager.save_checkpoint(idx, processed_images)
        
        image_results = defaultdict(list)
        for record in results:
            image_path = record["image_path"][1]  # Extract image_path from record
            image_results[image_path].append(record)

        # After processing this batch, check for any images that have all their patches processed
        completed_images = []
        for image_path in list(image_results.keys()):
            image_records = image_results[image_path]
            # Build predicted_defective_set for this image
            predicted_defective_set = set()
            for record in image_records:
                x, y = record["patch_coords"][1]
                anomaly_map = record["anomaly_map_arithmetic_binary"][1]
                anomaly_pixels = np.sum(anomaly_map)
                if anomaly_pixels > 0:
                    grid_row = y // patch_size
                    grid_col = x // patch_size
                    predicted_defective_set.add((grid_row, grid_col))
            # Now you can safely do:
            # Compute ground_truth_defective as before
            # ...
            overlapping = predicted_defective_set.intersection(ground_truth_defective)
            if not checkpoint_manager or not checkpoint_manager.is_image_processed(image_path):
                    # Load ground truth annotations for this image
                    ground_truth_defective = ground_truth_map.get(image_path, set())
                    save_image_results_from_records(
                        checkpoint_manager,
                        image_path,
                        image_results[image_path],
                        predicted_defective_set,
                        ground_truth_defective,
                        overlapping,
                        enable_save_optional_image_results,
                        patch_size
                    )
                    if checkpoint_manager:
                        checkpoint_manager.mark_image_processed(image_path)
                    # Mark for removal from dict
                    completed_images.append(image_path)
        # Remove completed images from dict
        for image_path in completed_images:
            del image_results[image_path]

        # Memory optimization: Clear cache every 10 batches
        if idx % 10 == 0 and idx > 0:
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
            import gc
            gc.collect()
            if checkpoint_manager:
                checkpoint_manager.save_checkpoint(idx, processed_images)

    except KeyboardInterrupt:
        print("\nProcess interrupted. Saving checkpoint...")
        if checkpoint_manager:
            checkpoint_manager.save_checkpoint(idx, processed_images)
        raise
    except Exception as e:
        print(f"\nError occurred: {e}")
        if checkpoint_manager:
            checkpoint_manager.save_checkpoint(idx, processed_images)
        raise

    # Print epoch-wise statistics (if enabled)
    if enable_epoch_stats:
        metrics.print_epoch_stats()
    else:
        print("Skipping epoch statistics (disabled)")

    return results, optional_results

def main():
    REPO_ROOT = os.environ.get("REPO_ROOT", None)
    if REPO_ROOT is not None:
        os.chdir(os.path.dirname(REPO_ROOT))
        print("Current path:", os.getcwd())
    if "ipykernel_launcher" in sys.argv[0]:
        print("Running in IPython kernel")
        sys.argv = [
            "",
            "--dataset",
            "pcb",
            "--data-dir",
            os.path.expanduser(
                "~/dataset/PCB/Huang/PCB_DATASET/PCB-gray-128___deco-diff"
            ),
            "--model-size",
            "UNet_L",
            "--object-class",
            "all",
            "--anomaly-class",
            "all",
            "--image-size",
            "128",
            "--center-size",
            "128",
            "--center-crop",
            "False",
            "--batch-num",
            "1",
            "--pretrained",
            "DeCo-Diff_pcb_all_UNet_L_128_CenterCrop/001-UNet_L/checkpoints/best.pt",
            "--split",
            "train",
            "--perturbation",
            "noise",
        ]
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dataset", type=str, choices=["mvtec", "visa", "pcb"], default="mvtec"
    )
    parser.add_argument("--data-dir", type=str, default="./mvtec-dataset/")
    parser.add_argument(
        "--model-size",
        type=str,
        choices=["UNet_XS", "UNet_S", "UNet_M", "UNet_L", "UNet_XL"],
        default="UNet_L",
    )
    parser.add_argument("--image-size", type=int, default=288)
    parser.add_argument("--center-size", type=int, default=256)
    parser.add_argument("--batch-num", type=int, default=12)
    parser.add_argument(
        "--center-crop",
        type=lambda v: True if v.lower() in ("yes", "true", "t", "y", "1") else False,
        default=True,
    )
    parser.add_argument(
        "--vae-type", type=str, choices=["ema", "mse"], default="ema"
    )  # Choice doesn't affect training
    parser.add_argument(
        "--num-workers", 
        type=int, 
        default=None,
        help="Number of workers for DataLoader (0 for single-threaded, useful on Windows). Default: 0 on Windows, 4 on other platforms"
    )
    parser.add_argument("--object-class", type=str, default="all")
    parser.add_argument("--pretrained", type=str, default=".")
    parser.add_argument("--anomaly-class", type=str, default="all")
    parser.add_argument("--reverse-steps", type=int, default=5)
    parser.add_argument("--split", type=str, default="test")
    parser.add_argument(
        "--perturbation",
        type=str,
        choices=[None, "brightness", "shift_x", "shift_y", "noise", "blur", "scratch"],
        default=None,
    )
    parser.add_argument("--split-csv-path", type=str, default=None)
    parser.add_argument(
        "--input-json",
        type=str,
        help="Path to JSON file containing multiple test configurations"
    )
    parser.add_argument(
        "--irregular-images",
        action="store_true",
        help="Process irregular-sized images by splitting into 128x128 patches"
    )
    parser.add_argument(
        "--annotation-dir",
        type=str,
        help="Directory containing JSON annotation files for defective regions"
    )
    parser.add_argument(
        "--anomaly-binary-threshold",
        type=int,
        default=5,
        help="Threshold for determining if a patch is defective based on number of anomaly pixels (default: 5)"
    )
    parser.add_argument(
        "--anomaly-pixel-num-threshold",
        type=int,
        default=0,
        help="Threshold for determining if a patch is defective based on number of anomaly pixels (default: 0)"
    )

    parser.add_argument(
        "--memory-optimization",
        action="store_true",
        help="Enable memory optimization for large datasets (reduces batch size and workers)"
    )
    parser.add_argument(
        "--enable-epoch-stats",
        action="store_true",
        default=False,
        help="Enable epoch statistics printing (memory intensive for large datasets)"
    )
    parser.add_argument(
        "--enable-excel-report",
        action="store_true",
        default=False,
        help="Enable Excel report generation (memory intensive for large datasets)"
    )
    parser.add_argument(
        "--enable-save-optional-image-results",
        action="store_true",
        default=False,
        help="Enable saving optional image results (memory intensive for large datasets)"
    )
    parser.add_argument(
        "--force-rerun",
        action="store_true",
        default=False,
        help="Force rerun evaluation even if checkpoint exists (clears existing checkpoint)"
    )
    parser.add_argument(
        "--patch-size",
        type=int,
        default=128,
        help="Patch size for splitting images (default: 128)"
    )
    args = parser.parse_args()
    
    # Handle input JSON if provided
    if args.input_json:
        import json
        with open(args.input_json, 'r') as f:
            test_configs = json.load(f)
            
        # Run evaluation for each test configuration
        for test_name, test_args in test_configs.items():
            print(f"\nRunning evaluation for {test_name}")
            print(test_args)
            # Update args with test configuration
            for key, value in test_args.items():
                # Convert key from kebab-case to snake_case
                key = key.replace('-', '_')
                if hasattr(args, key):
                    # Convert string values to appropriate types
                    if key in ['image_size', 'center_size', 'patch_size', 'batch_num', 'reverse_steps', 'anomaly_binary_threshold', 'anomaly_pixel_num_threshold']:
                        value = int(value)
                    elif key == 'center_crop':
                        value = value.lower() in ('yes', 'true', 't', 'y', '1')
                    elif key in ['pretrained', 'data_dir', 'split_csv_path', 'annotation_dir']:
                        value = os.path.expanduser(value)
                    elif key in ['irregular_images', 'memory_optimization', 'enable_epoch_stats', 'enable_excel_report', 'enable_save_optional_image_results', 'force_rerun']:
                        value = value.lower() in ('yes', 'true', 't', 'y', '1')
                    setattr(args, key, value)
            
            # Set up derived arguments
            if args.dataset == "mvtec":
                args.num_classes = 15
            elif args.dataset == "visa":
                args.num_classes = 12
            elif args.dataset == "pcb":
                args.num_classes = 1
            current_time = datetime.now().strftime("%y%m%d_%H%M%S")
            args.results_dir = f"results/{test_name}_{current_time}"
            os.makedirs(args.results_dir, exist_ok=True)
            # INSERT_YOUR_CODE
            # Save the current test_args (key-value pairs) into the results_dir as a JSON file
            config_save_path = os.path.join(args.results_dir, "config.json")
            with open(config_save_path, "w") as config_file:
                json.dump(test_args, config_file, indent=2)
            if args.center_crop:
                args.actual_image_size = args.center_size
            else:
                args.actual_image_size = args.image_size

            # Set up object classes
            if args.object_class == "all" and args.dataset == "mvtec":
                args.object_classes = [
                    "bottle", "cable", "capsule", "hazelnut", "metal_nut",
                    "pill", "screw", "toothbrush", "transistor", "zipper",
                    "carpet", "grid", "leather", "tile", "wood",
                ]
            elif args.object_class == "all" and args.dataset == "visa":
                args.object_classes = [
                    "candle", "cashew", "fryum", "macaroni2", "pcb2", "pcb4",
                    "capsules", "chewinggum", "macaroni1", "pcb1", "pcb3", "pipe_fryum",
                ]
            elif args.object_class == "all" and args.dataset == "pcb":
                args.object_classes = ["pcb"]
            else:
                args.object_classes = [args.object_class]
                
            # Run evaluation for this configuration
            evaluation(args)
    else:
        # Original single configuration evaluation
        if args.dataset == "mvtec":
            args.num_classes = 15
        elif args.dataset == "visa":
            args.num_classes = 12
        elif args.dataset == "pcb":
            args.num_classes = 1
        args.results_dir = f"./DeCo-Diff_{args.dataset}_{args.object_class}_{args.model_size}_{args.center_size}"
        if args.center_crop:
            args.results_dir += "_CenterCrop"
            args.actual_image_size = args.center_size
        else:
            args.actual_image_size = args.image_size

        if args.object_class == "all" and args.dataset == "mvtec":
            args.object_classes = [
                "bottle",
                "cable",
                "capsule",
                "hazelnut",
                "metal_nut",
                "pill",
                "screw",
                "toothbrush",
                "transistor",
                "zipper",
                "carpet",
                "grid",
                "leather",
                "tile",
                "wood",
            ]
        elif args.object_class == "all" and args.dataset == "visa":
            args.object_classes = [
                "candle",
                "cashew",
                "fryum",
                "macaroni2",
                "pcb2",
                "pcb4",
                "capsules",
                "chewinggum",
                "macaroni1",
                "pcb1",
                "pcb3",
                "pipe_fryum",
            ]
        elif args.object_class == "all" and args.dataset == "pcb":
            args.object_classes = [
                "pcb",
            ]
        else:
            args.object_classes = [args.object_class]

        evaluation(args)


# %%
if __name__ == "__main__":
    main()
# Below are cell makrkers used in VSCode
# %%
#
# %%