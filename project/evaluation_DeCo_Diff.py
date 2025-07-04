# %%
from __future__ import annotations
import warnings
warnings.filterwarnings(
    "ignore",
    message="A new version of Albumentations is available.*",
    category=UserWarning
)
from datetime import datetime
import torch
from skimage.transform import resize
from diffusion import create_diffusion
from diffusers.models.autoencoders.autoencoder_kl import AutoencoderKL
from models import UNET_models
import argparse
import numpy as np
import torch.nn.functional as F

from glob import glob

from torch.utils.data import DataLoader
from torchvision import transforms
from MVTECDataLoader import MVTECDataset
from VISADataLoader import VISADataset
from PCBDataLoader import PCBDataset
from scipy.ndimage import gaussian_filter

from anomalib import metrics
from sklearn.metrics import average_precision_score
from numpy import ndarray
import pandas as pd
from skimage import measure
from sklearn.metrics import auc

import os
import sys
from typing import List
import matplotlib.pyplot as plt
from collections import OrderedDict, defaultdict
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from typing import Sequence

from io import BytesIO
from pathlib import Path

from typing import Any, Tuple, cast

from torchmetrics.functional.image import (
    learned_perceptual_image_patch_similarity as _lpips,
    structural_similarity_index_measure as _ssim,
)
from sklearn.metrics import roc_curve
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay
import json

torch.set_grad_enabled(False)
device = "cuda" if torch.cuda.is_available() else "cpu"
if device == "cpu":
    print("GPU not found. Using CPU instead.")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_DIFF_SCALE = 2.0
_THRESHOLD = 5.0 / 255.0
_LATENT_SCALE = 0.18215

Kinded = Tuple[str, Any]  # (kind, value)
Record = OrderedDict[str, Kinded]

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------


def add_metric_fields(rec: Record, *, device=None) -> None:
    device = device or (
        torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
    )

    def to4d(x):
        if isinstance(x, np.ndarray):
            x = torch.from_numpy(x)
            if x.dtype != torch.float32:
                x = x.float()
            if x.ndim == 3 and x.shape[-1] == 3:  # HWC ➜ CHW
                x = x.permute(2, 0, 1)
            if x.ndim == 2:
                x = x.unsqueeze(0)
            if x.ndim == 3:
                x = x.unsqueeze(0)
            return x.to(device).clamp(-1, 1)

    a = to4d(rec["encoded_recon"][1])
    b = to4d(rec["dod_recon"][1])
    rec["lpips"] = ("metric", _lpips(a, b, net_type="alex").item())
    rec["ssim"] = ("metric", _ssim(a, b).item())
    rec["mse"] = ("metric", F.mse_loss(a, b).item())


def make_record(**kwargs) -> Record:
    """Return an **ordered** dict whose values are (kind, value) pairs."""
    return OrderedDict(kwargs)


def _compute_diff(a: torch.Tensor, b: torch.Tensor) -> torch.Tensor:
    """Return the mean channel‑wise difference *scaled* by ``_DIFF_SCALE``."""
    return (a - b).mean(dim=1, keepdim=True) / _DIFF_SCALE


def _binary_mask(diff: torch.Tensor, threshold: float = _THRESHOLD) -> torch.Tensor:
    """Return a binary mask in ``{-1, 1}`` based on *absolute* diff magnitude."""
    return (diff.abs() > threshold).float() * 2.0 - 1.0


def _to_numpy(
    t: torch.Tensor,
) -> "Sequence | torch.Tensor":  # keep Images API compatibility
    """Detach, move to CPU and convert to ``numpy`` if ``t`` is a tensor."""
    return t.detach().cpu().numpy() if isinstance(t, torch.Tensor) else t


def _tensor_to_xlimage(arr, size: int) -> XLImage:
    if isinstance(arr, torch.Tensor):
        arr = arr.detach().cpu().numpy()
    arr = np.squeeze(arr)
    if arr.ndim == 2:
        arr = arr[..., None]
    if arr.ndim == 3 and arr.shape[0] in (1, 3, 4):
        arr = np.transpose(arr, (1, 2, 0))
    c = arr.shape[2]
    if c == 1:
        arr = np.repeat(arr, 3, axis=2)
    elif c == 4:
        # Split the image into 4 quadrants if c == 4 (e.g., 4-channel image)
        # We'll arrange the 4 channels as 2x2 grid: [0|1]
        #                                             [2|3]
        # h, w = arr.shape[0], arr.shape[1]
        # h2, w2 = h // 2, w // 2
        # If the image is not square, just split in half along each axis
        # Each channel is a grayscale image, so we tile them
        q0 = arr[..., 0]
        q1 = arr[..., 1]
        q2 = arr[..., 2]
        q3 = arr[..., 3]
        # Stack as 2x2 grid
        top = np.concatenate([q0, q1], axis=1)
        bottom = np.concatenate([q2, q3], axis=1)
        arr = np.stack(
            [np.concatenate([top, bottom], axis=0)] * 3, axis=2
        )  # make 3 channels
    elif c != 3:
        raise ValueError(f"unsupported channels: {c}")
    arr = ((np.clip(arr, -1, 1) + 1) / 2 * 255).astype(np.uint8)

    buf = BytesIO()
    PILImage.fromarray(arr, mode="RGB").save(buf, format="PNG")
    buf.seek(0)
    img = XLImage(buf)
    img.width = img.height = size
    return img


def _write_row(ws, row_idx: int, rec: dict, size: int):
    scalars, embeds = [], []
    for col_idx, key in enumerate(rec.keys(), 1):
        kind, val = rec[key]
        if kind == "image":
            embeds.append((col_idx, _tensor_to_xlimage(val, size)))
            scalars.append("")
        else:
            scalars.append(val)
    ws.append(scalars)
    ws.row_dimensions[row_idx].height = size * 0.75
    for col_idx, img in embeds:
        ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")


# ---------------------------------------------------------------------------
# Functions
# ---------------------------------------------------------------------------


def process_split(
    dataloader,
    split: str,
    diffusion,
    model,
    vae,
    reverse_steps: int,
    center_size: int,
    batch_num: int,
    device: torch.device | None = None,
) -> List[Record]:
    """Run a forward‑&‑reverse pass on *one* dataset split and collect metrics.

    Parameters
    ----------
    dataloader : torch.utils.data.DataLoader
        Yields ``(x, seg, object_cls, image_paths, anomaly_classes)``.
    split : str
        Name of the split (e.g. ``"train"`` or ``"test"``).
    diffusion, model, vae : nn.Module‑like
        Components used for the DDIM deviation sampling pipeline.
    reverse_steps : int
        DDIM reverse steps.
    center_size : int
        Spatial size to which the latent anomaly map is resized.
    batch_num : int
        Maximum number of *mini‑batches* to process.
    device : torch.device | None, default = ``cuda`` if available
        Target device for all computations.

    Returns
    -------
    List[Record]
        One entry per *image* in the processed subset.
    """

    device = device or torch.device("cuda" if torch.cuda.is_available() else "cpu")

    results: List[Record] = []

    for idx, (x, seg, object_cls, image_paths, anomaly_classes) in enumerate(  # noqa: B905
        tqdm(dataloader, desc=f"{split} split")
    ):
        if idx >= batch_num:
            break

        with torch.no_grad():
            # -----------------------------------------------------------------
            # Forward pass through VAE encoder (to latent space)
            # -----------------------------------------------------------------
            x = x.to(device)
            object_cls = object_cls.to(device)

            encoded = vae.encode(x).latent_dist.mean * _LATENT_SCALE

            # -----------------------------------------------------------------
            # Reverse DDIM sampling conditioned on encoder latents
            # -----------------------------------------------------------------
            model_kwargs = {"context": object_cls.unsqueeze(1), "mask": None}
            # latent_samples = diffusion.ddim_deviation_sample_loop(
            #    model,
            #    shape=encoded.shape,
            #    noise=encoded,
            #    clip_denoised=False,
            #    start_t=reverse_steps,
            #    model_kwargs=model_kwargs,
            #    progress=False,
            #    device=device,
            #    eta=0.0,
            # )
            latent_samples_list = []
            for samples in diffusion.ddim_deviation_sample_loop_progressive(
                model,
                shape=encoded.shape,
                noise=encoded,
                clip_denoised=False,
                start_t=reverse_steps,
                model_kwargs=model_kwargs,
                progress=False,
                device=device,
                eta=0.0,
            ):
                latent_samples_list.append(samples["sample"])
            latent_samples_final = latent_samples_list[-1]

            image_samples_list = []
            # lat_slices_list = []
            for latent_samples in latent_samples_list:
                # lat_slices_list.append([latent_samples[:, i:i+1] for i in range(4)])
                image_samples_list.append(
                    vae.decode(latent_samples / _LATENT_SCALE).sample
                )

            # -----------------------------------------------------------------
            # Reconstructions & other intermediate images
            # -----------------------------------------------------------------
            image_samples = vae.decode(latent_samples_final / _LATENT_SCALE).sample
            x0 = vae.decode(encoded / _LATENT_SCALE).sample

            # -----------------------------------------------------------------
            # Difference / binary maps
            # -----------------------------------------------------------------
            orig_dodrecon_diff = _compute_diff(x, image_samples)
            orig_encodedrecon_diff = _compute_diff(x, x0)
            encodedrecon_dodrecon_diff = _compute_diff(x0, image_samples)

            orig_dodrecon_binary = _binary_mask(orig_dodrecon_diff)
            orig_encodedrecon_binary = _binary_mask(orig_encodedrecon_diff)
            encodedrecon_dodrecon_binary = _binary_mask(encodedrecon_dodrecon_diff)

            encoded_latent_diff = (
                (latent_samples_final - encoded).max(dim=1, keepdim=True).values
            )
            encoded_latent_binary = _binary_mask(encoded_latent_diff)

            encoded_latent_abs_diff_resized = F.interpolate(
                encoded_latent_diff.abs(),
                size=(center_size, center_size),
                mode="bilinear",
                align_corners=False,
            )

            # -----------------------------------------------------------------
            # Composite anomaly maps
            # -----------------------------------------------------------------
            anomaly_map_arithmetic = 0.5 * (
                encodedrecon_dodrecon_diff + encoded_latent_abs_diff_resized
            )
            anomaly_map_arithmetic_binary = _binary_mask(anomaly_map_arithmetic)
            anomaly_map_geometric = (
                encodedrecon_dodrecon_diff * encoded_latent_abs_diff_resized
            )
            anomaly_map_geometric_binary = _binary_mask(anomaly_map_geometric)
            # lat_slices = [latent_samples[:, i:i+1] for i in range(4)]
        # ---------------------------------------------------------------------
        # Per‑sample aggregation (no unsqueeze gymnastics)
        # ---------------------------------------------------------------------
        batch_size = x.size(0)
        for b in range(batch_size):
            rec = make_record(
                split=("meta", split),
                image_path=("meta", image_paths[b]),
                anomaly_class=("meta", anomaly_classes[b]),
                orig=("image", _to_numpy(x[b])),
                dod_recon=("image", _to_numpy(image_samples[b])),
                encoded_recon=("image", _to_numpy(x0[b])),
                orig_dodrecon_diff=("image", _to_numpy(orig_dodrecon_diff[b])),
                orig_dodrecon_binary=("image", _to_numpy(orig_dodrecon_binary[b])),
                orig_encodedrecon_diff=("image", _to_numpy(orig_encodedrecon_diff[b])),
                orig_encodedrecon_binary=(
                    "image",
                    _to_numpy(orig_encodedrecon_binary[b]),
                ),
                encodedrecon_dodrecon_diff=(
                    "image",
                    _to_numpy(encodedrecon_dodrecon_diff[b]),
                ),
                encodedrecon_dodrecon_binary=(
                    "image",
                    _to_numpy(encodedrecon_dodrecon_binary[b]),
                ),
                encoded_latent_diff=("image", _to_numpy(encoded_latent_diff[b])),
                encoded_latent_binary=("image", _to_numpy(encoded_latent_binary[b])),
                anomaly_map_arithmetic=("image", _to_numpy(anomaly_map_arithmetic[b])),
                anomaly_map_geometric=("image", _to_numpy(anomaly_map_geometric[b])),
                anomaly_map_arithmetic_binary=(
                    "image",
                    _to_numpy(anomaly_map_arithmetic_binary[b]),
                ),
                anomaly_map_geometric_binary=(
                    "image",
                    _to_numpy(anomaly_map_geometric_binary[b]),
                ),
                encoded=("image", _to_numpy(encoded[b])),
            )
            for i in range(len(image_samples_list)):
                if i < 5 or i % 10 == 0 or i >= len(image_samples_list) - 5:
                    rec[f"encoded_samples_{i}"] = (
                        "image",
                        _to_numpy(latent_samples_list[i][b]),
                    )
                    if i == 0:
                        rec[f"encoded_samples_diff_{i}"] = (
                            "image",
                            _to_numpy(latent_samples_list[i][b] - encoded[b]),
                        )
                    else:
                        rec[f"encoded_samples_diff_{i}"] = (
                            "image",
                            _to_numpy(
                                latent_samples_list[i][b]
                                - latent_samples_list[i - 1][b]
                            ),
                        )
                    rec[f"image_samples_{i}"] = (
                        "image",
                        _to_numpy(image_samples_list[i][b]),
                    )
                    if i == 0:
                        rec[f"image_samples_diff_{i}"] = (
                            "image",
                            _to_numpy(image_samples_list[i][b] - x[b]),
                        )
                    else:
                        rec[f"image_samples_diff_{i}"] = (
                            "image",
                            _to_numpy(
                                image_samples_list[i][b] - image_samples_list[i - 1][b]
                            ),
                        )

            add_metric_fields(rec, device=device)
            results.append(rec)

    return results


def compute_pro(masks: ndarray, amaps: ndarray, num_th: int = 200) -> None:
    """Compute the area under the curve of per-region overlaping (PRO) and 0 to 0.3 FPR
    Args:
        category (str): Category of product
        masks (ndarray): All binary masks in test. masks.shape -> (num_test_data, h, w)
        amaps (ndarray): All anomaly maps in test. amaps.shape -> (num_test_data, h, w)
        num_th (int, optional): Number of thresholds
    """

    assert isinstance(amaps, ndarray), "type(amaps) must be ndarray"
    assert isinstance(masks, ndarray), "type(masks) must be ndarray"
    assert amaps.ndim == 3, "amaps.ndim must be 3 (num_test_data, h, w)"
    assert masks.ndim == 3, "masks.ndim must be 3 (num_test_data, h, w)"
    assert amaps.shape == masks.shape, "amaps.shape and masks.shape must be same"
    assert set(masks.flatten()) == {0, 1}, "set(masks.flatten()) must be {0, 1}"
    assert isinstance(num_th, int), "type(num_th) must be int"

    df = pd.DataFrame([], columns=["pro", "fpr", "threshold"])
    binary_amaps = np.zeros_like(amaps, dtype=bool)

    min_th = amaps.min()
    max_th = amaps.max()
    delta = (max_th - min_th) / num_th

    for th in np.arange(min_th, max_th, delta):
        binary_amaps[amaps <= th] = 0
        binary_amaps[amaps > th] = 1

        pros = []
        for binary_amap, mask in zip(binary_amaps, masks):
            for region in measure.regionprops(measure.label(mask)):
                axes0_ids = region.coords[:, 0]
                axes1_ids = region.coords[:, 1]
                tp_pixels = binary_amap[axes0_ids, axes1_ids].sum()
                pros.append(tp_pixels / region.area)

        inverse_masks = 1 - masks
        fp_pixels = np.logical_and(inverse_masks, binary_amaps).sum()
        fpr = fp_pixels / inverse_masks.sum()

        df = pd.concat(
            [
                df,
                pd.DataFrame({"pro": [np.mean(pros)], "fpr": [fpr], "threshold": [th]}),
            ],
            ignore_index=True,
        )

    # Normalize FPR from 0 ~ 1 to 0 ~ 0.3
    df = df[df["fpr"] < 0.3]
    df["fpr"] = df["fpr"] / df["fpr"].max()

    pro_auc = auc(df["fpr"], df["pro"])
    return pro_auc


def calculate_metrics(ground_truth, prediction):
    flat_gt = ground_truth.flatten()
    flat_pred = prediction.flatten()

    auprc = metrics.AUPR()
    auprc_score = auprc(
        torch.from_numpy(flat_pred), torch.from_numpy(flat_gt.astype(int))
    )

    # aupro_score = 0
    aupro = metrics.AUPRO(fpr_limit=0.3)
    aupro_score = compute_pro(ground_truth, prediction)

    auroc = metrics.AUROC()
    auroc_score = auroc(
        torch.from_numpy(flat_pred), torch.from_numpy(flat_gt.astype(int))
    )

    f1max = metrics.F1Max()
    f1_max_score = f1max(
        torch.from_numpy(flat_pred), torch.from_numpy(flat_gt.astype(int))
    )

    ap = average_precision_score(ground_truth.flatten(), prediction.flatten())

    gt_list_sp = []
    pr_list_sp = []
    for idx in range(len(ground_truth)):
        gt_list_sp.append(np.max(ground_truth[idx]))
        sp_score = np.max(prediction[idx])
        pr_list_sp.append(sp_score)

    gt_list_sp = np.array(gt_list_sp).astype(np.int32)
    pr_list_sp = np.array(pr_list_sp)

    apsp = average_precision_score(gt_list_sp, pr_list_sp)
    aurocsp = auroc(torch.from_numpy(pr_list_sp), torch.from_numpy(gt_list_sp))
    f1sp = f1max(torch.from_numpy(pr_list_sp), torch.from_numpy(gt_list_sp))

    return (
        auroc_score.numpy(),
        aupro_score,
        f1_max_score.numpy(),
        ap,
        aurocsp.numpy(),
        apsp,
        f1sp.numpy(),
    )


def smooth_mask(mask, sigma=1.0):
    smoothed_mask = gaussian_filter(mask, sigma=sigma)
    return smoothed_mask


def calculate_anomaly_maps(
    x0_s, encoded_s, image_samples_s, latent_samples_s, center_size=256
):
    pred_geometric = []
    pred_arithmetic = []
    image_differences = []
    encoded_latent_differences = []
    input_images = []
    output_images = []
    for x, encoded, image_samples, latent_samples in zip(
        x0_s, encoded_s, image_samples_s, latent_samples_s
    ):
        input_image = (
            (np.clip(x[0].detach().cpu().numpy(), -1, 1).transpose(1, 2, 0)) * 127.5
            + 127.5
        ).astype(np.uint8)
        output_image = (
            (np.clip(image_samples[0].detach().cpu().numpy(), -1, 1).transpose(1, 2, 0))
            * 127.5
            + 127.5
        ).astype(np.uint8)
        input_images.append(input_image)
        output_images.append(output_image)

        image_difference = (
            (((torch.abs(image_samples - x)).to(torch.float32)).mean(axis=0))
            .detach()
            .cpu()
            .numpy()
            .transpose(1, 2, 0)
            .max(axis=2)
        )
        image_difference = (np.clip(image_difference, 0.0, 0.4)) * 2.5
        image_difference = smooth_mask(image_difference, sigma=3)
        image_differences.append(image_difference)

        encoded_latent_difference = (
            (((torch.abs(latent_samples - encoded)).to(torch.float32)).mean(axis=0))
            .detach()
            .cpu()
            .numpy()
            .transpose(1, 2, 0)
            .mean(axis=2)
        )
        encoded_latent_difference = (np.clip(encoded_latent_difference, 0.0, 0.4)) * 2.5
        encoded_latent_difference = smooth_mask(encoded_latent_difference, sigma=1)
        encoded_latent_difference = resize(
            encoded_latent_difference, (center_size, center_size)
        )
        encoded_latent_differences.append(encoded_latent_difference)

        final_anomaly = image_difference * encoded_latent_difference
        final_anomaly = np.sqrt(final_anomaly)
        final_anomaly = smooth_mask(final_anomaly, sigma=1)
        final_anomaly2 = 1 / 2 * image_difference + 1 / 2 * encoded_latent_difference
        final_anomaly2 = smooth_mask(final_anomaly2, sigma=1)
        pred_geometric.append(final_anomaly)
        pred_arithmetic.append(final_anomaly2)

    pred_geometric = np.stack(pred_geometric, axis=0)
    pred_arithmetic = np.stack(pred_arithmetic, axis=0)
    encoded_latent_differences = np.stack(encoded_latent_differences, axis=0)
    image_differences = np.stack(image_differences, axis=0)

    return {
        "anomaly_arithmetic": pred_arithmetic,
        "anomaly_geometric": pred_geometric,
        "latent_discrepancy": encoded_latent_differences,
        "image_discrepancy": image_differences,
    }


def evaluate_anomaly_maps(anomaly_maps, segmentation):
    for key in anomaly_maps.keys():
        auroc_score, aupro_score, f1_max_score, ap, aurocsp, apsp, f1sp = (
            calculate_metrics(segmentation, anomaly_maps[key])
        )
        (
            auroc_score,
            aupro_score,
            f1_max_score,
            ap,
            aurocsp,
            apsp,
            f1sp,
        ) = (
            np.round(auroc_score, 4),
            np.round(aupro_score, 4),
            np.round(f1_max_score, 4),
            np.round(ap, 4),
            np.round(aurocsp, 4),
            np.round(apsp, 4),
            np.round(f1sp, 4),
        )
        print(
            "{}: auroc:{:.4f}, aupro:{:.4f}, f1_max:{:.4f}, ap:{:.4f}, aurocsp:{:.4f}, apsp:{:.4f}, f1sp:{:.4f}".format(
                key, auroc_score, aupro_score, f1_max_score, ap, aurocsp, apsp, f1sp
            )
        )


def evaluation(args):
    if os.path.exists("./models/config.json"):
        vae = cast(AutoencoderKL, AutoencoderKL.from_pretrained("./models", local_files_only=True)).to(
            device
        )
    else:
        vae = cast(AutoencoderKL, AutoencoderKL.from_pretrained(f"stabilityai/sd-vae-ft-{args.vae_type}")).to(
            device
        )
    vae.eval()
    try:
        if args.pretrained != "":
            ckpt = args.pretrained
        else:
            path = f"./DeCo-Diff_{args.dataset}_{args.object_class}_{args.model_size}_{args.center_size}"
            try:
                ckpt = sorted(glob(f"{path}/last.pt"))[-1]
            except (IndexError, FileNotFoundError):
                ckpt = sorted(glob(f"{path}/*/last.pt"))[-1]
    except (IndexError, FileNotFoundError, OSError):
        raise Exception("Please provide the model's pretrained path using --pretrained")

    latent_size = int(args.center_size) // 8
    model = UNET_models[args.model_size](latent_size=latent_size)

    state_dict = torch.load(ckpt)["model"]
    print(model.load_state_dict(state_dict))
    model.eval()  # important!
    model.cuda()
    print("model loaded")

    print("==" * 30)
    print("Starting Evaluation...")
    print("==" * 30)
    diffusion = create_diffusion(
        f"ddim{args.reverse_steps}",
        predict_deviation=True,
        sigma_small=False,
        predict_xstart=False,
        diffusion_steps=1000,
    )

    for object_class in args.object_classes:
        transform = transforms.Compose(
            [
                transforms.ToTensor(),
                transforms.Normalize(
                    mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5], inplace=True
                ),
            ]
        )
        common_args = dict(
            split=args.split,
            object_class=object_class,
            rootdir=args.data_dir,
            transform=transform,
            anomaly_class=args.anomaly_class,
            image_size=args.image_size,
            center_size=args.actual_image_size,
            center_crop=True,
            process_split_fn=process_split,
            diffusion=diffusion,
            model=model,
            vae=vae,
            reverse_steps=args.reverse_steps,
            batch_num=args.batch_num,
            device=device,
            split_csv_path=args.split_csv_path,
        )
        if args.dataset == "pcb":
            common_args["dataset_class"] = PCBDataset
        elif args.dataset == "mvtec":
            common_args["dataset_class"] = MVTECDataset
        elif args.dataset == "visa":
            common_args["dataset_class"] = VISADataset
        else:
            raise ValueError(f"Invalid dataset: {args.dataset}")
        if args.perturbation is not None:
            if args.perturbation == "brightness":
                param_values = np.arange(-20, 21, 1)
                record_pairs = collect_records_for_params(
                    param_name="brightness", param_values=param_values, **common_args
                )
            if args.perturbation == "shift_x":
                param_values = np.arange(-20, 21, 1)
                record_pairs = collect_records_for_params(
                    param_name="shift_x", param_values=param_values, **common_args
                )
            if args.perturbation == "shift_y":
                param_values = np.arange(-20, 21, 1)
                record_pairs = collect_records_for_params(
                    param_name="shift_y", param_values=param_values, **common_args
                )
            if args.perturbation == "noise":
                param_values = np.arange(0, 21, 1)
                record_pairs = collect_records_for_params(
                    param_name="noise", param_values=param_values, **common_args
                )
            if args.perturbation == "blur":
                param_values = np.arange(1, 42, 2)
                record_pairs = collect_records_for_params(
                    param_name="blur", param_values=param_values, **common_args
                )
            if args.perturbation == "scratch":
                param_values = [0]
                record_pairs = collect_records_for_params(
                    param_name="brightness", param_values=param_values, **common_args
                )
            y_true_score_list = compute_y_true_y_score(record_pairs)
            roc_stats = compute_metrics_from_y_true_y_score(y_true_score_list)
            save_perturbation_results(
                param_name=args.perturbation,
                roc_stats=roc_stats,
                param_values=param_values,
                save_dir=args.results_dir,
            )

            plot_accuracy_results(
                param_name=args.perturbation,
                param_values=param_values,
                accuracies=roc_stats["accuracies"],
                color="red",
                save_dir=args.results_dir,
            )
            #plot_roc_curves(
            #    param_name=args.perturbation,
            #    roc_stats=roc_stats,
            #    param_values=param_values,
            #    save_dir=args.results_dir,
            #)
            #plot_confusion_matrices(
            #    param_name=args.perturbation,
            #    roc_stats=roc_stats,
            #    param_values=param_values,
            #    save_dir=args.results_dir,
            #)
            #records = []
            #for i in range(len(record_pairs[0][0])):
            #    for record_pair in record_pairs:
            #        record_train = record_pair[0][i]
            #        records.append(record_train)
            #        record_train_defect = record_pair[1][i]
            #        record_diff = diff_records(record_train, record_train_defect)
            #        records.append(record_train_defect)
            #        records.append(record_diff)
            #make_excel(records, args.image_size, save_dir=args.results_dir)
            #plot_distribution(records, save_dir=args.results_dir)

        # Create diffusion object:

        # records = []
        # if args.dataset == 'mvtec':
        #    train_dataset = MVTECDataset('train', object_class=category, rootdir=args.data_dir, transform=transform, normal=True, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        #    val_dataset = MVTECDataset('val', object_class=category, rootdir=args.data_dir, transform=transform, normal=False, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        #    test_dataset = MVTECDataset('test', object_class=category, rootdir=args.data_dir, transform=transform, normal=False, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        # elif args.dataset == 'visa':
        #    train_dataset = VISADataset('train', object_class=category, rootdir=args.data_dir, transform=transform, normal=True, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        #    val_dataset = VISADataset('val', object_class=category, rootdir=args.data_dir, transform=transform, normal=False, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        #    test_dataset = VISADataset('test', object_class=category, rootdir=args.data_dir, transform=transform, normal=False, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        # elif args.dataset == 'pcb':
        #    dataset = PCBDataset(args.split, object_class=category, rootdir=args.data_dir, transform=transform, normal=True, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        #    dataset_defect = PCBDataset(args.split, object_class=category, rootdir=args.data_dir, transform=transform, normal=True, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True, scratch=True)
        #
        #    #test_dataset = PCBDataset('test', object_class=category, rootdir=args.data_dir, transform=transform, normal=False, anomaly_class=args.anomaly_class, image_size=args.image_size, center_size=args.actual_image_size, center_crop=True)
        # loader = DataLoader(dataset, batch_size=8, shuffle=False, num_workers=4, drop_last=False)
        # loader_defect = DataLoader(dataset_defect, batch_size=8, shuffle=False, num_workers=4, drop_last=False)
        ##test_loader = DataLoader(test_dataset, batch_size=8, shuffle=False, num_workers=4, drop_last=False)

        # records = process_split(loader, args.split, diffusion, model, vae, args.reverse_steps, args.center_size, args.batch_num, device)
        # records_defect = process_split(loader_defect, args.split, diffusion, model, vae, args.reverse_steps, args.center_size, args.batch_num, device)
        ##records_test = process_split(test_loader, 'test', diffusion, model, vae, args.reverse_steps, args.center_size, args.batch_num, device)

        # plot_roc_curve_confusion_matrix(records, records_defect)
        # plot_distribution(records)
        #
        # make_excel(records, args.image_size)

        """
        encoded_s = []
        image_samples_s = []
        latent_samples_s = []
        x0_s = []
        x_s = []
        segmentation_s = []
        for ii, (x, seg, object_cls) in enumerate(test_loader):
            with torch.no_grad():
                # Map input images to latent space + normalize latents:
                encoded = vae.encode(x.to(device)).latent_dist.mean.mul_(0.18215)
                model_kwargs = {
                'context':object_cls.to(device).unsqueeze(1),
                'mask': None
                }
                latent_samples = diffusion.ddim_deviation_sample_loop(
                    model, encoded.shape, noise = encoded, clip_denoised=False, 
                    start_t = args.reverse_steps,
                    model_kwargs=model_kwargs, progress=False, device=device,
                    eta = 0
                )

                image_samples = vae.decode(latent_samples / 0.18215).sample 
                x0 = vae.decode(encoded / 0.18215).sample 

            segmentation_s += [_seg.squeeze() for _seg in seg]
            encoded_s += [_encoded.unsqueeze(0) for _encoded in encoded]
            image_samples_s += [_image_samples.unsqueeze(0) for _image_samples in image_samples]
            latent_samples_s += [_latent_samples.unsqueeze(0) for _latent_samples in latent_samples]
            x0_s += [_x0.unsqueeze(0) for _x0 in x0]
            x_s += [_x.unsqueeze(0) for _x in x]

        records = [
            Images(
                split="test",
                original_image=torch.clamp(img1, -1.0, 1.0).cpu().numpy() if hasattr(img1, 'cpu') else np.clip(img1, -1.0, 1.0),
                reconstructed_image=torch.clamp(img2, -1.0, 1.0).cpu().numpy() if hasattr(img2, 'cpu') else np.clip(img2, -1.0, 1.0)
            )
            for img1, img2 in zip(x_s, image_samples_s)
        ]

        plot_distribution(records, device=device)
        anomaly_maps = calculate_anomaly_maps(x0_s, encoded_s,  image_samples_s, latent_samples_s, center_size=args.center_size)
        
        evaluate_anomaly_maps(anomaly_maps, np.stack(segmentation_s, axis=0))
        """
        print("==" * 30)


def diff_records(a: Record, b: Record) -> Record:
    diff = make_record(
        split=("meta", "diff"),
        image_path=a["image_path"],
        anomaly_class=("meta", "diff"),
    )
    for k in a:
        if k in ("split", "image_path", "anomaly_class"):
            continue
        kind, va = a[k]
        vb = b[k][1]
        if kind == "image":
            diff[k] = ("image", va - vb)
        elif kind == "metric":
            diff[k] = ("metric", va - vb)
    # add_metric_fields(diff)
    return diff


def diff_records2(records_train: dict, records_train_defect: dict):
    record = {}
    # Use _COLS to determine kind: 'image', 'metric', or 'other'
    for key, item in records_train.items():
        kind, value = item
        if kind == "image" or kind == "metric":
            record[key] = (kind, value - records_train_defect[key][1])
        else:
            record[key] = (kind, value)
    return record


def cal_similarity(img1, img2, device=None, similarity_type="lpips"):
    if device is None:
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    def to_tensor(img):
        if isinstance(img, np.ndarray):
            img = torch.from_numpy(img)
        if img.dtype != torch.float32:
            img = img.float()
        if img.ndim == 3 and img.shape[-1] == 3:
            img = img.permute(2, 0, 1)
        if img.ndim == 4 and img.shape[-1] == 3:
            img = img.permute(0, 3, 1, 2)
        if img.ndim == 2:
            img = img.unsqueeze(0)
        if img.ndim == 3:
            img = img.unsqueeze(0)
        return img

    img1 = to_tensor(img1).to(device).clamp(min=-1, max=1)
    img2 = to_tensor(img2).to(device).clamp(min=-1, max=1)

    if similarity_type == "lpips":
        sim = _lpips(img1, img2, net_type="alex")
    elif similarity_type == "ssim":
        sim = _ssim(img1, img2)
    elif similarity_type == "mse":
        sim = F.mse_loss(img1, img2)
    else:
        raise ValueError(f"Invalid similarity type: {similarity_type}")
    return sim.cpu().item()


def plot_distribution(
    records: List[Record],
    save_dir="similarity_distribution",
    save_filename=datetime.now().strftime("%y%m%d_%H%M%S"),
):
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    splits = defaultdict(list)

    for similarity_type in ["ssim", "lpips", "mse"]:
        print(f"Processing {similarity_type} distribution")
        for rec in records:
            splits[f"{rec['split']}_{rec['anomaly_class']}"].append(
                rec[similarity_type][1]
            )

        plt.figure(figsize=(8, 6))
        for split, vals in splits.items():
            vals = np.array(vals)
            plt.hist(
                vals, bins=30, alpha=0.6, label=f"{split} (n={len(vals)})", density=True
            )
        plt.xlabel(similarity_type.upper())
        plt.ylabel("Density")
        plt.title(f"Distribution of {similarity_type.upper()} by Split")
        plt.legend()
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, f"{similarity_type}_{save_filename}.png"))
        plt.close()


def make_excel(
    records: List[Record],
    image_size: int,
    save_dir: str | Path = "report",
    save_filename: str | None = datetime.now().strftime("%y%m%d_%H%M%S"),
) -> Path:
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    # Header comes from the first record's keys (order preserved)
    header = list(records[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(header)

    for r, rec in enumerate(records, start=2):
        _write_row(ws, r, rec, image_size)
    for c in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    out_path = save_dir / f"report_{save_filename}.xlsx"
    wb.save(out_path)
    print(f"Report saved to {out_path}")
    return out_path


def compute_roc_stats(y_true, y_score):
    """
    Compute ROC statistics and best threshold using Youden's J statistic.
    Returns: fpr, tpr, thresholds, best_threshold, best_idx, auc_score
    """
    fpr, tpr, thresholds = roc_curve(y_true, y_score)
    youden_j = tpr - fpr
    best_idx = np.argmax(youden_j)
    best_threshold = thresholds[best_idx]
    auc_score = auc(fpr, tpr)
    return fpr, tpr, thresholds, best_threshold, best_idx, auc_score


def plot_roc_curve_confusion_matrix(
    records_good,
    records_anomaly,
    save_dir="roc_curve_confusion_matrix",
    save_filename=None,
):
    if not all(isinstance(rec, dict) for rec in records_good + records_anomaly):
        raise TypeError("All elements in records must be of type dict")
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    print("Processing ROC curve")
    y_true = []
    y_score = []
    for rec in records_good:
        y_true.append(0)
        mask = (
            rec["anomaly_map_arithmetic_binary"][1]
            if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
            else rec["anomaly_map_arithmetic_binary"]
        )
        num_white = np.sum(mask == 1)
        y_score.append(num_white)
    for rec in records_anomaly:
        y_true.append(1)
        mask = (
            rec["anomaly_map_arithmetic_binary"][1]
            if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
            else rec["anomaly_map_arithmetic_binary"]
        )
        num_white = np.sum(mask == 1)
        y_score.append(num_white)
    y_true = np.array(y_true)
    y_score = np.array(y_score)

    fpr, tpr, thresholds, best_threshold, best_idx, roc_auc = compute_roc_stats(
        y_true, y_score
    )

    plt.figure()
    plt.plot(
        fpr, tpr, color="darkorange", lw=2, label=f"ROC curve (area = {roc_auc:.2f})"
    )
    plt.plot([0, 1], [0, 1], color="navy", lw=2, linestyle="--")
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title("ROC Curve (White Pixel Count in Anomaly Map)")
    plt.legend(loc="lower right")
    plt.savefig(os.path.join(save_dir, f"roc_curve_{save_filename}.png"))
    plt.close()

    print(f"Best threshold: {best_threshold}")

    # Predict using the best threshold
    y_pred = (y_score >= best_threshold).astype(int)

    # Compute confusion matrix
    cm = confusion_matrix(y_true, y_pred)
    disp = ConfusionMatrixDisplay(
        confusion_matrix=cm, display_labels=["Good", "Anomaly"]
    )
    disp.plot(cmap=plt.cm.Blues)
    plt.title(f"Confusion Matrix (Threshold={best_threshold:.2f})")
    plt.savefig(os.path.join(save_dir, f"confusion_matrix_{save_filename}.png"))
    plt.close()


def collect_records_for_params(
    *,
    param_name: str,
    param_values,
    split: str,
    object_class: str,
    rootdir: str,
    transform,
    anomaly_class: str,
    image_size: int,
    center_size: int,
    center_crop: bool,
    process_split_fn,
    diffusion,
    model,
    vae,
    reverse_steps,
    batch_num,
    device=None,
    dataset_class,
    split_csv_path: str = None,
):
    common_args = dict(
        mode=split,
        object_class=object_class,
        rootdir=rootdir,
        transform=transform,
        anomaly_class=anomaly_class,
        image_size=image_size,
        center_size=center_size,
        center_crop=center_crop,
        split_csv_path=split_csv_path,
    )
    all_records = []
    for val in param_values:
        print(f"Processing {param_name} = {val}")
        kwargs = common_args.copy()
        kwargs[param_name] = val
        dataset = dataset_class(**kwargs)
        loader = DataLoader(
            dataset, batch_size=8, shuffle=False, num_workers=4, drop_last=False
        )
        records = process_split_fn(
            loader,
            split,
            diffusion,
            model,
            vae,
            reverse_steps,
            center_size,
            batch_num,
            device,
        )

        kwargs["scratch"] = True
        dataset_defect = dataset_class(**kwargs)
        loader_defect = DataLoader(
            dataset_defect, batch_size=8, shuffle=False, num_workers=4, drop_last=False
        )
        records_defect = process_split_fn(
            loader_defect,
            split,
            diffusion,
            model,
            vae,
            reverse_steps,
            center_size,
            batch_num,
            device,
        )
        all_records.append((records, records_defect))
    return all_records


def compute_y_true_y_score(all_records):
    """
    For each param value, compute y_true and y_score arrays from records.
    Returns: list of (y_true, y_score) tuples, one for each param value.
    """
    y_true_score_list = []
    for i, (records, records_defect) in enumerate(all_records):
        y_true = []
        y_score = []
        for rec in records:
            y_true.append(0)
            mask = (
                rec["anomaly_map_arithmetic_binary"][1]
                if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
                else rec["anomaly_map_arithmetic_binary"]
            )
            num_white = np.sum(mask == 1)
            y_score.append(num_white)
        for rec in records_defect:
            y_true.append(1)
            mask = (
                rec["anomaly_map_arithmetic_binary"][1]
                if isinstance(rec["anomaly_map_arithmetic_binary"], tuple)
                else rec["anomaly_map_arithmetic_binary"]
            )
            num_white = np.sum(mask == 1)
            y_score.append(num_white)
        y_score = np.array(y_score)
        y_true = np.array(y_true)
        y_true_score_list.append((y_true, y_score))
    return y_true_score_list


def compute_metrics_from_y_true_y_score(y_true_score_list):
    """
    For each (y_true, y_score), compute accuracy, threshold, and ROC stats.
    Returns: accuracies, thresholds, and a dict of lists for each ROC metric.
    """
    accuracies = []
    fpr_list = []
    tpr_list = []
    thresholds_list = []
    best_thresholds = []
    best_idxs = []
    aucs = []
    y_trues = []
    y_preds = []
    y_scores = []

    for i, (y_true, y_score) in enumerate(y_true_score_list):
        fpr, tpr, thresholds_, best_threshold, best_idx, auc_score = compute_roc_stats(
            y_true, y_score
        )
        y_pred = (y_score >= best_threshold).astype(int)
        y_preds.append(y_pred)
        accuracy = np.mean(y_pred == y_true)
        accuracies.append(accuracy)
        fpr_list.append(fpr)
        tpr_list.append(tpr)
        # thresholds_list.append(thresholds_)
        best_thresholds.append(best_threshold)
        best_idxs.append(best_idx)
        aucs.append(auc_score)
        y_trues.append(y_true)
        y_scores.append(y_score)
        print(f"Accuracy {accuracy:.4f} (threshold={best_threshold})")

    roc_stats = {
        "fpr": fpr_list,
        "tpr": tpr_list,
        # "thresholds": thresholds_list,
        "best_threshold": best_thresholds,
        "best_idx": best_idxs,
        "auc": aucs,
        "y_true": y_trues,
        "y_pred": y_preds,
        "y_score": y_scores,
        "accuracies": accuracies,
    }
    return roc_stats


def plot_accuracy_results(
    param_values,
    accuracies,
    param_name: str,
    save_dir="accuracy_vs_param",
    save_filename=datetime.now().strftime("%y%m%d_%H%M%S"),
    title=None,
    xlabel=None,
    ylabel="Accuracy",
    grid=True,
    marker="o",
    **plot_kwargs,
):
    """
    Plot accuracy results with customizable parameters.

    Args:
        param_values: List of parameter values
        accuracies: List of corresponding accuracies
        param_name: Name of the parameter being varied
        save_dir: Directory to save the plot
        save_filename: Filename for saving the plot
        title: Custom title for the plot
        xlabel: Custom x-axis label
        ylabel: Custom y-axis label
        grid: Whether to show grid
        marker: Marker style for the plot
        **plot_kwargs: Additional plotting parameters
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    plt.figure()
    plt.plot(param_values, accuracies, marker=marker, **plot_kwargs)
    plt.xlabel(xlabel or param_name)
    plt.ylabel(ylabel)
    plt.title(title or f"Accuracy vs {param_name.capitalize()} (synthetic defect)")
    plt.ylim(0.5, 1.0)
    plt.grid(grid)
    out_path = os.path.join(save_dir, f"accuracy_vs_{param_name}_{save_filename}.png")
    plt.savefig(out_path)
    print(f"Accuracy vs {param_name} saved to {out_path}")
    plt.close()


def plot_roc_curves(
    roc_stats,
    param_values=None,
    param_name="param",
    save_dir="roc_curves",
    save_filename=datetime.now().strftime("%y%m%d_%H%M%S"),
    figsize=(10, 6),
    colors=None,
    linestyles=None,
):
    """
    Plot multiple ROC curves on the same figure using roc_stats.

    Args:
        roc_stats: Dictionary containing ROC statistics (from compute_metrics_from_y_true_y_score)
        param_values: List of parameter values (for legend labels)
        param_name: Name of the parameter being varied
        save_dir: Directory to save the plot
        save_filename: Filename for saving the plot
        figsize: Figure size (width, height)
        colors: List of colors for each curve (optional)
        linestyles: List of line styles for each curve (optional)
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    plt.figure(figsize=figsize)

    # Default colors and linestyles if not provided
    if colors is None:
        colors = plt.cm.viridis(np.linspace(0, 1, len(roc_stats["fpr"])))
    if linestyles is None:
        linestyles = ["-"] * len(roc_stats["fpr"])
    # Plot each ROC curve
    for i in range(len(roc_stats["fpr"])):
        fpr = roc_stats["fpr"][i]
        tpr = roc_stats["tpr"][i]
        auc = roc_stats["auc"][i]

        # Create label with parameter value and AUC
        if param_values is not None:
            label = f"{param_name}={param_values[i]:.2f} (AUC={auc:.3f})"
        else:
            label = f"Curve {i + 1} (AUC={auc:.3f})"

        plt.plot(fpr, tpr, color=colors[i], linestyle=linestyles[i], lw=2, label=label)

    # Plot diagonal line
    plt.plot([0, 1], [0, 1], color="navy", lw=2, linestyle="--", label="Random")

    # Set plot properties
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title("ROC Curves for Different Parameter Values")

    # Add legend
    plt.legend(loc="lower right", bbox_to_anchor=(1.15, 0))

    # Adjust layout to prevent label cutoff
    plt.tight_layout()

    # Save plot
    out_path = os.path.join(save_dir, f"roc_curves_{save_filename}.png")
    plt.savefig(out_path, bbox_inches="tight")
    print(f"ROC curves saved to {out_path}")
    plt.close()


def plot_confusion_matrices(
    roc_stats,
    param_values=None,
    param_name="param",
    save_dir="confusion_matrices",
    save_filename=datetime.now().strftime("%y%m%d_%H%M%S"),
    figsize=(15, 5),
    cmap="Blues",
):
    """
    Plot multiple confusion matrices side by side using roc_stats.

    Args:
        roc_stats: Dictionary containing ROC statistics (from compute_metrics_from_y_true_y_score)
        param_values: List of parameter values (for subplot titles)
        param_name: Name of the parameter being varied
        save_dir: Directory to save the plot
        save_filename: Filename for saving the plot
        figsize: Figure size (width, height)
        cmap: Colormap for the confusion matrices
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    n_matrices = len(roc_stats["y_true"])
    fig, axes = plt.subplots(1, n_matrices, figsize=figsize)
    if n_matrices == 1:
        axes = [axes]

    for i, ax in enumerate(axes):
        y_true = roc_stats["y_true"][i]
        y_pred = roc_stats["y_pred"][i]
        threshold = roc_stats["best_threshold"][i]

        # Compute confusion matrix
        cm = confusion_matrix(y_true, y_pred)

        # Create confusion matrix display
        disp = ConfusionMatrixDisplay(
            confusion_matrix=cm, display_labels=["Normal", "Anomaly"]
        )

        # Plot confusion matrix
        disp.plot(ax=ax, cmap=cmap)

        # Set title with parameter value and threshold
        if param_values is not None:
            title = f"{param_name}={param_values[i]:.2f}\nThreshold={threshold:.2f}"
        else:
            title = f"Matrix {i + 1}\nThreshold={threshold:.2f}"
        ax.set_title(title)

        # Add accuracy to the plot
        accuracy = roc_stats["accuracies"][i]
        ax.text(
            0.5,
            -0.2,
            f"Accuracy: {accuracy:.4f}",
            horizontalalignment="center",
            transform=ax.transAxes,
        )

    # Adjust layout
    plt.tight_layout()

    # Save plot
    out_path = os.path.join(save_dir, f"confusion_matrices_{save_filename}.png")
    plt.savefig(out_path, bbox_inches="tight")
    print(f"Confusion matrices saved to {out_path}")
    plt.close()


def save_perturbation_results(
    param_name: str,
    roc_stats: dict,
    param_values: list,
    save_dir: str,
):
    """
    Save perturbation experiment data (roc_stats and param_values) to a specified folder in JSON format.

    Args:
        param_name: Name of the perturbation parameter
        roc_stats: Dictionary containing ROC statistics
        param_values: List of parameter values
        save_dir: Directory to save the results
    """
    save_dir = Path(save_dir).expanduser()
    save_dir.mkdir(parents=True, exist_ok=True)

    # Convert numpy arrays to lists for JSON serialization
    def convert_for_json(obj):
        if hasattr(obj, 'tolist'):  # numpy arrays
            return obj.tolist()
        elif isinstance(obj, list):
            return [convert_for_json(item) for item in obj]
        elif isinstance(obj, dict):
            return {key: convert_for_json(value) for key, value in obj.items()}
        else:
            return obj

    # Convert roc_stats to JSON-serializable format
    roc_stats_json = convert_for_json(roc_stats)
    
    # Convert param_values to JSON-serializable format
    param_values_json = convert_for_json(param_values)

    # Save both roc_stats and param_values in a single JSON file
    results_data = {
        "param_name": param_name,
        "param_values": param_values_json,
        "roc_stats": roc_stats_json
    }
    
    json_path = os.path.join(save_dir, f"{param_name}_results.json")
    with open(json_path, "w") as f:
        json.dump(results_data, f, indent=2)

    print(f"Perturbation results saved to {json_path}")


def main():
    REPO_ROOT = os.environ.get("REPO_ROOT", None)
    if REPO_ROOT is not None:
        os.chdir(os.path.dirname(REPO_ROOT))
        print("Current path:", os.getcwd())
    if "ipykernel_launcher" in sys.argv[0]:
        print("Running in IPython kernel")
        sys.argv = [
            "",
            "--dataset",
            "pcb",
            "--data-dir",
            os.path.expanduser(
                "~/dataset/PCB/Huang/PCB_DATASET/PCB-gray-128___deco-diff"
            ),
            "--model-size",
            "UNet_L",
            "--object-class",
            "all",
            "--anomaly-class",
            "all",
            "--image-size",
            "128",
            "--center-size",
            "128",
            "--center-crop",
            "False",
            "--batch-num",
            "1",
            "--pretrained",
            "DeCo-Diff_pcb_all_UNet_L_128_CenterCrop/001-UNet_L/checkpoints/best.pt",
            "--split",
            "train",
            "--perturbation",
            "noise",
        ]
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dataset", type=str, choices=["mvtec", "visa", "pcb"], default="mvtec"
    )
    parser.add_argument("--data-dir", type=str, default="./mvtec-dataset/")
    parser.add_argument(
        "--model-size",
        type=str,
        choices=["UNet_XS", "UNet_S", "UNet_M", "UNet_L", "UNet_XL"],
        default="UNet_L",
    )
    parser.add_argument("--image-size", type=int, default=288)
    parser.add_argument("--center-size", type=int, default=256)
    parser.add_argument("--batch-num", type=int, default=12)
    parser.add_argument(
        "--center-crop",
        type=lambda v: True if v.lower() in ("yes", "true", "t", "y", "1") else False,
        default=True,
    )
    parser.add_argument(
        "--vae-type", type=str, choices=["ema", "mse"], default="ema"
    )  # Choice doesn't affect training
    parser.add_argument("--num-workers", type=int, default=4)
    parser.add_argument("--object-class", type=str, default="all")
    parser.add_argument("--pretrained", type=str, default=".")
    parser.add_argument("--anomaly-class", type=str, default="all")
    parser.add_argument("--reverse-steps", type=int, default=5)
    parser.add_argument("--split", type=str, default="test")
    parser.add_argument(
        "--perturbation",
        type=str,
        choices=[None, "brightness", "shift_x", "shift_y", "noise", "blur", "scratch"],
        default=None,
    )
    parser.add_argument("--split-csv-path", type=str, default=None)
    parser.add_argument(
        "--input-json",
        type=str,
        help="Path to JSON file containing multiple test configurations"
    )

    args = parser.parse_args()
    
    # Handle input JSON if provided
    if args.input_json:
        import json
        with open(args.input_json, 'r') as f:
            test_configs = json.load(f)
            
        # Run evaluation for each test configuration
        for test_name, test_args in test_configs.items():
            print(f"\nRunning evaluation for {test_name}")
            print(test_args)
            # Update args with test configuration
            for key, value in test_args.items():
                # Convert key from kebab-case to snake_case
                key = key.replace('-', '_')
                if hasattr(args, key):
                    # Convert string values to appropriate types
                    if key in ['image_size', 'center_size', 'batch_num', 'reverse_steps']:
                        value = int(value)
                    elif key == 'center_crop':
                        value = value.lower() in ('yes', 'true', 't', 'y', '1')
                    elif key in ['pretrained', 'data_dir', 'split_csv_path']:
                        value = os.path.expanduser(value)
                    setattr(args, key, value)
            
            # Set up derived arguments
            if args.dataset == "mvtec":
                args.num_classes = 15
            elif args.dataset == "visa":
                args.num_classes = 12
            elif args.dataset == "pcb":
                args.num_classes = 1
            current_time = datetime.now().strftime("%y%m%d_%H%M%S")
            args.results_dir = f"results/{test_name}_{current_time}"
            os.makedirs(args.results_dir, exist_ok=True)
            # INSERT_YOUR_CODE
            # Save the current test_args (key-value pairs) into the results_dir as a JSON file
            config_save_path = os.path.join(args.results_dir, "config.json")
            with open(config_save_path, "w") as config_file:
                json.dump(test_args, config_file, indent=2)
            if args.center_crop:
                args.actual_image_size = args.center_size
            else:
                args.actual_image_size = args.image_size

            # Set up object classes
            if args.object_class == "all" and args.dataset == "mvtec":
                args.object_classes = [
                    "bottle", "cable", "capsule", "hazelnut", "metal_nut",
                    "pill", "screw", "toothbrush", "transistor", "zipper",
                    "carpet", "grid", "leather", "tile", "wood",
                ]
            elif args.object_class == "all" and args.dataset == "visa":
                args.object_classes = [
                    "candle", "cashew", "fryum", "macaroni2", "pcb2", "pcb4",
                    "capsules", "chewinggum", "macaroni1", "pcb1", "pcb3", "pipe_fryum",
                ]
            elif args.object_class == "all" and args.dataset == "pcb":
                args.object_classes = ["pcb"]
            else:
                args.object_classes = [args.object_class]
                
            # Run evaluation for this configuration
            evaluation(args)
    else:
        # Original single configuration evaluation
        if args.dataset == "mvtec":
            args.num_classes = 15
        elif args.dataset == "visa":
            args.num_classes = 12
        elif args.dataset == "pcb":
            args.num_classes = 1
        args.results_dir = f"./DeCo-Diff_{args.dataset}_{args.object_class}_{args.model_size}_{args.center_size}"
        if args.center_crop:
            args.results_dir += "_CenterCrop"
            args.actual_image_size = args.center_size
        else:
            args.actual_image_size = args.image_size

        if args.object_class == "all" and args.dataset == "mvtec":
            args.object_classes = [
                "bottle",
                "cable",
                "capsule",
                "hazelnut",
                "metal_nut",
                "pill",
                "screw",
                "toothbrush",
                "transistor",
                "zipper",
                "carpet",
                "grid",
                "leather",
                "tile",
                "wood",
            ]
        elif args.object_class == "all" and args.dataset == "visa":
            args.object_classes = [
                "candle",
                "cashew",
                "fryum",
                "macaroni2",
                "pcb2",
                "pcb4",
                "capsules",
                "chewinggum",
                "macaroni1",
                "pcb1",
                "pcb3",
                "pipe_fryum",
            ]
        elif args.object_class == "all" and args.dataset == "pcb":
            args.object_classes = [
                "pcb",
            ]
        else:
            args.object_classes = [args.object_class]

        evaluation(args)


# %%
if __name__ == "__main__":
    main()
# Below are cell makrkers used in VSCode
# %%
#
# %%
